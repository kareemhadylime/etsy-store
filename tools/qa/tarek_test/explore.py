"""Inspect the workbook structure to understand sheets, cells, layouts."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
import openpyxl
from openpyxl.utils import get_column_letter

PATH = r"C:\ETSY\etsy-store\tools\qa\tarek_test\tarek.xlsx"

wb = openpyxl.load_workbook(PATH, data_only=False)
print("=== Sheets ===")
for s in wb.sheetnames:
    ws = wb[s]
    print(f"  {s} -- dim {ws.dimensions} max_row={ws.max_row} max_col={ws.max_column}")

print()
for s in wb.sheetnames:
    ws = wb[s]
    print(f"\n===== {s} =====")
    rows_to_show = min(ws.max_row, 40)
    cols_to_show = min(ws.max_column, 14)
    for r in range(1, rows_to_show + 1):
        line_parts = []
        for c in range(1, cols_to_show + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is None:
                continue
            sv = str(v)
            if len(sv) > 38:
                sv = sv[:35] + "..."
            line_parts.append(f"{get_column_letter(c)}{r}={sv!r}")
        if line_parts:
            print("  " + " | ".join(line_parts))
