"""
Live multi-persona simulation suite for the Debt Payoff Planner bundle.

5 real-life debt scenarios, each:
  1. Hand-computed via Python month-by-month cascade simulation (ground truth)
  2. Written into the AI Edition workbook via openpyxl
  3. Recalculated via LibreOffice headless
  4. Evaluated cells read and compared to ground truth
  5. Reported as a verdict per metric

Output: tools/qa/output/live-personas-report.md
"""
import sys, os, math, subprocess, json
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from datetime import datetime

SOFFICE = r"C:\Program Files\LibreOffice\program\soffice.com"
WORKBOOK = "tools/sheets-gen/output/debt-payoff-planner-ai-edition.xlsx"
SCRATCH = "tools/qa/scratch/personas"
os.makedirs(SCRATCH, exist_ok=True)


# ============================================================
# THE 5 PERSONAS
# ============================================================

PERSONAS = [
    {
        "id": "P1",
        "name": "Recent grad — $42k student loans + 2 credit cards",
        "context": (
            "26yo software engineer, 2 years out of college. Federal student loans on IBR, "
            "moved to standard payoff because income went up. Two credit cards from college "
            "that never got fully paid off."
        ),
        "income": 6500,  # $78k gross, monthly
        "extra": 400,
        "debts": [
            ("Sallie Mae Federal", "Student Loan",  42000, 0.0489, 380, 15),
            ("Capital One Quicksilver", "Credit Card", 3800, 0.2299, 95,  3),
            ("Discover It",        "Credit Card",  2400, 0.1899,  60, 22),
        ],
        "expected_strategy_winner": "avalanche",
        "explain": "High-APR CC's should be killed first; student loan low APR can take its time",
    },
    {
        "id": "P2",
        "name": "Dual-income couple — mortgage + auto + 1 high-APR CC",
        "context": (
            "Combined household income $13.5k/mo. Mortgage on a starter home, one financed "
            "car, one stubborn CC they kept paying minimums on for 3 years."
        ),
        "income": 13500,
        "extra": 800,
        "debts": [
            ("Wells Fargo Mortgage", "Mortgage",   285000, 0.0625, 1850, 1),
            ("Toyota Financial",     "Car Loan",   18500, 0.0589,  420, 10),
            ("Chase Sapphire",       "Credit Card", 8700, 0.2299,  220,  5),
        ],
        "expected_strategy_winner": "avalanche",
        "explain": "Mortgage 30yr fixed — don't aggressively prepay; Avalanche flags the CC as priority",
    },
    {
        "id": "P3",
        "name": "Single parent — $14k medical + payday + 2 CCs",
        "context": (
            "Single parent, 1 kid. Surgery a year ago left $14k medical debt. Payday loan "
            "from a tight month. Two CCs near limit."
        ),
        "income": 4200,
        "extra": 100,  # very limited extra
        "debts": [
            ("ER Bill (Collections)",  "Medical Debt",  14000, 0.00,  100,  1),
            ("Speedy Cash",            "Personal Loan", 1850, 0.3990, 220, 28),  # payday-loan APR
            ("Old Navy Visa",          "Credit Card",   2100, 0.2599, 65,  12),
            ("Walmart Mastercard",     "Credit Card",   2800, 0.2799, 80,  17),
        ],
        "expected_strategy_winner": "avalanche",
        "explain": "Payday loan at 40% APR is the bleeding wound — kill first; medical at 0% can wait",
    },
    {
        "id": "P4",
        "name": "Pre-retiree — clearing everything by age 65",
        "context": (
            "Age 61, planning retirement in 4 years. Wants debt-free entry. Higher income, "
            "high extra payment availability."
        ),
        "income": 11000,
        "extra": 1500,
        "debts": [
            ("HELOC",                  "Personal Loan", 48000, 0.0825, 380, 1),
            ("BMW Financial",          "Car Loan",      27500, 0.0599, 540, 12),
            ("Chase Slate (BT)",       "Credit Card",   6200, 0.00,  186, 8),  # 0% balance transfer
            ("AMEX Platinum",          "Credit Card",   3400, 0.2499, 85,  20),
        ],
        "expected_strategy_winner": "avalanche",
        "explain": "AMEX at 24.99% first; HELOC 8% second; BMW at 6%; BT card last (0% APR until promo)",
    },
    {
        "id": "P5",
        "name": "Maxed-cards-and-BT — 3 CCs + 1 BT card + auto",
        "context": (
            "Just consolidated $9k onto a 0% balance transfer card. 3 other CCs still active. "
            "Promo expires in 18 months."
        ),
        "income": 7800,
        "extra": 500,
        "debts": [
            ("Citi BT Card (promo)",   "Credit Card",  9000, 0.00,  180,  5),
            ("Capital One Venture",    "Credit Card",  4200, 0.2399, 105, 12),
            ("US Bank Visa",           "Credit Card",  3500, 0.2199, 90,  18),
            ("Bank of America CC",     "Credit Card",  2100, 0.2599, 55,  25),
            ("Honda Financial",        "Car Loan",     11500, 0.0699, 280, 1),
        ],
        "expected_strategy_winner": "avalanche",
        "explain": "Three high-APR CCs first; auto loan mid; BT card last (free money until promo ends)",
    },
]


# ============================================================
# GROUND-TRUTH SIMULATOR (true cascade)
# ============================================================

def true_cascade_sim(debts, order, extra):
    """Month-by-month true cascade. Returns (months, total_interest).
    Debt tuple shape: (name, type, balance, apr, min, due) → idx 0=name, 1=type, 2=balance, 3=apr, 4=min, 5=due.
    """
    d = [list(x) for x in debts]
    ORIG_MINS = sum(x[4] for x in debts)  # x[4] = min, NOT x[3] which is APR
    TOTAL_PAY = ORIG_MINS + extra
    total_int = 0.0
    month = 0
    while any(x[2] > 0.01 for x in d):
        month += 1
        m_int = 0.0
        # Accrue interest on every active debt
        for x in d:
            if x[2] > 0:
                int_this = x[2] * x[3] / 12
                x[2] += int_this
                m_int += int_this
        total_int += m_int
        pool = TOTAL_PAY
        # Apply each active debt's own minimum
        for x in d:
            if x[2] > 0:
                pay = min(x[4], x[2])
                x[2] -= pay
                pool -= pay
        # Cascade remaining pool to debts in strategy order
        for idx in order:
            if pool > 0.01 and d[idx][2] > 0:
                pay = min(pool, d[idx][2])
                d[idx][2] -= pay
                pool -= pay
        if month > 600:
            break
    return month, total_int


def compute_strategy_orders(debts):
    """Return (snowball_indices, avalanche_indices) by debt index."""
    # debts: list of (name, type, balance, apr, min, due)
    idxs = list(range(len(debts)))
    snow = sorted(idxs, key=lambda i: debts[i][2])  # ascending balance
    ava = sorted(idxs, key=lambda i: -debts[i][3])  # descending APR
    return snow, ava


def hand_compute(persona):
    debts = persona["debts"]
    extra = persona["extra"]
    snow_order, ava_order = compute_strategy_orders(debts)

    snow_m, snow_i = true_cascade_sim(debts, snow_order, extra)
    ava_m, ava_i = true_cascade_sim(debts, ava_order, extra)

    total_bal = sum(d[2] for d in debts)
    total_min = sum(d[4] for d in debts)
    weighted_apr = sum(d[2] * d[3] for d in debts) / total_bal if total_bal else 0
    dti = total_min / persona["income"]

    return {
        "total_bal": total_bal,
        "total_min": total_min,
        "weighted_apr": weighted_apr,
        "dti": dti,
        "snow_months": snow_m,
        "snow_interest": snow_i,
        "snow_order_names": [debts[i][0] for i in snow_order],
        "ava_months": ava_m,
        "ava_interest": ava_i,
        "ava_order_names": [debts[i][0] for i in ava_order],
        "savings_av_vs_sn": max(0, snow_i - ava_i),
    }


# ============================================================
# WORKBOOK WRITE + RECALC + READ
# ============================================================

def write_persona_to_workbook(persona, out_path):
    wb = openpyxl.load_workbook(WORKBOOK)
    dl = wb['📋 Debt List']
    # Clear rows 11-30
    for r in range(11, 31):
        for c in range(2, 12):
            dl.cell(row=r, column=c).value = None
    # Write persona debts
    for i, (name, typ, bal, apr, mn, due) in enumerate(persona["debts"]):
        r = 11 + i
        dl.cell(row=r, column=2).value = name
        dl.cell(row=r, column=3).value = typ
        dl.cell(row=r, column=4).value = bal
        dl.cell(row=r, column=5).value = apr
        dl.cell(row=r, column=6).value = mn
        dl.cell(row=r, column=7).value = due
    # Set income at K40
    dl['K40'] = persona["income"]
    # Set extra at Extra Payment Simulator!D6
    wb['🎯 Extra Payment Simulator']['D6'] = persona["extra"]
    wb.save(out_path)


def recalc_via_lo(in_path, out_dir):
    # LO requires Windows-style absolute paths (forward slashes break load on this host).
    in_abs = os.path.abspath(in_path).replace("/", "\\")
    out_abs = os.path.abspath(out_dir).replace("/", "\\")
    out_path = os.path.join(out_dir, os.path.basename(in_path))
    out_path_abs = os.path.abspath(out_path).replace("/", "\\")
    for attempt in range(3):
        try:
            if os.path.exists(out_path_abs):
                os.remove(out_path_abs)
        except Exception:
            pass
        result = subprocess.run([SOFFICE, "--headless", "--calc", "--convert-to", "xlsx",
                                 "--outdir", out_abs, in_abs],
                                capture_output=True, timeout=120, text=True)
        if os.path.exists(out_path_abs):
            return out_path_abs
        import time
        time.sleep(2)
    raise RuntimeError(f"LibreOffice recalc failed after 3 attempts. stderr: {result.stderr if 'result' in dir() else 'n/a'}")


def read_evaluated(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sim = wb['_Strategy Sim']
    sc = wb['📊 Strategy Comparison']
    dl = wb['📋 Debt List']
    sn = wb['❄️ Snowball Method']
    av = wb['🌊 Avalanche Method']
    eps = wb['🎯 Extra Payment Simulator']
    ai = wb['🤖 AI Credit Coach']
    mt = wb['🏆 Milestone Tracker']

    # Snowball ranking — read names from C11..C20
    snow_names = [sn[f'C{r}'].value for r in range(11, 21) if sn[f'C{r}'].value]
    ava_names  = [av[f'C{r}'].value for r in range(11, 21) if av[f'C{r}'].value]

    return {
        "total_bal_d32": dl['D32'].value,
        "total_min_f32": dl['F32'].value,
        "snow_months": sim['C3'].value,
        "snow_interest": sim['D3'].value,
        "ava_months": sim['C4'].value,
        "ava_interest": sim['D4'].value,
        "snow_names": snow_names,
        "ava_names": ava_names,
        "sc_snow_months_disp": sc['C11'].value,
        "sc_ava_months_disp": sc['D11'].value,
        "sc_savings": sc['K2'].value,
        "eps_months_left": eps['I2'].value,
        "ai_health_score": ai['B10'].value,
        "milestone_b9_div0_check": mt['B9'].value,
        "milestone_e18_div0_check": mt['E18'].value,
    }


# ============================================================
# REPORT GENERATION
# ============================================================

def fmt_money(x):
    if x is None: return "—"
    return f"${x:,.0f}"

def fmt_pct(x):
    if x is None: return "—"
    return f"{x*100:.1f}%"

def fmt_months(x):
    if x is None: return "—"
    return f"{x} mo"

def verdict_symbol(workbook_val, ground_truth, tolerance_pct=5):
    if workbook_val is None or ground_truth is None or ground_truth == 0:
        return "❓"
    delta = abs(workbook_val - ground_truth)
    pct_off = delta / abs(ground_truth) * 100 if ground_truth else 0
    if pct_off <= tolerance_pct:
        return "✅"
    elif pct_off <= 15:
        return "⚠️"
    return "❌"


def run_all():
    results = []
    for persona in PERSONAS:
        print(f"\n=== {persona['id']}: {persona['name']} ===")
        truth = hand_compute(persona)
        print(f"  Ground truth: SB={truth['snow_months']}mo/${truth['snow_interest']:,.0f}, "
              f"AV={truth['ava_months']}mo/${truth['ava_interest']:,.0f}")

        in_path = os.path.join(SCRATCH, f"persona-{persona['id']}.xlsx")
        write_persona_to_workbook(persona, in_path)

        out_dir = os.path.join(SCRATCH, "recalc")
        os.makedirs(out_dir, exist_ok=True)
        recalc_path = recalc_via_lo(in_path, out_dir)

        wb_vals = read_evaluated(recalc_path)
        print(f"  Workbook:    SB={wb_vals['snow_months']}mo/${wb_vals['snow_interest']:,.0f}, "
              f"AV={wb_vals['ava_months']}mo/${wb_vals['ava_interest']:,.0f}")
        b9_str = str(wb_vals['milestone_b9_div0_check'])[:50]
        print(f"  Milestone day-one: B9={b9_str!r} E18={wb_vals['milestone_e18_div0_check']!r}")

        results.append({"persona": persona, "truth": truth, "workbook": wb_vals})
    return results


# ============================================================
# WRITE MARKDOWN REPORT
# ============================================================

def write_report(results):
    lines = []
    lines.append("# Live Multi-Persona Simulation Report")
    lines.append(f"\n**Generated:** {datetime.now().isoformat(timespec='seconds')}")
    lines.append(f"**Workbook:** `{WORKBOOK}`")
    lines.append(f"**Recalc engine:** LibreOffice headless (real ground-truth path)")
    lines.append(f"**Personas tested:** {len(results)}\n")

    lines.append("## Summary verdict\n")
    lines.append("| Persona | Description | SB months (WB/truth) | SB interest (WB/truth) | AV months (WB/truth) | AV interest (WB/truth) | DIV/0? |")
    lines.append("|---|---|---:|---:|---:|---:|:---:|")
    for r in results:
        p, t, w = r["persona"], r["truth"], r["workbook"]
        div0 = "✅ clean" if (w["milestone_b9_div0_check"] and "#DIV" not in str(w["milestone_b9_div0_check"])) else "❌ #DIV/0!"
        lines.append(
            f"| **{p['id']}** | {p['name'][:50]} | {w['snow_months']}/{t['snow_months']} {verdict_symbol(w['snow_months'], t['snow_months'])} "
            f"| {fmt_money(w['snow_interest'])}/{fmt_money(t['snow_interest'])} {verdict_symbol(w['snow_interest'], t['snow_interest'])} "
            f"| {w['ava_months']}/{t['ava_months']} {verdict_symbol(w['ava_months'], t['ava_months'])} "
            f"| {fmt_money(w['ava_interest'])}/{fmt_money(t['ava_interest'])} {verdict_symbol(w['ava_interest'], t['ava_interest'])} "
            f"| {div0} |"
        )

    lines.append("\n---\n")

    for r in results:
        p, t, w = r["persona"], r["truth"], r["workbook"]
        lines.append(f"## {p['id']}: {p['name']}\n")
        lines.append(f"**Context:** {p['context']}\n")
        lines.append(f"**Strategy guidance:** {p['explain']}\n")

        lines.append("### Inputs\n")
        lines.append("| Debt | Type | Balance | APR | Min |")
        lines.append("|---|---|---:|---:|---:|")
        for n, t_, b, a, m, _ in p["debts"]:
            lines.append(f"| {n} | {t_} | {fmt_money(b)} | {fmt_pct(a)} | {fmt_money(m)} |")
        lines.append(f"\nMonthly income: **{fmt_money(p['income'])}** · Extra: **{fmt_money(p['extra'])}/mo**")

        lines.append("\n### Results vs ground truth\n")
        lines.append("| Metric | Hand-computed | Workbook (LO recalc) | Δ | Verdict |")
        lines.append("|---|---:|---:|---:|:---:|")

        def row(label, true_v, wb_v, fmt=fmt_money, tol=5):
            d = (wb_v - true_v) if (wb_v is not None and true_v is not None) else None
            d_str = fmt(d) if d is not None else "—"
            return f"| {label} | {fmt(true_v)} | {fmt(wb_v)} | {d_str} | {verdict_symbol(wb_v, true_v, tol)} |"

        lines.append(row("Total balance", t["total_bal"], w["total_bal_d32"], fmt_money, tol=0.5))
        lines.append(row("Total min", t["total_min"], w["total_min_f32"], fmt_money, tol=0.5))
        lines.append(row("Snowball months", t["snow_months"], w["snow_months"], fmt_months, tol=10))
        lines.append(row("Snowball interest", t["snow_interest"], w["snow_interest"], fmt_money, tol=10))
        lines.append(row("Avalanche months", t["ava_months"], w["ava_months"], fmt_months, tol=10))
        lines.append(row("Avalanche interest", t["ava_interest"], w["ava_interest"], fmt_money, tol=10))

        lines.append("\n### Snowball ordering\n")
        lines.append(f"- **Hand-computed:** {' → '.join(t['snow_order_names'])}")
        lines.append(f"- **Workbook:**       {' → '.join(w['snow_names']) if w['snow_names'] else '—'}")
        match = (t['snow_order_names'] == w['snow_names'][:len(t['snow_order_names'])])
        lines.append(f"- **Match:** {'✅ exact' if match else '❌ diverges'}")

        lines.append("\n### Avalanche ordering\n")
        lines.append(f"- **Hand-computed:** {' → '.join(t['ava_order_names'])}")
        lines.append(f"- **Workbook:**       {' → '.join(w['ava_names']) if w['ava_names'] else '—'}")
        match = (t['ava_order_names'] == w['ava_names'][:len(t['ava_order_names'])])
        lines.append(f"- **Match:** {'✅ exact' if match else '❌ diverges'}")

        lines.append("\n### Direction check\n")
        wb_direction = "Avalanche" if (w['ava_interest'] is not None and w['snow_interest'] is not None and w['ava_interest'] < w['snow_interest']) else "Snowball"
        truth_direction = "Avalanche" if t['ava_interest'] < t['snow_interest'] else "Snowball"
        dir_match = wb_direction == truth_direction
        lines.append(f"- **Truth says winner:** {truth_direction} (saves {fmt_money(t['savings_av_vs_sn'])})")
        lines.append(f"- **Workbook says winner:** {wb_direction}")
        lines.append(f"- **Match:** {'✅ correct direction' if dir_match else '❌ DIRECTION FLIPPED'}\n")

        lines.append("### Day-one health checks\n")
        b9 = str(w['milestone_b9_div0_check'])
        e18 = str(w['milestone_e18_div0_check'])
        lines.append(f"- Milestone Tracker B9 (progress bar): `{b9[:80]}` — {'✅ clean' if '#DIV' not in b9 else '❌ #DIV/0!'}")
        lines.append(f"- Milestone Tracker E18 (25% status): `{e18}` — {'✅ clean' if '#DIV' not in str(e18) else '❌ #DIV/0!'}")
        lines.append(f"- EPS MONTHS LEFT KPI: `{w['eps_months_left']}`")
        lines.append(f"- AI Health Score: `{w['ai_health_score']}`")
        lines.append("\n---\n")

    # Write
    out_path = "tools/qa/output/live-personas-report.md"
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"\n✓ Report written: {out_path}")
    return out_path


if __name__ == "__main__":
    results = run_all()
    write_report(results)
