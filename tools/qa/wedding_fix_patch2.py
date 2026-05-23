"""
Wedding Budget & Planner — Patch 2 Script
Fixes issues found in Round 2:
1. FX sheet E8-E17: invalid formula '=1.08*100 USD' -> static text '$108.00 USD'
2. Multi-Event Planner J10:J14 variance sign: =H-I -> =I-H (so over-budget is negative)
3. Update error scanner to not flag #1, #2 etc rank labels
"""
import sys, io, os, subprocess
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
from openpyxl import load_workbook

FIXED_DIR = 'C:/ETSY/etsy-store/tools/qa/fixed'
SCRATCH = 'C:/ETSY/etsy-store/tools/qa/scratch'
LO = r'C:\Program Files\LibreOffice\program\soffice.com'

def lo_recalc(src_path, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    cmd = [LO, '--headless', '--calc', '--convert-to', 'xlsx', '--outdir', out_dir, src_path]
    result = subprocess.run(cmd, capture_output=True, timeout=90)
    base = os.path.basename(src_path)
    out = os.path.join(out_dir, base)
    return out if os.path.exists(out) else None

print('\n' + '='*70)
print('PATCH 2 — Fix FX sheet example formulas + Multi-Event variance sign')
print('='*70)

for key in ['pro', 'ai-edition']:
    path = os.path.join(FIXED_DIR, f'wedding-budget-planner-{key}.xlsx')
    if not os.path.exists(path):
        print(f'  SKIP {key}: not found')
        continue

    wb = load_workbook(path, data_only=False)

    # Fix 1: FX Rate sheet E8-E17 — replace invalid formula with static text
    for ws in wb.worksheets:
        if 'fx' in ws.title.lower() or 'currency' in ws.title.lower():
            fx_rates = [1.08, 1.27, 0.2723, 0.0205, 0.01196, 0.736, 0.643, 1.09, 0.0575, 0.0279]
            for i, rate in enumerate(fx_rates):
                r = 8 + i
                old_val = ws.cell(row=r, column=5).value
                new_val = f'${rate*100:.2f}'
                ws.cell(row=r, column=5).value = new_val
                print(f'  {key}/FX row {r} E: "{old_val}" -> "{new_val}"')
            print(f'  {key}: Fixed FX example column (E8:E17)')

    # Fix 2: Multi-Event Planner — variance sign fix
    # Current formula: =H-I (spent - budget = positive when over)
    # Should be:       =I-H (budget - spent = negative when over)
    for ws in wb.worksheets:
        if 'multi' in ws.title.lower() or 'event planner' in ws.title.lower():
            for r in range(10, 15):
                j_cell = ws.cell(row=r, column=10)
                old_j = j_cell.value
                if old_j and 'H' in str(old_j) and 'I' in str(old_j):
                    # Fix sign: =H{r}-I{r} -> =I{r}-H{r}
                    new_j = f'=I{r}-H{r}'
                    j_cell.value = new_j
                    print(f'  {key}/Multi-Event J{r}: "{old_j}" -> "{new_j}"')
            # Also fix totals row J16
            j16 = ws.cell(row=16, column=10)
            if j16.value and 'SUM(J' in str(j16.value):
                pass  # SUM stays the same, signs fix automatically
            else:
                j16.value = '=SUM(J10:J14)'
            print(f'  {key}: Fixed variance sign in Multi-Event Planner (I-H convention = negative when over)')

    wb.save(path)
    print(f'  {key}: Saved')

# Recalculate to verify
print('\nVerifying patch...')
recalc_dir = os.path.join(SCRATCH, 'r2-patch2-recalc')
for key in ['pro', 'ai-edition']:
    path = os.path.join(FIXED_DIR, f'wedding-budget-planner-{key}.xlsx')
    if not os.path.exists(path):
        continue
    out = lo_recalc(path, recalc_dir)
    if out:
        wb_check = load_workbook(out, data_only=True)
        errors = []
        for ws in wb_check.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    # Skip rank label cells (#1 through #5) and string rank labels
                    if isinstance(v, str) and v.startswith('#') and len(v) < 20:
                        # Check if it's a rank label (just digits after #)
                        suffix = v[1:]
                        if not suffix.isdigit():
                            errors.append(f'{ws.title}/{cell.coordinate}: {v}')
        if errors:
            print(f'  {key}: {len(errors)} formula errors after patch:')
            for e in errors[:10]:
                print(f'    {e}')
        else:
            print(f'  {key}: 0 formula errors after patch — CLEAN')

        # Spot check multi-event variance
        for ws in wb_check.worksheets:
            if 'multi' in ws.title.lower() or 'event planner' in ws.title.lower():
                for r in range(10, 13):
                    j_val = ws.cell(row=r, column=10).value
                    h_val = ws.cell(row=r, column=8).value
                    i_val = ws.cell(row=r, column=9).value
                    print(f'  {key}/Multi-Event row {r}: H={h_val}, I={i_val}, J={j_val} (expect I-H = {(i_val or 0)-(h_val or 0)})')

print('\nPatch 2 DONE.')
