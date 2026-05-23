"""
Wedding Budget & Planner — Fix + Complement Script
Applies all Round-1 CRITICAL and HIGH fixes to copies in tools/qa/fixed/
Python 3.14, openpyxl 3.1.5
"""
import sys, io, os, shutil, copy, datetime, subprocess, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = {
    'essentials': 'C:/ETSY/etsy-store/tools/sheets-gen/output/wedding-budget-planner-essentials.xlsx',
    'pro':        'C:/ETSY/etsy-store/tools/sheets-gen/output/wedding-budget-planner-pro.xlsx',
    'ai':         'C:/ETSY/etsy-store/tools/sheets-gen/output/wedding-budget-planner-ai-edition.xlsx',
}
FIXED_DIR = 'C:/ETSY/etsy-store/tools/qa/fixed'
SCRATCH    = 'C:/ETSY/etsy-store/tools/qa/scratch'
LO         = r'C:\Program Files\LibreOffice\program\soffice.com'

os.makedirs(FIXED_DIR, exist_ok=True)
os.makedirs(SCRATCH, exist_ok=True)

changelog = []
fix_count = 0

def log(issue_id, file_key, sheet, cell_range, before, after, rationale):
    global fix_count
    fix_count += 1
    entry = f'[FIX:{issue_id}] {file_key} | {sheet} | {cell_range}\n  BEFORE: {before[:100]}\n  AFTER:  {after[:100]}\n  WHY:    {rationale}\n'
    changelog.append(entry)
    print(f'  Fixed {issue_id}: {file_key}/{sheet}/{cell_range}')

def log_complement(tag, file_key, sheet, cell_range, description):
    entry = f'[COMPLEMENT:{tag}] {file_key} | {sheet} | {cell_range}\n  DESCRIPTION: {description}\n'
    changelog.append(entry)
    print(f'  Complement {tag}: {file_key}/{sheet}/{cell_range} — {description}')


# ══════════════════════════════════════════════════════════════════════════════
# FIX 1 — TEXT() format string double-quote issue (#VALUE! errors)
# Pattern: ""$"#,##0"  →  "$#,##0"
# This affects all three workbooks, all KPI banner cells in row 2 of each tab.
# Root cause: ExcelJS/generator used Python-escaped quotes that are valid in
# Excel's internal XML but fail in LibreOffice's formula parser.
# ══════════════════════════════════════════════════════════════════════════════
print('\n' + '='*70)
print('FIX 1 — TEXT() format string fix (WBP-001 through WBP-060)')
print('='*70)

def fix_text_format(formula_str):
    """
    Replace all occurrences of the double-escaped dollar format in TEXT().
    The generator encoded dollar signs as ""$"" which produces #VALUE! in LibreOffice.
    In number format strings, $ does not need quoting in either Excel or LibreOffice.

    Two patterns (using explicit character sequences):
      Pattern B (with negative variant): ""$"#,##0;\\-"$"#,##0"  -> "$#,##0;-$#,##0"
      Pattern A (simple positive-only):  ""$"#,##0"               -> "$#,##0"
    """
    if formula_str is None:
        return formula_str
    s = str(formula_str)
    if '""' not in s:
        return formula_str
    # Build the exact target strings using chr() to avoid escape ambiguity
    q = chr(34)   # "
    # Pattern B: "" $ " # , # # 0 ; \ - " $ " # , # # 0 "
    pat_b = q + q + '$' + q + '#,##0;' + chr(92) + '-' + q + '$' + q + '#,##0' + q
    rep_b = q + '$#,##0;-$#,##0' + q
    # Pattern A: "" $ " # , # # 0 "
    pat_a = q + q + '$' + q + '#,##0' + q
    rep_a = q + '$#,##0' + q
    s = s.replace(pat_b, rep_b)
    s = s.replace(pat_a, rep_a)
    return s

def apply_text_fixes(wb, file_key):
    cell_count = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and '""' in v and 'TEXT(' in v:
                    new_v = fix_text_format(v)
                    if new_v != v:
                        cell.value = new_v
                        cell_count += 1
    print(f'  {file_key}: fixed {cell_count} TEXT() format cells')
    return cell_count

for key, src_path in SRC.items():
    out_path = os.path.join(FIXED_DIR, f'wedding-budget-planner-{key if key != "ai" else "ai-edition"}.xlsx')
    if key == 'essentials':
        out_path = os.path.join(FIXED_DIR, 'wedding-budget-planner-essentials.xlsx')
    elif key == 'pro':
        out_path = os.path.join(FIXED_DIR, 'wedding-budget-planner-pro.xlsx')
    else:
        out_path = os.path.join(FIXED_DIR, 'wedding-budget-planner-ai-edition.xlsx')

    shutil.copy(src_path, out_path)
    wb = load_workbook(out_path, data_only=False)
    count = apply_text_fixes(wb, key)
    wb.save(out_path)

    if count > 0:
        log(f'WBP-001-060 ({key})', key, 'All sheets', 'Row 2 KPI cells',
            '=""BUDGET CAP""&CHAR(10)&(TEXT(E16,\"\"\"$\"\"#,##0\"))',
            '=""BUDGET CAP""&CHAR(10)&(TEXT(E16,\"$#,##0\"))',
            'TEXT() format strings used double-escaped dollar sign ""$"" which LibreOffice cannot parse. Fix: use single-quoted "$#,##0"')


# ══════════════════════════════════════════════════════════════════════════════
# FIX 2 — Vendor Tracker: add headers to row 8 (currently blank)
# and add Cake as a standalone Budget Category line item
# ══════════════════════════════════════════════════════════════════════════════
print('\n' + '='*70)
print('FIX 2 — Vendor Tracker: expose column headers (WBP-065/066/067)')
print('='*70)

VENDOR_HEADERS = ['Vendor Name', 'Category', 'Contact', 'Deposit Paid ($)', 'Total ($)',
                  'Balance Due ($)', 'Balance Due Date', 'Status', 'Contract', 'Notes', 'Days to Due']

for key in ['essentials', 'pro', 'ai-edition']:
    out_path = os.path.join(FIXED_DIR, f'wedding-budget-planner-{key}.xlsx')
    if not os.path.exists(out_path):
        continue
    wb = load_workbook(out_path, data_only=False)
    for ws in wb.worksheets:
        if 'vendor' in ws.title.lower() and 'comparison' not in ws.title.lower():
            # Row 8 is merged B8:L8 — unmerge it first, then add column headers
            # Remove merge for row 8
            merges_to_remove = []
            for merge in list(ws.merged_cells.ranges):
                if merge.min_row == 8 and merge.max_row == 8:
                    merges_to_remove.append(str(merge))
            for merge_range in merges_to_remove:
                ws.unmerge_cells(merge_range)
                print(f'  {key}: Unmerged {merge_range} in {ws.title}')

            cols = [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]  # B through L
            headers = ['Vendor Name', 'Category', 'Contact', 'Deposit $', 'Total $',
                       'Balance $', 'Balance Due Date', 'Status', 'Contract URL', 'Notes', 'Days to Due']
            header_fill = PatternFill(start_color='FF8B5A6B', end_color='FF8B5A6B', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFFFF')
            for col_idx, header in zip(cols, headers):
                cell = ws.cell(row=8, column=col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            print(f'  {key}/{ws.title}: added column headers in row 8')
            log_complement('VT-HEADERS', key, ws.title, 'B8:L8',
                           'Added visible column headers to Vendor Tracker header row (unmerged row 8 first)')
    wb.save(out_path)


# ══════════════════════════════════════════════════════════════════════════════
# COMPLEMENT 1 — Add "Cake" as standalone category
# Currently "Catering + Bar" does not break out Cake separately.
# We add it as the last category row.
# ══════════════════════════════════════════════════════════════════════════════
# Note: The existing 14 categories (rows 10-23) already sum to the budget.
# We add Cake as an informational note — the percentage column is editable so
# the user can re-split Catering and Cake allocations. We simply add a comment.
# Actually, looking at the structure, row 23 is Contingency. We can't add rows
# without breaking the fixed-range formulas. Instead we document this limitation.
print('\n' + '='*70)
print('COMPLEMENT 1 — Cake category note (WBP-071/072)')
print('='*70)

for key in ['essentials', 'pro', 'ai-edition']:
    out_path = os.path.join(FIXED_DIR, f'wedding-budget-planner-{key}.xlsx')
    if not os.path.exists(out_path):
        continue
    wb = load_workbook(out_path, data_only=False)
    for ws in wb.worksheets:
        if 'budget cat' in ws.title.lower():
            # Update "Catering + Bar" cell note to mention Cake
            b11 = ws['B11']
            old_val = str(b11.value or '')
            if 'catering' in old_val.lower():
                b11.value = 'Catering + Bar + Cake'
                log('WBP-071', key, ws.title, 'B11',
                    'Catering + Bar',
                    'Catering + Bar + Cake',
                    'Cake is canonically a separate wedding spend category; merged into Catering row since fixed row structure cannot be extended without breaking formulas. Label updated to surface cake.')
    wb.save(out_path)


# ══════════════════════════════════════════════════════════════════════════════
# COMPLEMENT 2 — Multi-Event Sub-Budget Tab (Persona 2 / WBP-062/063)
# Add a "Multi-Event Planner" sheet to Pro and AI editions
# ══════════════════════════════════════════════════════════════════════════════
print('\n' + '='*70)
print('COMPLEMENT 2 — Multi-Event Sub-Budget Tab (WBP-062/063)')
print('='*70)

MULTI_EVENT_CONTENT = [
    # (row, col, value, bold, fill_hex)
    (1,  1, '    Lime Premium Studios', False, None),
    (1,  4, 'Wedding Budget & Planner', False, None),
    (1, 10, 'Multi-Event Planner', True, None),
    (2,  1, '="EVENTS"&CHAR(10)&COUNTA(B10:B14)', False, None),
    (2,  3, '="TOTAL COMMITTED"&CHAR(10)&TEXT(SUMPRODUCT((E10:E14+F10:F14+G10:G14)),"$#,##0")', False, None),
    (2,  5, '="OVERALL BUDGET"&CHAR(10)&TEXT(\'🧭 Setup Wizard\'!E16,"$#,##0")', False, None),
    (2,  7, '="REMAINING"&CHAR(10)&TEXT(MAX(0,\'🧭 Setup Wizard\'!E16-SUMPRODUCT((E10:E14+F10:F14+G10:G14))),"$#,##0")', False, None),
    (3,  1, 'Track up to 5 pre-wedding and wedding events with separate guest counts, sub-budgets, and variance tracking.', False, None),
    (6,  2, 'Event', True, 'FF8B5A6B'),
    (6,  3, 'Date', True, 'FF8B5A6B'),
    (6,  4, 'Guest Count', True, 'FF8B5A6B'),
    (6,  5, 'Venue ($)', True, 'FF8B5A6B'),
    (6,  6, 'Food ($)', True, 'FF8B5A6B'),
    (6,  7, 'Other ($)', True, 'FF8B5A6B'),
    (6,  8, 'Event Total ($)', True, 'FF8B5A6B'),
    (6,  9, 'Sub-Budget ($)', True, 'FF8B5A6B'),
    (6, 10, 'Variance ($)', True, 'FF8B5A6B'),
    (6, 11, 'Per-Guest ($)', True, 'FF8B5A6B'),
    (6, 12, 'Status', True, 'FF8B5A6B'),
    # Event rows with formulas
    (10, 2, 'Katb-Kitab / Nikah'),
    (11, 2, 'Henna / Mehndi Night'),
    (12, 2, 'Rehearsal Dinner'),
    (13, 2, 'Ceremony'),
    (14, 2, 'Reception'),
]

EVENT_ROW_FORMULAS = {
    'H': '=E{r}+F{r}+G{r}',
    'J': '=H{r}-I{r}',
    'K': '=IFERROR(H{r}/MAX(1,D{r}),0)',
    'L': '=IF(J{r}<0,"Over Budget",IF(J{r}/MAX(1,I{r})<0.05,"Near Limit","Under Budget"))',
}

def add_multi_event_sheet(wb, edition_name):
    # Check if sheet already exists
    for ws in wb.worksheets:
        if 'event' in ws.title.lower() or 'multi' in ws.title.lower():
            print(f'  Multi-event sheet already exists: {ws.title}')
            return

    ws = wb.create_sheet('Multi-Event Planner', 2)
    ws.sheet_properties.tabColor = openpyxl.styles.colors.Color(rgb='FFD4A574')

    header_fill = PatternFill(start_color='FF8B5A6B', end_color='FF8B5A6B', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF', size=11)
    kpi_fill = PatternFill(start_color='FFC9A0A0', end_color='FFC9A0A0', fill_type='solid')
    kpi_font = Font(bold=True, size=11)

    # Header row 1
    ws['A1'] = '    Lime Premium Studios'
    ws['D1'] = f'Wedding Budget & Planner — {edition_name}'
    ws['J1'] = 'Multi-Event Planner'
    ws['J1'].font = Font(bold=True, size=12)

    # KPI row 2
    kpi_cells = {
        'A2': '="EVENTS"&CHAR(10)&COUNTA(B10:B14)',
        'C2': '="TOTAL COMMITTED"&CHAR(10)&TEXT(SUM(H10:H14),"$#,##0")',
        'E2': '="OVERALL BUDGET"&CHAR(10)&TEXT(\'🧭 Setup Wizard\'!E16,"$#,##0")',
        'G2': '="REMAINING"&CHAR(10)&TEXT(MAX(0,\'🧭 Setup Wizard\'!E16-SUM(H10:H14)),"$#,##0")',
        'I2': '="TOTAL GUESTS"&CHAR(10)&SUM(D10:D14)&" guests"',
        'K2': '="OVER BUDGET"&CHAR(10)&COUNTIF(J10:J14,"<0")&" event(s)"',
    }
    for addr, formula in kpi_cells.items():
        c = ws[addr]
        c.value = formula
        c.fill = kpi_fill
        c.font = kpi_font
        c.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws.row_dimensions[2].height = 45

    # Description row 3
    ws['B3'] = 'Track up to 5 pre-wedding and wedding events. Each event has its own guest count, venue/food/other spend, sub-budget, and variance. Reception over-budget will flag even if overall budget is fine.'

    # Column headers row 6
    headers = [
        ('B6', 'Event Name'), ('C6', 'Event Date'), ('D6', 'Guests'),
        ('E6', 'Venue ($)'), ('F6', 'Food + Bar ($)'), ('G6', 'Other ($)'),
        ('H6', 'Total Spent ($)'), ('I6', 'Sub-Budget ($)'), ('J6', 'Variance ($)'),
        ('K6', 'Per-Guest ($)'), ('L6', 'Status'),
    ]
    for addr, label in headers:
        c = ws[addr]
        c.value = label
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.row_dimensions[6].height = 30

    # Pre-filled event rows 10-14
    event_names = [
        'Katb-Kitab / Contract Signing',
        'Henna / Mehndi Night',
        'Rehearsal Dinner',
        'Ceremony',
        'Reception',
    ]
    for i, name in enumerate(event_names):
        r = 10 + i
        ws.cell(row=r, column=2).value = name  # B
        ws.cell(row=r, column=8).value = f'=E{r}+F{r}+G{r}'  # H = total spent
        ws.cell(row=r, column=10).value = f'=H{r}-I{r}'  # J = variance
        ws.cell(row=r, column=11).value = f'=IFERROR(H{r}/MAX(1,D{r}),0)'  # K = per guest
        ws.cell(row=r, column=12).value = (
            f'=IF(J{r}<0,"Over Budget",IF(AND(I{r}>0,J{r}/MAX(1,I{r})<0.05),"Near Limit","On Track"))'
        )  # L = status

        # Zero out input columns (so formulas start clean)
        for col in [3, 4, 5, 6, 7, 9]:  # C D E F G I
            ws.cell(row=r, column=col).value = 0 if col != 3 else None

    # Totals row 16
    ws['B16'] = 'TOTALS'
    ws['B16'].font = Font(bold=True)
    ws['D16'] = '=SUM(D10:D14)'  # total guests
    ws['H16'] = '=SUM(H10:H14)'  # total spent
    ws['I16'] = '=SUM(I10:I14)'  # total sub-budget
    ws['J16'] = '=SUM(J10:J14)'  # total variance

    # Cultural event guide row 19
    ws['B19'] = 'Cultural Event Guide'
    ws['B19'].font = Font(bold=True, size=11)
    ws['B20'] = 'Katb-Kitab (Arabic: contract signing ceremony) — typically intimate, 30-80 guests, officiant (Ma\'dhoun), light catering.'
    ws['B21'] = 'Henna / Mehndi Night — typically women-focused or mixed, pre-wedding celebration with henna artist, music, catering.'
    ws['B22'] = 'Sangeet (South Asian) — musical evening with dance performances, typically 1-2 nights before the wedding.'
    ws['B23'] = 'Rehearsal Dinner — Western convention, evening before wedding for wedding party + close family.'

    # Column widths
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 16
    ws.column_dimensions['J'].width = 14
    ws.column_dimensions['K'].width = 14
    ws.column_dimensions['L'].width = 18

    # Freeze panes
    ws.freeze_panes = 'A7'

    print(f'  Added Multi-Event Planner sheet to {edition_name}')

for key, edition in [('pro', 'Pro'), ('ai-edition', 'AI Edition')]:
    out_path = os.path.join(FIXED_DIR, f'wedding-budget-planner-{key}.xlsx')
    if not os.path.exists(out_path):
        continue
    wb = load_workbook(out_path, data_only=False)
    add_multi_event_sheet(wb, edition)
    wb.save(out_path)
    log_complement('MULTI-EVENT', key, 'Multi-Event Planner', 'A1:L25',
                   f'Added Multi-Event sub-budget tab to {edition} with 5 event rows, per-event variance, per-guest cost, cultural event guide (Katb-Kitab, Henna, Sangeet, Rehearsal Dinner)')


# ══════════════════════════════════════════════════════════════════════════════
# COMPLEMENT 3 — Contribution Tracker Tab (Persona 4 / WBP-068)
# Add "Family Contributions" sheet to Pro and AI editions
# ══════════════════════════════════════════════════════════════════════════════
print('\n' + '='*70)
print('COMPLEMENT 3 — Contribution Tracker Tab (WBP-068)')
print('='*70)

def add_contribution_sheet(wb, edition_name):
    for ws in wb.worksheets:
        if 'contrib' in ws.title.lower() or 'contribution' in ws.title.lower():
            print(f'  Contribution sheet already exists: {ws.title}')
            return

    ws = wb.create_sheet('Contribution Tracker', 3)
    ws.sheet_properties.tabColor = openpyxl.styles.colors.Color(rgb='FFD4A574')

    header_fill = PatternFill(start_color='FF8B5A6B', end_color='FF8B5A6B', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF', size=11)
    kpi_fill = PatternFill(start_color='FFC9A0A0', end_color='FFC9A0A0', fill_type='solid')

    ws['A1'] = '    Lime Premium Studios'
    ws['D1'] = f'Wedding Budget & Planner — {edition_name}'
    ws['J1'] = 'Contribution Tracker'
    ws['J1'].font = Font(bold=True, size=12)

    # KPI row 2
    kpi_cells = {
        'A2': '="TOTAL PLEDGED"&CHAR(10)&TEXT(SUM(D10:D14),"$#,##0")',
        'C2': '="TOTAL RECEIVED"&CHAR(10)&TEXT(SUM(E10:E14),"$#,##0")',
        'E2': '="TOTAL BUDGET"&CHAR(10)&TEXT(\'🧭 Setup Wizard\'!E16,"$#,##0")',
        'G2': '="COVERED"&CHAR(10)&TEXT(IFERROR(SUM(E10:E14)/MAX(1,\'🧭 Setup Wizard\'!E16),0),"0.0%")',
        'I2': '="UNCOVERED GAP"&CHAR(10)&TEXT(MAX(0,\'🧭 Setup Wizard\'!E16-SUM(E10:E14)),"$#,##0")',
        'K2': '="SOURCES"&CHAR(10)&COUNTA(B10:B14)',
    }
    for addr, formula in kpi_cells.items():
        c = ws[addr]
        c.value = formula
        c.fill = kpi_fill
        c.font = Font(bold=True, size=11)
        c.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws.row_dimensions[2].height = 45

    ws['B3'] = 'Track who is contributing to the wedding budget. Pledged = amount promised. Received = cash in hand. Gap = budget minus total received.'

    # Headers row 6
    headers = [
        ('B6', 'Contributor'), ('C6', 'Relationship'), ('D6', 'Pledged ($)'),
        ('E6', 'Received ($)'), ('F6', 'Outstanding ($)'), ('G6', '% of Budget'),
        ('H6', 'Date Expected'), ('I6', 'Notes'),
    ]
    for addr, label in headers:
        c = ws[addr]
        c.value = label
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.row_dimensions[6].height = 30

    # Pre-filled contributor rows 10-14
    contributors = [
        ("Bride's Family / Parents", "Bride's side"),
        ("Groom's Family / Parents", "Groom's side"),
        ("Couple", "Self"),
        ("Cash Gifts (received)", "Guests"),
        ("Other / Sponsors", "Other"),
    ]
    for i, (name, rel) in enumerate(contributors):
        r = 10 + i
        ws.cell(row=r, column=2).value = name   # B
        ws.cell(row=r, column=3).value = rel    # C
        ws.cell(row=r, column=4).value = 0      # D pledged
        ws.cell(row=r, column=5).value = 0      # E received
        ws.cell(row=r, column=6).value = f'=D{r}-E{r}'  # F outstanding
        ws.cell(row=r, column=7).value = f'=IFERROR(E{r}/MAX(1,\'🧭 Setup Wizard\'!E16),0)'  # G % of budget

    # Totals row 16
    ws['B16'] = 'TOTALS'
    ws['B16'].font = Font(bold=True)
    ws['D16'] = '=SUM(D10:D14)'
    ws['E16'] = '=SUM(E10:E14)'
    ws['F16'] = '=SUM(F10:F14)'
    ws['G16'] = f'=IFERROR(E16/MAX(1,\'🧭 Setup Wizard\'!E16),0)'

    # Budget comparison section row 19
    ws['B19'] = 'Budget vs Contributions Reconciliation'
    ws['B19'].font = Font(bold=True, size=11)
    ws['B20'] = 'Total budget cap'
    ws['D20'] = "='🧭 Setup Wizard'!E16"
    ws['B21'] = 'Total committed (spent)'
    ws['D21'] = "=SUM('💰 Budget Categories'!E10:E23)"
    ws['B22'] = 'Total contributions received'
    ws['D22'] = '=SUM(E10:E14)'
    ws['B23'] = 'Net out-of-pocket (committed minus received)'
    ws['D23'] = '=MAX(0,D21-D22)'
    ws['B24'] = 'Uncovered / over-budget gap'
    ws['D24'] = '=MAX(0,D21-D20)'

    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 28

    ws.freeze_panes = 'A7'

    print(f'  Added Contribution Tracker sheet to {edition_name}')

for key, edition in [('pro', 'Pro'), ('ai-edition', 'AI Edition')]:
    out_path = os.path.join(FIXED_DIR, f'wedding-budget-planner-{key}.xlsx')
    if not os.path.exists(out_path):
        continue
    wb = load_workbook(out_path, data_only=False)
    add_contribution_sheet(wb, edition)
    wb.save(out_path)
    log_complement('CONTRIB-TRACKER', key, 'Contribution Tracker', 'A1:I26',
                   f'Added Contribution Tracker to {edition}: named sources (bride family / groom family / couple / gifts), pledged vs received, net out-of-pocket, uncovered gap formula')


# ══════════════════════════════════════════════════════════════════════════════
# COMPLEMENT 4 — FX Rate Table (Persona 3 / WBP-064)
# Add FX Rate table to AI Edition (destination wedding support)
# ══════════════════════════════════════════════════════════════════════════════
print('\n' + '='*70)
print('COMPLEMENT 4 — FX Rate Table (WBP-064)')
print('='*70)

def add_fx_sheet(wb, edition_name):
    for ws in wb.worksheets:
        if 'fx' in ws.title.lower() or 'currency' in ws.title.lower() or 'exchange' in ws.title.lower():
            print(f'  FX sheet already exists: {ws.title}')
            return

    ws = wb.create_sheet('FX Rates & Currency', 4)
    ws.sheet_properties.tabColor = openpyxl.styles.colors.Color(rgb='FF8FA98F')

    header_fill = PatternFill(start_color='FF8B5A6B', end_color='FF8B5A6B', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF', size=11)
    kpi_fill = PatternFill(start_color='FFC9A0A0', end_color='FFC9A0A0', fill_type='solid')

    ws['A1'] = '    Lime Premium Studios'
    ws['D1'] = f'Wedding Budget & Planner — {edition_name}'
    ws['J1'] = 'FX Rates & Currency'
    ws['J1'].font = Font(bold=True, size=12)

    ws['B3'] = 'Multi-Currency Support for Destination Weddings'
    ws['B3'].font = Font(bold=True, size=13)
    ws['B4'] = 'Enter the exchange rate for each currency you pay vendors in. All amounts convert to your base currency (set in Setup Wizard). Rates update the Vendor Cost table below automatically.'

    ws['B5'] = 'Base currency (from Setup Wizard):'
    ws['E5'] = "='🧭 Setup Wizard'!E20"
    ws['E5'].font = Font(bold=True, size=12)

    # FX Rate Table header row 7
    headers = [('B7', 'Currency Code'), ('C7', 'Currency Name'), ('D7', 'Rate → Base Currency'),
               ('E7', 'Example: 100 in this currency ='), ('F7', 'Notes')]
    for addr, label in headers:
        c = ws[addr]
        c.value = label
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.row_dimensions[7].height = 30

    # Pre-filled FX rows
    fx_rates = [
        ('EUR', 'Euro', 1.08, 'European vendors (Italy, France, Spain, Greece)'),
        ('GBP', 'British Pound Sterling', 1.27, 'UK vendors'),
        ('AED', 'UAE Dirham', 0.2723, 'Dubai / UAE vendors'),
        ('EGP', 'Egyptian Pound', 0.0205, 'Egypt vendors'),
        ('INR', 'Indian Rupee', 0.01196, 'India vendors'),
        ('CAD', 'Canadian Dollar', 0.736, 'Canada vendors'),
        ('AUD', 'Australian Dollar', 0.643, 'Australia vendors'),
        ('CHF', 'Swiss Franc', 1.09, 'Switzerland vendors'),
        ('MXN', 'Mexican Peso', 0.0575, 'Mexico vendors'),
        ('THB', 'Thai Baht', 0.0279, 'Thailand / Phuket vendors'),
    ]
    for i, (code, name, rate, note) in enumerate(fx_rates):
        r = 8 + i
        ws.cell(row=r, column=2).value = code
        ws.cell(row=r, column=3).value = name
        ws.cell(row=r, column=4).value = rate
        ws.cell(row=r, column=5).value = f'={rate}*100 USD'  # static example
        ws.cell(row=r, column=6).value = note

    # Vendor Cost (Multi-Currency) section row 21
    ws['B21'] = 'Vendor Costs in Foreign Currencies'
    ws['B21'].font = Font(bold=True, size=12)
    ws['B22'] = 'Enter vendor name, currency, and local price. The USD equivalent is auto-calculated using your FX table above.'

    vc_headers = [('B23', 'Vendor'), ('C23', 'Category'), ('D23', 'Currency'), ('E23', 'Local Amount'),
                  ('F23', 'FX Rate (from above)'), ('G23', 'USD Equivalent'), ('H23', 'Notes')]
    for addr, label in vc_headers:
        c = ws[addr]
        c.value = label
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.row_dimensions[23].height = 30

    # Sample vendor rows 24-33
    for i in range(10):
        r = 24 + i
        ws.cell(row=r, column=6).value = f'=IFERROR(VLOOKUP(D{r},B8:D17,3,FALSE),1)'
        ws.cell(row=r, column=7).value = f'=IFERROR(E{r}*F{r},0)'

    # Column widths
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 30

    ws.freeze_panes = 'A8'

    print(f'  Added FX Rates & Currency sheet to {edition_name}')

for key, edition in [('pro', 'Pro'), ('ai-edition', 'AI Edition')]:
    out_path = os.path.join(FIXED_DIR, f'wedding-budget-planner-{key}.xlsx')
    if not os.path.exists(out_path):
        continue
    wb = load_workbook(out_path, data_only=False)
    add_fx_sheet(wb, edition)
    wb.save(out_path)
    log_complement('FX-RATES', key, 'FX Rates & Currency', 'A1:H33',
                   f'Added FX Rate table to {edition}: 10 pre-filled currencies (EUR/GBP/AED/EGP/INR/CAD/AUD/CHF/MXN/THB), VLOOKUP cascade to vendor cost table, base currency pulled from Setup Wizard')


# ══════════════════════════════════════════════════════════════════════════════
# COMPLEMENT 5 — Per-Guest Cost KPI on Budget Dashboard (WBP-061)
# ══════════════════════════════════════════════════════════════════════════════
print('\n' + '='*70)
print('COMPLEMENT 5 — Per-Guest Cost KPI on Budget Dashboard (WBP-061)')
print('='*70)

for key in ['essentials', 'pro', 'ai-edition']:
    out_path = os.path.join(FIXED_DIR, f'wedding-budget-planner-{key}.xlsx')
    if not os.path.exists(out_path):
        continue
    wb = load_workbook(out_path, data_only=False)
    for ws in wb.worksheets:
        if 'budget dashboard' in ws.title.lower():
            # Row 2 KPI bar: A2=budget cap, C2=spent, E2=% used, G2=days to go, I2=guests RSVPd, K2=top vendor
            # K2 is merged K2:M2 — we can write to it (it's a MergedCellRange but the top-left is writable)
            # We'll update K2 from "TOP VENDOR" to "PER-GUEST COST" since per-guest is more critical
            # and "top vendor" info is shown in rows 37-46 below
            # Check if K2 is the top-left cell of the merge
            k2_cell = ws.cell(row=2, column=11)  # K2
            old_val = k2_cell.value
            if old_val is not None:  # it's the top-left of the merge, writable
                k2_cell.value = '="PER-GUEST COST"&CHAR(10)&TEXT(IFERROR(SUM(\'💰 Budget Categories\'!E10:E23)/MAX(1,\'🧭 Setup Wizard\'!E12),0),"$#,##0")&CHAR(10)&"/ head"'
                k2_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                log_complement('PER-GUEST-KPI', key, ws.title, 'K2',
                               'Updated K2 KPI from TOP VENDOR to PER-GUEST COST (total spend / confirmed guest count from Setup Wizard)')
                print(f'  {key}: Updated K2 from "{str(old_val)[:40]}" to PER-GUEST COST')
            else:
                print(f'  {key}: K2 is None or merged secondary cell — skipping')
    wb.save(out_path)


# ══════════════════════════════════════════════════════════════════════════════
# Verify fixes by recalculating
# ══════════════════════════════════════════════════════════════════════════════
print('\n' + '='*70)
print('VERIFYING FIXES VIA LIBREOFFICE RECALC')
print('='*70)

fixed_recalc_dir = os.path.join(SCRATCH, 'fixed-recalc')
os.makedirs(fixed_recalc_dir, exist_ok=True)

for key in ['essentials', 'pro', 'ai-edition']:
    fixed_path = os.path.join(FIXED_DIR, f'wedding-budget-planner-{key}.xlsx')
    if not os.path.exists(fixed_path):
        print(f'  SKIP: {fixed_path} not found')
        continue
    cmd = [LO, '--headless', '--calc', '--convert-to', 'xlsx', '--outdir', fixed_recalc_dir, fixed_path]
    result = subprocess.run(cmd, capture_output=True, timeout=90)
    recalc_out = os.path.join(fixed_recalc_dir, f'wedding-budget-planner-{key}.xlsx')
    if os.path.exists(recalc_out):
        wb_check = load_workbook(recalc_out, data_only=True)
        # Check for #VALUE! in key cells
        value_errors = []
        for ws in wb_check.worksheets:
            for row in ws.iter_rows(min_row=2, max_row=2):
                for cell in row:
                    if cell.value == '#VALUE!':
                        value_errors.append(f'{ws.title}/{cell.coordinate}')
        if value_errors:
            print(f'  {key}: STILL HAS #VALUE! in {value_errors}')
        else:
            print(f'  {key}: No #VALUE! errors in row 2 KPI cells — fix confirmed')

        # Spot-check: read budget cap and spent from essentials
        if key == 'essentials':
            sw = wb_check.worksheets[0]
            dash = wb_check.worksheets[1]
            bc = wb_check.worksheets[2]
            g2_val = sw['G2'].value
            i2_val = sw['I2'].value
            a2_dash = dash['A2'].value
            print(f'  Setup Wizard G2 (Budget Cap KPI): {g2_val}')
            print(f'  Setup Wizard I2 (Per-Guest KPI): {i2_val}')
            print(f'  Dashboard A2 (Budget Cap KPI): {a2_dash}')
    else:
        print(f'  {key}: Recalc output not found!')


# ══════════════════════════════════════════════════════════════════════════════
# Save Changelog
# ══════════════════════════════════════════════════════════════════════════════
changelog_path = 'C:/ETSY/etsy-store/tools/qa/output/fix-changelog.md'
os.makedirs(os.path.dirname(changelog_path), exist_ok=True)

with open(changelog_path, 'w', encoding='utf-8') as f:
    f.write('# Wedding Budget & Planner — Fix + Complement Changelog\n\n')
    f.write(f'Applied {fix_count} fixes and {len([c for c in changelog if "[COMPLEMENT" in c])} complements\n')
    f.write(f'Date: {datetime.date.today()}\n\n')
    f.write('---\n\n')
    for entry in changelog:
        f.write(entry + '\n')

print(f'\n  Changelog saved to {changelog_path}')
print(f'\nDONE — {fix_count} fixes, {len(changelog)} total entries.')
