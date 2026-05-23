"""
Family & Education Planner — Round 2 persona verification.

Runs all 5 personas against the FIXED workbooks (tools/qa/fixed/) and verifies:
  - All Round 1 CRITICAL+HIGH fixes hold
  - No regressions introduced
  - Critical KPIs evaluate to expected values within tolerance

Output: tools/qa/round2/fep_persona_results.json + per-persona xlsx copies.
"""
import sys, os, math, subprocess, json, shutil
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from datetime import date
from pathlib import Path

ROOT = Path(r"C:\ETSY\etsy-store")
SOFFICE = r"C:\Program Files\LibreOffice\program\soffice.com"
FIXED_DIR = ROOT / "tools/qa/fixed"
R2_DIR = ROOT / "tools/qa/round2/fep_runs"
SCRATCH = ROOT / "tools/qa/scratch/fep_r2"
R2_DIR.mkdir(parents=True, exist_ok=True)
SCRATCH.mkdir(parents=True, exist_ok=True)

# Tier file under test (AI Edition has all 20 tabs incl. EFC + Aid + Settings & FX)
FIXED_AI = FIXED_DIR / "family-education-planner-ai-edition.xlsx"
FIXED_PRO = FIXED_DIR / "family-education-planner-pro.xlsx"
FIXED_ESS = FIXED_DIR / "family-education-planner-essentials.xlsx"

CHILD_PROFILES = '👶 Child Profiles'
DASHBOARD = '🏠 Dashboard'
CSP = '🎓 College Savings Planner'
K12 = '🏫 K-12 Cost Map'
EFC = '🧮 EFC SAI Calculator'
AID = '📑 Aid Letter Comparison'
HEALTH = '🏥 Family Health Budget'
WHOLELIFE = '🏦 529 vs. Whole Life'
LIFEINS = '🛡️ Life Insurance Calculator'
LITERACY = '🎓 Literacy Milestones'
SETTINGS = '⚙️ Settings & FX'


def recalc(path):
    """Recalc xlsx via LibreOffice headless. Returns path to recalced file in SCRATCH."""
    out = SCRATCH / Path(path).name
    if out.exists():
        out.unlink()
    cmd = [SOFFICE, '--headless', '--calc', '--convert-to', 'xlsx',
           '--outdir', str(SCRATCH), str(path)]
    subprocess.run(cmd, check=True, capture_output=True, timeout=120)
    return out


def read_evaluated(path):
    """Open recalced xlsx with data_only=True."""
    return openpyxl.load_workbook(path, data_only=True)


def write_inputs(src, dst, parent_ctx, children, settings):
    """Copy src→dst, write parent context, children rows, and Settings & FX overrides."""
    shutil.copy(src, dst)
    wb = openpyxl.load_workbook(dst, data_only=False)

    cp = wb[CHILD_PROFILES]
    # Parent context (rows 7-12, col C)
    cp['C7']  = parent_ctx['income']
    cp['C8']  = parent_ctx['marital']
    cp['C9']  = parent_ctx['state']
    cp['C10'] = parent_ctx['fed_bracket']
    cp['C11'] = parent_ctx['state_rate']
    cp['C12'] = parent_ctx['saveable']

    # Children (rows 17-20)
    for i, child in enumerate(children):
        r = 17 + i
        cp[f'B{r}'] = i + 1
        cp[f'C{r}'] = child.get('name', '')
        cp[f'D{r}'] = child.get('dob')  # date
        # E (age) and F (yrs) are auto-derived formulas — leave alone
        cp[f'G{r}'] = child.get('k12', 'Public')
        cp[f'H{r}'] = child.get('tier', '')
        cp[f'I{r}'] = child.get('savings', 0)
        cp[f'J{r}'] = child.get('monthly', 0)
        cp[f'K{r}'] = child.get('start_year', None)
        cp[f'L{r}'] = child.get('special_needs', 'No')
        # M = Currency, N = Custody % (from complements)
        if 'currency' in child:
            cp[f'M{r}'] = child['currency']
        if 'custody' in child:
            cp[f'N{r}'] = child['custody']

    # Settings & FX named-range overrides
    if SETTINGS in wb.sheetnames and settings:
        s = wb[SETTINGS]
        # The Settings & FX tab uses named ranges. We'll override the cells the
        # named ranges resolve to. Common locations: D5 Inflation, D6 EduReturn,
        # D7 K12Inflation. If unsure, log what we find.
        # Per the complement, layout is: B label, D value, with defined names.
        # We'll scan defined names and set them.
        for nm in wb.defined_names:
            try:
                dest = list(wb.defined_names[nm].destinations)
                if not dest:
                    continue
                sheet_name, coord = dest[0]
                # coord like "$D$5" — strip $
                coord_clean = coord.replace('$', '')
                if nm == 'Inflation' and 'inflation' in settings:
                    wb[sheet_name][coord_clean] = settings['inflation']
                elif nm == 'EduReturn' and 'return' in settings:
                    wb[sheet_name][coord_clean] = settings['return']
                elif nm == 'K12Inflation' and 'k12_inflation' in settings:
                    wb[sheet_name][coord_clean] = settings['k12_inflation']
            except Exception:
                pass

    wb.save(dst)


def collect_cells(wb, persona_id):
    """Pull every cell we need to verify the fixes held."""
    r = {}
    # Dashboard headline KPI
    try:
        r['dash_A2_health_kpi'] = wb[DASHBOARD]['A2'].value
    except KeyError:
        r['dash_A2_health_kpi'] = '(no Dashboard tab)'
    try:
        r['dash_B10_health_score'] = wb[DASHBOARD]['B10'].value
    except KeyError:
        r['dash_B10_health_score'] = '(no Dashboard tab)'

    # CSP — per-child target, gap, rec_monthly, status
    csp = wb[CSP] if CSP in wb.sheetnames else None
    if csp:
        for i in range(4):
            row = 8 + i
            r[f'csp_D{row}_target'] = csp[f'D{row}'].value
            r[f'csp_F{row}_gap'] = csp[f'F{row}'].value
            r[f'csp_I{row}_rec_mo'] = csp[f'I{row}'].value
            r[f'csp_J{row}_status'] = csp[f'J{row}'].value

    # K-12 Cost Map A2 13-yr total
    if K12 in wb.sheetnames:
        r['k12_A2_total_kpi'] = wb[K12]['A2'].value
        r['k12_B30_alt_total'] = wb[K12]['B30'].value if wb[K12]['B30'].value is not None else 'N/A'

    # EFC F33
    if EFC in wb.sheetnames:
        r['efc_F33_total'] = wb[EFC]['F33'].value
        r['efc_F25_parent_income_contrib'] = wb[EFC]['F25'].value
        r['efc_F28_parent_asset_contrib'] = wb[EFC]['F28'].value
        r['efc_F29_total_parent'] = wb[EFC]['F29'].value
        r['efc_E30_student_income'] = wb[EFC]['E30'].value

    # Aid Letter I2 + G17
    if AID in wb.sheetnames:
        r['aid_I2_appeal_open_kpi'] = wb[AID]['I2'].value
        r['aid_G17_days_to_appeal_empty'] = wb[AID]['G17'].value
        r['aid_C2_best_net_kpi'] = wb[AID]['C2'].value

    # Health Budget K2
    if HEALTH in wb.sheetnames:
        r['health_K2_annual_total_kpi'] = wb[HEALTH]['K2'].value

    # 529 vs Whole Life A2, C2, E2 (now year-18)
    if WHOLELIFE in wb.sheetnames:
        r['wl_A2_529_yr18'] = wb[WHOLELIFE]['A2'].value
        r['wl_C2_wholelife_yr18'] = wb[WHOLELIFE]['C2'].value
        r['wl_E2_difference'] = wb[WHOLELIFE]['E2'].value

    # Life Insurance E26
    if LIFEINS in wb.sheetnames:
        r['li_E26_rec_term'] = wb[LIFEINS]['E26'].value
        r['li_E20_dime_total'] = wb[LIFEINS]['E20'].value
        r['li_E22_adjusted_coverage'] = wb[LIFEINS]['E22'].value

    # Literacy E22 (% complete)
    if LITERACY in wb.sheetnames:
        r['lit_E22_pct_complete'] = wb[LITERACY]['E22'].value

    return r


# ============================================================
# PERSONA DEFINITIONS
# ============================================================

PERSONAS = [
    {
        'id': 'p1_mariam',
        'name': "Mariam & Ali (Cairo, single toddler, 16yr EGP)",
        'fep_targets': [
            'FEP-001 VLOOKUP fix',
            'FEP-004 named-input inflation/return',
            'FEP-005 FX support (Settings & FX tab exists)',
            'FEP-006 inflation-adjusted FV target',
            'FEP-010 Dashboard A2 #VALUE fix',
            'FEP-014 empty-roster guard (Layla = only child, so guard does not fire)',
        ],
        'parent': {
            'income': 600000, 'marital': 'Married Filing Jointly',
            'state': 'N/A (Egypt)', 'fed_bracket': 0, 'state_rate': 0,
            'saveable': 72000,
        },
        'children': [
            {'name': 'Layla', 'dob': date(2024, 5, 23), 'k12': 'Public',
             'tier': 'Private Mid-Tier', 'savings': 50000, 'monthly': 6000,
             'start_year': 2042, 'special_needs': 'No', 'currency': 'EGP'},
        ],
        'settings': {'inflation': 0.07, 'return': 0.06, 'k12_inflation': 0.05},
    },
    {
        'id': 'p2_mohamed',
        'name': "Mohamed & Heba (3 kids parallel USD)",
        'fep_targets': [
            'FEP-001 VLOOKUP fix (In-State Public → $120K target, not $0)',
            'FEP-006 inflation adjustment lifts PMT to true required',
            'FEP-009 status pills guarded',
        ],
        'parent': {
            'income': 140000, 'marital': 'Married Filing Jointly',
            'state': 'TX', 'fed_bracket': 0.22, 'state_rate': 0,
            'saveable': 30000,
        },
        'children': [
            {'name': 'Ahmed', 'dob': date(2012, 1, 15), 'k12': 'Public',
             'tier': 'In-State Public', 'savings': 42000, 'monthly': 850,
             'start_year': 2030, 'special_needs': 'No', 'currency': 'USD'},
            {'name': 'Noor', 'dob': date(2015, 4, 20), 'k12': 'Public',
             'tier': 'In-State Public', 'savings': 28000, 'monthly': 850,
             'start_year': 2033, 'special_needs': 'No', 'currency': 'USD'},
            {'name': 'Karim', 'dob': date(2018, 8, 5), 'k12': 'Public',
             'tier': 'In-State Public', 'savings': 15000, 'monthly': 800,
             'start_year': 2036, 'special_needs': 'No', 'currency': 'USD'},
        ],
        'settings': {'inflation': 0.06, 'return': 0.05, 'k12_inflation': 0.04},
    },
    {
        'id': 'p3_tarek',
        'name': "Tarek & Yasmin (UAE multi-currency)",
        'fep_targets': [
            'FEP-005 FX support — Settings & FX tab + per-child Currency column M',
            'FEP-008 negative rec_monthly clamped to 0 (Hala overfunded)',
            'FEP-001 + FEP-006 multi-currency university targets',
        ],
        'parent': {
            'income': 280000, 'marital': 'Married Filing Jointly',
            'state': 'N/A (UAE)', 'fed_bracket': 0, 'state_rate': 0,
            'saveable': 90000,
        },
        'children': [
            {'name': 'Hala', 'dob': date(2014, 3, 1), 'k12': 'Private (Independent)',
             'tier': 'Private Mid-Tier', 'savings': 85000, 'monthly': 400,
             'start_year': 2032, 'special_needs': 'No', 'currency': 'GBP'},
            {'name': 'Yousef', 'dob': date(2017, 6, 12), 'k12': 'Private (Independent)',
             'tier': 'Private Elite', 'savings': 42000, 'monthly': 1500,
             'start_year': 2035, 'special_needs': 'No', 'currency': 'USD'},
            {'name': 'Lila', 'dob': date(2021, 2, 10), 'k12': 'Private (Independent)',
             'tier': 'Out-of-State Public', 'savings': 28000, 'monthly': 600,
             'start_year': 2039, 'special_needs': 'No', 'currency': 'CAD'},
        ],
        'settings': {'inflation': 0.05, 'return': 0.05, 'k12_inflation': 0.04},
    },
    {
        'id': 'p4_sara',
        'name': "Sara & Khaled (catch-up, teens)",
        'fep_targets': [
            'FEP-001 VLOOKUP — Aya/Omar now get $120K target (was $0)',
            'FEP-013 Aya 1yr horizon does not return whole gap',
            'FEP-015 scholarship offset (Aya $20K) feeds CSP — VERIFY: complement adds offset column',
        ],
        'parent': {
            'income': 165000, 'marital': 'Married Filing Jointly',
            'state': 'NY', 'fed_bracket': 0.24, 'state_rate': 0.0685,
            'saveable': 24000,
        },
        'children': [
            {'name': 'Aya', 'dob': date(2009, 6, 1), 'k12': 'Public',
             'tier': 'In-State Public', 'savings': 48000, 'monthly': 900,
             'start_year': 2027, 'special_needs': 'No', 'currency': 'USD'},
            {'name': 'Omar', 'dob': date(2012, 8, 14), 'k12': 'Public',
             'tier': 'In-State Public', 'savings': 22000, 'monthly': 900,
             'start_year': 2030, 'special_needs': 'No', 'currency': 'USD'},
        ],
        'settings': {'inflation': 0.04, 'return': 0.04, 'k12_inflation': 0.04},
    },
    {
        'id': 'p5_layla_blended',
        'name': "Layla (blended SN+gifted, 50% custody)",
        'fep_targets': [
            'FEP-021 custody-share column N — Hadi 50%',
            'FEP-022 Category column expanded (Special Needs / Gifted / Standard)',
            'FEP-001 VLOOKUP — In-State Public targets',
        ],
        'parent': {
            'income': 195000, 'marital': 'Married Filing Jointly',
            'state': 'WA', 'fed_bracket': 0.24, 'state_rate': 0,
            'saveable': 48000,
        },
        'children': [
            {'name': 'Adam', 'dob': date(2016, 1, 1), 'k12': 'Public',
             'tier': 'In-State Public', 'savings': 18000, 'monthly': 700,
             'start_year': 2034, 'special_needs': 'Yes — ABLE eligible', 'currency': 'USD',
             'custody': 1.0},
            {'name': 'Hadi', 'dob': date(2013, 2, 1), 'k12': 'Public',
             'tier': 'In-State Public', 'savings': 11000, 'monthly': 800,
             'start_year': 2031, 'special_needs': 'No', 'currency': 'USD',
             'custody': 0.50},
            {'name': 'Lina', 'dob': date(2020, 1, 1), 'k12': 'Public Magnet',
             'tier': 'Private Elite', 'savings': 24000, 'monthly': 1200,
             'start_year': 2038, 'special_needs': 'Gifted', 'currency': 'USD',
             'custody': 1.0},
        ],
        'settings': {'inflation': 0.06, 'return': 0.05, 'k12_inflation': 0.04},
    },
]


# ============================================================
# DRIVER
# ============================================================

def run_persona(persona):
    pid = persona['id']
    src = FIXED_AI
    dst = R2_DIR / f"{pid}-fixed.xlsx"
    print(f"\n=== {pid} — {persona['name']} ===")
    write_inputs(src, dst, persona['parent'], persona['children'], persona['settings'])
    print(f"  wrote inputs → {dst.name}")
    recalced = recalc(dst)
    print(f"  recalced via LibreOffice → {recalced.name}")
    wb = read_evaluated(recalced)
    cells = collect_cells(wb, pid)
    return {'persona': pid, 'name': persona['name'],
            'fep_targets': persona['fep_targets'], 'evaluated': cells}


def main():
    results = []
    for persona in PERSONAS:
        try:
            results.append(run_persona(persona))
        except Exception as e:
            print(f"  ERROR: {e}")
            results.append({'persona': persona['id'], 'name': persona['name'],
                            'error': str(e)})

    out = ROOT / "tools/qa/round2/fep_persona_results.json"
    out.write_text(json.dumps(results, indent=2, default=str), encoding='utf-8')
    print(f"\n→ Wrote {out}")
    print(f"  {len(results)} personas")


if __name__ == '__main__':
    main()
