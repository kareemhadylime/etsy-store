"""
Budget Tracker QA — deep-inspection dump.
Reads xlsx (openpyxl), pdf (pypdf), png (Pillow) and prints a structured dump
of everything reviewer needs to reason about.
"""
import os, sys, json, hashlib
from pathlib import Path

ROOT = Path("C:/ETSY/etsy-store")
OUT  = ROOT / "tools/qa/output/raw-inspection.txt"

XLSX = [
    ROOT / "tools/sheets-gen/output/budget-tracker-essentials.xlsx",
    ROOT / "tools/sheets-gen/output/budget-tracker-pro.xlsx",
    ROOT / "tools/sheets-gen/output/budget-tracker-ai-edition-v2.xlsx",
]
PDFS = [
    ROOT / "tools/pdf-gen/output/budget-tracker-ai-pdf.pdf",
    ROOT / "tools/pdf-gen/output/budget-tracker-quickstart.pdf",
]
PNGS = [
    ROOT / "tools/thumb-gen/output/budget-tracker-01-hero.png",
    ROOT / "tools/thumb-gen/output/budget-tracker-02-health-score.png",
    ROOT / "tools/thumb-gen/output/budget-tracker-03-methods.png",
    ROOT / "tools/thumb-gen/output/budget-tracker-04-ai-advisor.png",
    ROOT / "tools/thumb-gen/output/budget-tracker-05-privacy.png",
]

lines = []
def p(*a):
    s = " ".join(str(x) for x in a)
    lines.append(s)
    try:
        sys.stdout.buffer.write((s + "\n").encode("utf-8", errors="replace"))
    except Exception:
        pass

# ---------------- XLSX ----------------
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def dump_xlsx(path):
    p("="*80)
    p("XLSX:", path.name, "size=", path.stat().st_size)
    # 1) Pass with data_only=True for cached values; 2) without for formulas
    wb_f = load_workbook(path, data_only=False)
    wb_v = load_workbook(path, data_only=True)
    p("Sheet names:", wb_f.sheetnames)
    p("Defined names:")
    for dn in wb_f.defined_names:
        try:
            d = wb_f.defined_names[dn]
            p("  ", dn, "->", d.attr_text if hasattr(d,'attr_text') else d.value)
        except Exception as e:
            p("  ", dn, "ERR", e)

    for sn in wb_f.sheetnames:
        ws_f = wb_f[sn]; ws_v = wb_v[sn]
        p("-"*60)
        p("Sheet:", sn, " dim=", ws_f.dimensions, " max_row=", ws_f.max_row, " max_col=", ws_f.max_column,
          " tab_color=", (ws_f.sheet_properties.tabColor.rgb if ws_f.sheet_properties.tabColor else None))
        p("  freeze_panes=", ws_f.freeze_panes)
        # column widths
        widths = {}
        for col_letter, dim in ws_f.column_dimensions.items():
            widths[col_letter] = dim.width
        p("  col_widths=", widths)
        # row heights (only custom)
        heights = {r:d.height for r,d in ws_f.row_dimensions.items() if d.height}
        if heights:
            p("  row_heights=", heights)
        # print area / page breaks
        try:
            p("  print_area=", ws_f.print_area)
        except Exception:
            pass
        try:
            p("  print_title_rows=", ws_f.print_title_rows)
            p("  print_title_cols=", ws_f.print_title_cols)
        except Exception:
            pass
        # data validations
        if ws_f.data_validations and ws_f.data_validations.dataValidation:
            p("  data_validations:")
            for dv in ws_f.data_validations.dataValidation:
                p("    type=", dv.type, " formula1=", dv.formula1, " ranges=", [str(r) for r in dv.sqref.ranges] if dv.sqref else None)
        # conditional formatting
        if ws_f.conditional_formatting:
            p("  conditional_formatting:")
            for rng, rules in ws_f.conditional_formatting._cf_rules.items():
                rng_str = str(rng.sqref) if hasattr(rng,'sqref') else str(rng)
                for rule in rules:
                    p("    range=", rng_str, " type=", rule.type, " operator=", rule.operator,
                      " formula=", rule.formula, " text=", getattr(rule,'text',None),
                      " color=", (rule.dxf.fill.bgColor.rgb if rule.dxf and rule.dxf.fill and rule.dxf.fill.bgColor else None) if rule.dxf else None)
        # merged cells
        if ws_f.merged_cells.ranges:
            p("  merged=", [str(r) for r in ws_f.merged_cells.ranges][:20])

        # cells dump: first 60 rows, all cols up to 16
        p("  ---- cells ----")
        max_r = min(ws_f.max_row or 0, 80)
        max_c = min(ws_f.max_column or 0, 16)
        for r in range(1, max_r+1):
            for c in range(1, max_c+1):
                cf = ws_f.cell(r,c); cv = ws_v.cell(r,c)
                if cf.value is None and cv.value is None: continue
                fill = cf.fill.fgColor.rgb if cf.fill and cf.fill.fgColor else None
                font = (cf.font.name, cf.font.size, cf.font.bold, (cf.font.color.rgb if cf.font.color else None))
                p(f"    {get_column_letter(c)}{r}: f={cf.value!r}  v={cv.value!r}  fill={fill}  font={font}  fmt={cf.number_format!r}")
        # charts
        if ws_f._charts:
            p("  charts:", len(ws_f._charts))
            for i,ch in enumerate(ws_f._charts):
                p("    chart", i, type(ch).__name__, "title=", getattr(ch,'title',None))

# ---------------- PDF ----------------
def dump_pdf(path):
    p("="*80)
    p("PDF:", path.name, "size=", path.stat().st_size)
    from pypdf import PdfReader
    r = PdfReader(str(path))
    p("pages:", len(r.pages))
    meta = r.metadata or {}
    p("meta:", {k:str(v) for k,v in meta.items()})
    for i, pg in enumerate(r.pages):
        txt = pg.extract_text() or ""
        p("-"*60)
        p("Page", i+1, " size=", pg.mediabox.width, "x", pg.mediabox.height)
        # truncate but preserve full content per page
        for line in txt.splitlines():
            p("  ", line)

# ---------------- PNG ----------------
def dump_png(path):
    from PIL import Image
    p("="*80)
    p("PNG:", path.name, "size=", path.stat().st_size)
    img = Image.open(path)
    p("  format=", img.format, "mode=", img.mode, "size=", img.size, "dpi=", img.info.get('dpi'))
    # palette: dominant colors via quantize
    small = img.convert('RGB').resize((100,100))
    q = small.quantize(colors=8)
    pal = q.getpalette()[:24]
    counts = q.getcolors() or []
    counts.sort(reverse=True)
    p("  dominant_colors:")
    for cnt, idx in counts[:8]:
        r_,g_,b_ = pal[idx*3:idx*3+3]
        p(f"    rgb({r_},{g_},{b_}) #{r_:02x}{g_:02x}{b_:02x} pct={cnt/100:.1f}%")

# run
for f in XLSX:
    try: dump_xlsx(f)
    except Exception as e:
        import traceback; p("ERROR xlsx", f.name, e); p(traceback.format_exc())
for f in PDFS:
    try: dump_pdf(f)
    except Exception as e:
        import traceback; p("ERROR pdf", f.name, e); p(traceback.format_exc())
for f in PNGS:
    try: dump_png(f)
    except Exception as e:
        import traceback; p("ERROR png", f.name, e); p(traceback.format_exc())

OUT.write_text("\n".join(lines), encoding='utf-8')
print(f"\nWrote {OUT} ({OUT.stat().st_size} bytes)")
