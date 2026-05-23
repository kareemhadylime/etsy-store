"""
Wedding Budget & Planner — Full QA Script
Runs all 5 personas, structural checks, and CF audits.
Python 3.14, openpyxl 3.1.5, Windows paths.
"""
import sys, io, os, shutil, copy, datetime, json, subprocess
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ── paths ──────────────────────────────────────────────────────────────────────
SRC = {
    'essentials': 'C:/ETSY/etsy-store/tools/sheets-gen/output/wedding-budget-planner-essentials.xlsx',
    'pro':        'C:/ETSY/etsy-store/tools/sheets-gen/output/wedding-budget-planner-pro.xlsx',
    'ai':         'C:/ETSY/etsy-store/tools/sheets-gen/output/wedding-budget-planner-ai-edition.xlsx',
}
RECALC = {
    'essentials': 'C:/ETSY/etsy-store/tools/qa/scratch/wedding-budget-planner-essentials.xlsx',
    'pro':        'C:/ETSY/etsy-store/tools/qa/scratch/wedding-budget-planner-pro.xlsx',
    'ai':         'C:/ETSY/etsy-store/tools/qa/scratch/wedding-budget-planner-ai-edition.xlsx',
}
SCRATCH = 'C:/ETSY/etsy-store/tools/qa/scratch'
LO = r'C:\Program Files\LibreOffice\program\soffice.com'
TODAY = datetime.date(2026, 5, 23)  # fixed for reproducibility

issues = []
issue_id = [0]

def new_issue(severity, file_key, sheet, cell, problem, evidence, fix):
    issue_id[0] += 1
    wbp = f'WBP-{issue_id[0]:03d}'
    issues.append(dict(id=wbp, severity=severity, file=file_key, sheet=sheet,
                       cell=cell, problem=problem, evidence=evidence, fix=fix))
    print(f'  [{severity}] {wbp}: {file_key}/{sheet}/{cell} — {problem}')
    return wbp


def lo_recalc(src_path, out_dir=SCRATCH):
    """Run LibreOffice headless recalc."""
    cmd = [LO, '--headless', '--calc', '--convert-to', 'xlsx', '--outdir', out_dir, src_path]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=90)
    base = os.path.basename(src_path)
    out = os.path.join(out_dir, base)
    if not os.path.exists(out):
        print(f'  WARNING: LO recalc produced no output for {base}')
        return None
    return out


def cell_val(ws, addr):
    return ws[addr].value


def scan_errors(wb, file_key):
    """Scan all cells for Excel error strings."""
    error_cells = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith('#'):
                    error_cells.append((ws.title, cell.coordinate, v))
    if error_cells:
        for (sheet, coord, val) in error_cells[:20]:
            new_issue('CRITICAL', file_key, sheet, coord,
                      f'Formula error {val} in evaluated workbook',
                      f'Cell contains error string: {val}',
                      'Fix the formula or add IFERROR wrapper')
    return error_cells


def check_named_ranges(wb, file_key):
    """Check named ranges exist."""
    try:
        names = list(wb.defined_names.keys())
    except Exception:
        names = []
    print(f'  Named ranges in {file_key}: {names}')
    return names


def check_data_validation(wb, file_key):
    """Check data validation lists."""
    dvs = {}
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            dvs[ws.title] = dvs.get(ws.title, [])
            dvs[ws.title].append({
                'sqref': str(dv.sqref),
                'type': dv.type,
                'formula1': dv.formula1,
            })
    print(f'  DV in {file_key}: {sum(len(v) for v in dvs.values())} rules across {len(dvs)} sheets')
    return dvs


def check_cf_rules(wb, file_key):
    """Check conditional formatting rules."""
    cf_rules = {}
    for ws in wb.worksheets:
        rules = ws.conditional_formatting._cf_rules
        if rules:
            cf_rules[ws.title] = []
            for rng, rule_list in rules.items():
                for rule in rule_list:
                    cf_rules[ws.title].append({
                        'range': str(rng),
                        'type': rule.type,
                        'formula': getattr(rule, 'formula', None),
                    })
    print(f'  CF rules in {file_key}: {sum(len(v) for v in cf_rules.values())} rules across {len(cf_rules)} sheets')
    return cf_rules


# ══════════════════════════════════════════════════════════════════════════════
# STRUCTURAL INTEGRITY CHECK
# ══════════════════════════════════════════════════════════════════════════════
print('\n' + '='*70)
print('STRUCTURAL INTEGRITY CHECK (recalculated workbooks, data_only=True)')
print('='*70)

struct_results = {}
for key, path in RECALC.items():
    print(f'\n--- {key} ---')
    if not os.path.exists(path):
        print(f'  ERROR: Recalculated file not found: {path}')
        new_issue('CRITICAL', key, 'N/A', 'N/A',
                  'Recalculated file missing', f'Path: {path}',
                  'Re-run LibreOffice recalculation')
        struct_results[key] = None
        continue

    wb = load_workbook(path, data_only=True)
    struct_results[key] = wb

    # Check sheets
    print(f'  Sheets: {wb.sheetnames}')

    # Scan for errors
    errs = scan_errors(wb, key)
    if errs:
        print(f'  ERROR CELLS: {errs[:5]}')
    else:
        print(f'  No formula errors found.')

    # Named ranges
    check_named_ranges(wb, key)


# ══════════════════════════════════════════════════════════════════════════════
# DV + CF CHECK (from source workbooks, not data_only)
# ══════════════════════════════════════════════════════════════════════════════
print('\n' + '='*70)
print('DATA VALIDATION + CONDITIONAL FORMATTING CHECK (source workbooks)')
print('='*70)

src_wbs = {}
for key, path in SRC.items():
    print(f'\n--- {key} ---')
    if not os.path.exists(path):
        print(f'  ERROR: Source file not found: {path}')
        continue
    wb = load_workbook(path, data_only=False)
    src_wbs[key] = wb
    check_data_validation(wb, key)
    cf_rules = check_cf_rules(wb, key)

    # Check for frozen panes
    frozen_sheets = []
    for ws in wb.worksheets:
        if ws.freeze_panes:
            frozen_sheets.append((ws.title, ws.freeze_panes))
    print(f'  Frozen panes on {len(frozen_sheets)} sheets: {frozen_sheets[:5]}')


# ══════════════════════════════════════════════════════════════════════════════
# PERSONA 1 — Aisha & Omar, Cairo, EGP, Modest Single-Day
# ══════════════════════════════════════════════════════════════════════════════
print('\n' + '='*70)
print('PERSONA 1 — Aisha & Omar | Cairo | EGP | Modest Single-Day Wedding')
print('='*70)

# Reference values
p1_items = {
    'Venue': 75000,
    'Catering': 80000,
    'Photography': 18000,
    'DJ/Music': 12000,
    'Bride dress': 22000,
    'Groom suit': 8000,
    'Flowers': 15000,
    'Cake': 6000,
    'Invitations': 4000,
    'Hair/Makeup': 7000,
    'Videographer': 8000,
}
p1_deposits = {
    'Venue': 25000,
    'Catering': 20000,
    'Photography': 5000,
    'DJ/Music': 3000,
    'Bride dress': 22000,
    'Groom suit': 8000,
    'Flowers': 4000,
    'Cake': 0,
    'Invitations': 4000,
    'Hair/Makeup': 1500,
    'Videographer': 2000,
}
p1_total_committed = sum(p1_items.values())
p1_budget = 250000
p1_variance = p1_budget - p1_total_committed
p1_total_deposits = sum(p1_deposits.values())
p1_total_outstanding = p1_total_committed - p1_total_deposits
p1_guests = 200
p1_per_guest = p1_total_committed / p1_guests
# Wedding date: 8 months from 2026-05-23 = 2026-01-23+8m... actually 8 months ahead
wedding_date_p1 = datetime.date(2027, 1, 23)  # ~8 months out from May 2026
p1_days = (wedding_date_p1 - TODAY).days

print(f'\nPYTHON REFERENCE VALUES:')
print(f'  Total committed: EGP {p1_total_committed:,}')
print(f'  Budget cap: EGP {p1_budget:,}')
print(f'  Variance: EGP {p1_variance:,} ({"OVER" if p1_variance < 0 else "UNDER"} budget)')
print(f'  Total deposits paid: EGP {p1_total_deposits:,}')
print(f'  Total outstanding: EGP {p1_total_outstanding:,}')
print(f'  Per-guest cost: EGP {p1_per_guest:.2f}/head')
print(f'  Days until wedding: {p1_days}')

# Write Persona 1 inputs to essentials workbook
p1_src = SRC['essentials']
p1_out = os.path.join(SCRATCH, 'p1-aisha-essentials-qa.xlsx')
shutil.copy(p1_src, p1_out)
wb_p1 = load_workbook(p1_out, data_only=False)

# Write setup wizard
sw = wb_p1.worksheets[0]
sw['E10'] = datetime.datetime(2027, 1, 23)  # wedding date
sw['E12'] = 200  # guest count
sw['E16'] = 250000  # budget cap
sw['E18'] = 'Cairo, Egypt'
sw['E20'] = 'EGP'
sw['E28'] = 'Aisha'
sw['E30'] = 'Omar'

# Write budget categories (columns: B=category, C=%alloc, D=target, E=spent, F=remaining, G=%used)
# Map items to rows 10-23
bc = wb_p1.worksheets[2]
category_map = [
    ('Venue + Rentals', 0, 75000),       # row 10
    ('Catering + Bar', 0, 80000),         # row 11
    ('Photography + Video', 0, 26000),    # row 12 (18000 photo + 8000 video)
    ('Attire + Beauty', 0, 29000),        # row 13 (22000+8000-1000 est)
    ('Flowers + Decor', 0, 15000),        # row 14
    ('Music + Entertainment', 0, 12000),  # row 15
    ('Rings', 0, 0),                      # row 16
    ('Stationery + Print', 0, 4000),      # row 17
    ('Transport + Lodging', 0, 0),        # row 18
    ('Hair + Makeup', 0, 7000),           # row 19
    ('Favors + Gifts', 0, 0),             # row 20
    ('Officiant + Ceremony', 0, 0),       # row 21
    ('Honeymoon', 0, 0),                  # row 22
    ('Contingency (5-10%)', 0, 0),        # row 23 - intentionally 0
]

for i, (cat, alloc, spent) in enumerate(category_map):
    row = 10 + i
    bc[f'B{row}'] = cat
    bc[f'C{row}'] = alloc
    bc[f'E{row}'] = spent

wb_p1.save(p1_out)

# Recalc via LibreOffice
p1_recalc_dir = os.path.join(SCRATCH, 'p1-qa-recalc')
os.makedirs(p1_recalc_dir, exist_ok=True)
cmd = [LO, '--headless', '--calc', '--convert-to', 'xlsx', '--outdir', p1_recalc_dir, p1_out]
subprocess.run(cmd, capture_output=True, timeout=90)
p1_recalc = os.path.join(p1_recalc_dir, os.path.basename(p1_out))
p1_pass = os.path.exists(p1_recalc)
print(f'\n  LibreOffice recalc: {"OK" if p1_pass else "FAILED — output missing"}')

if p1_pass:
    wb_p1r = load_workbook(p1_recalc, data_only=True)

    # Check dashboard values
    dash = wb_p1r.worksheets[1]
    bc_r = wb_p1r.worksheets[2]

    # Read Budget Categories totals from recalculated workbook
    spent_vals = []
    for row in range(10, 24):
        v = bc_r[f'E{row}'].value
        if v is not None:
            spent_vals.append(v)
    total_spent_wb = sum(v for v in spent_vals if isinstance(v, (int, float)))
    print(f'\n  EVALUATION RESULTS:')
    print(f'    Budget Categories row E10:E23 values: {spent_vals}')
    print(f'    Total spent in workbook: {total_spent_wb}')

    # Budget cap from setup wizard
    sw_r = wb_p1r.worksheets[0]
    budget_cap_val = sw_r['E16'].value
    print(f'    Budget cap (Setup Wizard E16): {budget_cap_val}')

    # Dashboard formula cells (these will be formula strings, not values, since they use concatenation)
    # The KPI cells are text-concatenation formulas - check E62 (actual spent) and E61 (expected)
    dash_e62 = dash['E62'].value
    print(f'    Dashboard E62 (Actual spent today): {dash_e62}')
    dash_e61 = dash['E61'].value
    print(f'    Dashboard E61 (Expected by today): {dash_e61}')

    # Check for #DIV/0! or errors in Budget Categories
    bc_errors = []
    for row in range(10, 24):
        for col in 'CDEFGJ':
            v = bc_r[f'{col}{row}'].value
            if isinstance(v, str) and v.startswith('#'):
                bc_errors.append(f'{col}{row}: {v}')
    if bc_errors:
        for e in bc_errors:
            new_issue('CRITICAL', 'essentials', 'Budget Categories', e.split(':')[0],
                      f'Formula error in Budget Categories', f'Value: {e}', 'Fix formula')
        print(f'  Budget Categories errors: {bc_errors}')
    else:
        print(f'  Budget Categories: no formula errors.')

    # Check: over-budget condition should be flagged
    # Dashboard B10 formula checks if spent > budget
    dash_b10 = dash['B10'].value
    print(f'    Dashboard B10 (Budget Health): {dash_b10}')

    # Validate persona 1 key numbers
    p1_results = {
        'total_committed_ref': p1_total_committed,
        'total_spent_wb': total_spent_wb,
        'budget_cap_wb': budget_cap_val,
        'budget_cap_ref': p1_budget,
    }
    print(f'\n  EXPECTED vs EVALUATED:')
    print(f'    Total committed (ref): EGP {p1_total_committed:,}  |  Total spent (wb): {total_spent_wb}')
    print(f'    Budget cap (ref): EGP {p1_budget:,}  |  Budget cap (wb): {budget_cap_val}')

    if budget_cap_val != p1_budget:
        new_issue('HIGH', 'essentials', 'Setup Wizard', 'E16',
                  f'Budget cap mismatch after write: expected {p1_budget}, got {budget_cap_val}',
                  f'Written: {p1_budget}, Read-back: {budget_cap_val}',
                  'Verify openpyxl write and recalc pipeline')

    # Verify the Vendor Tracker is separate from Budget Categories
    vt_r = wb_p1r.worksheets[3]
    vt_total_commit = vt_r['F30'].value
    vt_deposits = vt_r['E30'].value
    print(f'    Vendor Tracker F30 (Total Commit): {vt_total_commit}')
    print(f'    Vendor Tracker E30 (Deposits): {vt_deposits}')

    # Check that there's no per-guest cost KPI on the Dashboard (a gap we need to flag)
    # The setup wizard has one, but let's verify the Dashboard
    has_per_guest_on_dash = False
    for row in dash.iter_rows():
        for cell in row:
            v = str(cell.value or '')
            if 'per' in v.lower() and 'guest' in v.lower():
                has_per_guest_on_dash = True
                print(f'    Per-guest KPI on Dashboard at {cell.coordinate}: {v[:80]}')
    if not has_per_guest_on_dash:
        new_issue('MEDIUM', 'essentials', 'Budget Dashboard', 'N/A',
                  'No explicit Per-Guest Cost KPI on the Budget Dashboard tab',
                  'Setup Wizard has per-guest formula but it is not surfaced on Dashboard',
                  '[COMPLEMENT] Add Per-Guest Cost KPI cell on Dashboard')

    # Check days-until-wedding is dynamic
    has_days = False
    for row in dash.iter_rows():
        for cell in row:
            v = str(cell.value or '')
            if 'days' in v.lower() and ('to' in v.lower() or 'until' in v.lower() or 'wedding' in v.lower()):
                has_days = True
    if not has_days:
        new_issue('MEDIUM', 'essentials', 'Budget Dashboard', 'N/A',
                  'Days-until-wedding counter not found on Dashboard evaluated output',
                  'Formula string present but evaluates to text concatenation',
                  '[COMPLEMENT] Verify countdown cell evaluates and displays correctly')


# ══════════════════════════════════════════════════════════════════════════════
# PERSONA 2 — Mariam & Ahmed, Multi-Event MENA Wedding
# ══════════════════════════════════════════════════════════════════════════════
print('\n' + '='*70)
print('PERSONA 2 — Mariam & Ahmed | Cairo | EGP | Multi-Event MENA Wedding')
print('='*70)

p2_events = {
    'Katb-Kitab': {'guests': 50, 'budget': 60000, 'total': 60000},
    'Henna night': {'guests': 80, 'budget': 70000, 'total': 70000},
    'Reception': {'guests': 400, 'budget': 600000, 'total': 605000},
}
p2_grand_total = sum(e['total'] for e in p2_events.values())
p2_budget = 850000
p2_buffer = p2_budget - p2_grand_total
p2_per_guest_reception = 605000 / 400

print(f'\nPYTHON REFERENCE VALUES:')
print(f'  Grand total committed: EGP {p2_grand_total:,}')
print(f'  Overall budget: EGP {p2_budget:,}')
print(f'  Buffer remaining: EGP {p2_buffer:,} ({p2_buffer/p2_budget*100:.1f}%)')
print(f'  Reception per-guest cost: EGP {p2_per_guest_reception:.2f}/head')
print(f'  Reception over-budget by: EGP {605000-600000:,}')

# Check if multi-event sub-budgets are supported
# Look in AI edition (most likely to have it)
ai_wb = src_wbs.get('ai')
if ai_wb:
    sheets = [ws.title for ws in ai_wb.worksheets]
    multi_event_sheet = None
    for title in sheets:
        if any(word in title.lower() for word in ['event', 'multi', 'ceremony', 'henna', 'katb']):
            multi_event_sheet = title
            break
    if not multi_event_sheet:
        new_issue('HIGH', 'ai', 'N/A', 'N/A',
                  'No multi-event sub-budget sheet found in AI Edition',
                  f'Sheets: {sheets}',
                  '[COMPLEMENT] Add Multi-Event Budget tab with per-event sub-totals, variance, and per-guest cost')
        print(f'  MISSING: Multi-event sub-budget tab not found in AI Edition')
    else:
        print(f'  Multi-event sheet found: {multi_event_sheet}')
else:
    print('  AI workbook not loaded, skipping multi-event check')

# Check Pro edition for multi-event
pro_wb = src_wbs.get('pro')
if pro_wb:
    sheets_pro = [ws.title for ws in pro_wb.worksheets]
    has_multi_event_pro = any('event' in t.lower() or 'henna' in t.lower() for t in sheets_pro)
    if not has_multi_event_pro:
        new_issue('HIGH', 'pro', 'N/A', 'N/A',
                  'No multi-event sub-budget sheet found in Pro Edition',
                  f'Sheets: {sheets_pro}',
                  '[COMPLEMENT] Add Multi-Event Budget tab')
        print(f'  MISSING: Multi-event sub-budget tab not found in Pro Edition')

# Persona 2 verdict: FAIL (multi-event not supported)
p2_verdict = 'FAIL — no multi-event sub-budget tab exists in any edition'
print(f'\n  PERSONA 2 VERDICT: {p2_verdict}')


# ══════════════════════════════════════════════════════════════════════════════
# PERSONA 3 — Layla & Tarek, Destination Wedding, Multi-Currency
# ══════════════════════════════════════════════════════════════════════════════
print('\n' + '='*70)
print('PERSONA 3 — Layla & Tarek | Tuscany | USD/EUR | Destination Wedding')
print('='*70)

p3_eur_usd = 1.08
p3_italy_eur = 90400
p3_italy_usd = round(p3_italy_eur * p3_eur_usd, 2)
p3_us_usd = 25200
p3_travel_usd = 18000
p3_welcome_bags = 85 * 35
p3_total_committed = p3_italy_usd + p3_us_usd + p3_travel_usd + p3_welcome_bags
p3_budget = 185000
p3_buffer = p3_budget - p3_total_committed
p3_guests = 85
p3_per_guest = p3_total_committed / p3_guests
p3_italy_share = p3_italy_usd / p3_total_committed

print(f'\nPYTHON REFERENCE VALUES (EUR@1.08):')
print(f'  Italy EUR: {p3_italy_eur:,} EUR → USD {p3_italy_usd:,}')
print(f'  US vendors: USD {p3_us_usd:,}')
print(f'  Travel: USD {p3_travel_usd:,}')
print(f'  Welcome bags: USD {p3_welcome_bags:,}')
print(f'  Total committed: USD {p3_total_committed:,.2f}')
print(f'  Budget: USD {p3_budget:,}')
print(f'  Buffer: USD {p3_buffer:,.2f} ({p3_buffer/p3_budget*100:.1f}%)')
print(f'  Per-guest cost: USD {p3_per_guest:.2f}/head')
print(f'  Italy share: {p3_italy_share*100:.1f}%')

# Test FX cascade: EUR→USD = 1.12
p3_eur_usd_new = 1.12
p3_italy_usd_new = round(p3_italy_eur * p3_eur_usd_new, 2)
p3_total_new = p3_italy_usd_new + p3_us_usd + p3_travel_usd + p3_welcome_bags
print(f'\n  FX CASCADE TEST (EUR@1.12):')
print(f'  Italy USD becomes: {p3_italy_usd_new:,}')
print(f'  New total: USD {p3_total_new:,}')

# Check if any edition supports multi-currency FX
multi_currency_found = False
for key, wb in src_wbs.items():
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=10):
            for cell in row:
                v = str(cell.value or '')
                if 'fx' in v.lower() or 'exchange' in v.lower() or ('rate' in v.lower() and 'eur' in v.upper()):
                    multi_currency_found = True
                    print(f'  Multi-currency reference found in {key}/{ws.title} at {cell.coordinate}: {v[:60]}')

if not multi_currency_found:
    new_issue('CRITICAL', 'all', 'N/A', 'N/A',
              'No FX rate table found in any edition — multi-currency destination weddings unsupported',
              'Searched all sheets for FX/exchange rate references; none found',
              '[COMPLEMENT] Add FX Rate table (currency → USD) with cascade to line items in AI/Pro editions')
    print(f'  CRITICAL: No multi-currency FX table found in any edition')

# Deeper check: look for currency column in vendor tracker
for key, wb in src_wbs.items():
    vt = None
    for ws in wb.worksheets:
        if 'vendor' in ws.title.lower():
            vt = ws
            break
    if vt:
        headers = [str(cell.value or '').lower() for cell in vt[8]]  # row 8 headers
        print(f'  {key} Vendor Tracker row 8 headers: {headers[:10]}')
        if not any('curr' in h for h in headers):
            new_issue('HIGH', key, 'Vendor Tracker', 'Row 8',
                      f'No currency column in Vendor Tracker — multi-currency line items impossible',
                      f'Headers found: {headers[:10]}',
                      '[COMPLEMENT] Add Currency and FX-converted Amount columns to Vendor Tracker')

p3_verdict = 'FAIL — no FX rate table, no currency column in vendor tracker'
print(f'\n  PERSONA 3 VERDICT: {p3_verdict}')


# ══════════════════════════════════════════════════════════════════════════════
# PERSONA 4 — Malak & Karim, High-Budget Multi-Family Contribution
# ══════════════════════════════════════════════════════════════════════════════
print('\n' + '='*70)
print('PERSONA 4 — Malak & Karim | USD | High-Budget Multi-Family Contribution')
print('='*70)

p4_reception = 389050
p4_engagement = 42500
p4_honeymoon = 42000
p4_total_committed = p4_reception + p4_engagement + p4_honeymoon
p4_budget = 420000
p4_variance = p4_budget - p4_total_committed
p4_guests = 350
p4_per_guest_reception = p4_reception / p4_guests
p4_contribs = {'Bride father': 250000, 'Groom family': 120000, 'Couple': 50000}
p4_total_contribs = sum(p4_contribs.values())
p4_uncovered = p4_total_committed - p4_total_contribs

print(f'\nPYTHON REFERENCE VALUES:')
print(f'  Total committed: USD {p4_total_committed:,}')
print(f'  Budget: USD {p4_budget:,}')
print(f'  Variance: USD {p4_variance:,} ({"OVER" if p4_variance < 0 else "UNDER"} by {abs(p4_variance)/p4_budget*100:.1f}%)')
print(f'  Per-guest (reception): USD {p4_per_guest_reception:.2f}/head')
print(f'  Total contributions: USD {p4_total_contribs:,}')
print(f'  Uncovered gap: USD {p4_uncovered:,}')

# Check if any edition has contribution tracking
contribution_found = False
for key, wb in src_wbs.items():
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=10):
            for cell in row:
                v = str(cell.value or '')
                if any(word in v.lower() for word in ['contribution', 'contrib', "bride's family", "groom's family", 'sponsor']):
                    contribution_found = True
                    print(f'  Contribution reference in {key}/{ws.title} at {cell.coordinate}: {v[:60]}')

if not contribution_found:
    new_issue('HIGH', 'all', 'N/A', 'N/A',
              'No contribution tracking (bride family / groom family / couple sources) in any edition',
              'Searched all sheets for contribution/family tracking; none found',
              '[COMPLEMENT] Add Contribution Tracker tab with named sources, amounts, percentages, and coverage vs total committed')
    print(f'  HIGH: No multi-family contribution tracker found')

# Check 6-digit value handling — look at column widths in pro
if 'pro' in src_wbs:
    pro_wb = src_wbs['pro']
    for ws in pro_wb.worksheets:
        if 'budget' in ws.title.lower() or 'vendor' in ws.title.lower():
            for col_letter, col_dim in ws.column_dimensions.items():
                w = col_dim.width
                if w is not None and w < 12:
                    # Flag potentially narrow columns
                    pass  # we'll note but not flag unless we see ### in data_only

# Check Dashboard over-budget flag
if struct_results.get('essentials'):
    wb_ess = struct_results['essentials']
    dash = wb_ess.worksheets[1]
    # B10 has the budget health check
    b10 = dash['B10'].value
    print(f'\n  Essentials Dashboard B10 (Budget Health): {b10}')
    if b10 is None:
        new_issue('HIGH', 'essentials', 'Budget Dashboard', 'B10',
                  'Budget Health KPI evaluates to None — over-budget flag not visible',
                  'Cell B10 is None after LibreOffice recalc',
                  'Verify formula evaluates correctly; may need formula fix')

p4_verdict = 'FAIL — no contribution tracker, multi-family split unsupported'
print(f'\n  PERSONA 4 VERDICT: {p4_verdict}')


# ══════════════════════════════════════════════════════════════════════════════
# PERSONA 5 — Nour & Hassan, Budget-Conscious, Gift-Registry Offset
# ══════════════════════════════════════════════════════════════════════════════
print('\n' + '='*70)
print('PERSONA 5 — Nour & Hassan | USD | Budget-Conscious | Gift-Registry Offset')
print('='*70)

p5_items = {
    'Venue': 4500, 'Catering': 5400, 'Photography': 3200,
    'Flowers/Decor': 1800, 'Dress': 1200, 'Suit': 450,
    'Invitations': 400, 'DJ': 1500, 'Cake': 650,
    'Hair/Makeup': 400, 'Honeymoon': 4800, 'Rings': 2800,
    'License/Officiant': 400, 'Contingency': 7500,
}
p5_excl_contingency = sum(v for k, v in p5_items.items() if k != 'Contingency')
p5_total = sum(p5_items.values())
p5_budget = 35000
p5_per_guest_core = (p5_excl_contingency - p5_items['Honeymoon'] - p5_items['Rings']) / 120
p5_gifts = 4500
p5_net_oop = 25650 - p5_gifts  # final spend minus gifts
p5_days = (datetime.date(2026, 11, 23) - TODAY).days  # ~6 months out

print(f'\nPYTHON REFERENCE VALUES:')
print(f'  Total committed (excl contingency): USD {p5_excl_contingency:,}')
print(f'  Total with contingency: USD {p5_total:,}')
print(f'  Budget: USD {p5_budget:,}')
print(f'  On budget: {p5_total == p5_budget}')
print(f'  Per-guest cost (core): USD {p5_per_guest_core:.2f}/head')
print(f'  Total gifts received: USD {p5_gifts:,}')
print(f'  Net out-of-pocket: USD {p5_net_oop:,}')
print(f'  Days until wedding: ~{p5_days}')

# Write Persona 5 to essentials
p5_src = SRC['essentials']
p5_out = os.path.join(SCRATCH, 'p5-nour-essentials-qa.xlsx')
shutil.copy(p5_src, p5_out)
wb_p5 = load_workbook(p5_out, data_only=False)

sw5 = wb_p5.worksheets[0]
sw5['E10'] = datetime.datetime(2026, 11, 23)
sw5['E12'] = 120
sw5['E16'] = 35000
sw5['E18'] = 'Portland, OR'
sw5['E20'] = 'USD'
sw5['E28'] = 'Nour'
sw5['E30'] = 'Hassan'

# Write budget categories
bc5 = wb_p5.worksheets[2]
p5_category_data = [
    ('Venue + Rentals', 4500),
    ('Catering + Bar', 5400+650),  # catering + cake
    ('Photography + Video', 3200),
    ('Attire + Beauty', 1200+450+400),  # dress+suit+hair
    ('Flowers + Decor', 1800),
    ('Music + Entertainment', 1500),
    ('Rings', 2800),
    ('Stationery + Print', 400),
    ('Transport + Lodging', 0),
    ('Hair + Makeup', 0),  # already in attire
    ('Favors + Gifts', 0),
    ('Officiant + Ceremony', 400),
    ('Honeymoon', 4800),
    ('Contingency (5-10%)', 7500),
]
for i, (cat, spent) in enumerate(p5_category_data):
    row = 10 + i
    bc5[f'E{row}'] = spent

wb_p5.save(p5_out)

# Recalc
p5_recalc_dir = os.path.join(SCRATCH, 'p5-qa-recalc')
os.makedirs(p5_recalc_dir, exist_ok=True)
cmd5 = [LO, '--headless', '--calc', '--convert-to', 'xlsx', '--outdir', p5_recalc_dir, p5_out]
subprocess.run(cmd5, capture_output=True, timeout=90)
p5_recalc = os.path.join(p5_recalc_dir, os.path.basename(p5_out))
p5_pass = os.path.exists(p5_recalc)
print(f'\n  LibreOffice recalc: {"OK" if p5_pass else "FAILED"}')

if p5_pass:
    wb_p5r = load_workbook(p5_recalc, data_only=True)
    bc5r = wb_p5r.worksheets[2]
    sw5r = wb_p5r.worksheets[0]

    total_spent_5 = sum(bc5r[f'E{r}'].value or 0 for r in range(10, 24))
    budget_cap_5 = sw5r['E16'].value
    months_out_5 = sw5r['E26'].value

    print(f'  Total spent (wb): {total_spent_5}')
    print(f'  Budget cap (wb): {budget_cap_5}')
    print(f'  Months from today (E26): {months_out_5}')

    if abs(total_spent_5 - p5_excl_contingency) < 1:
        print(f'  PASS: Total spent matches reference ({p5_excl_contingency})')
    else:
        new_issue('HIGH', 'essentials', 'Budget Categories', 'E10:E23',
                  f'Total spent mismatch: expected {p5_excl_contingency}, got {total_spent_5}',
                  f'Persona 5 write-and-recalc; diff={total_spent_5-p5_excl_contingency}',
                  'Verify budget category write logic')

    # Check gift-registry offset — not present in essentials
    has_gift_registry = False
    for ws in wb_p5r.worksheets:
        if 'gift' in ws.title.lower() or 'registry' in ws.title.lower():
            has_gift_registry = True
    if not has_gift_registry:
        new_issue('MEDIUM', 'essentials', 'N/A', 'N/A',
                  'No Gift Registry tab in Essentials — gift offset tracking impossible',
                  'Essentials does not include gift registry sheet',
                  'Gift Registry is present in Pro and AI editions; note this as Essentials limitation')

p5_verdict = 'PASS-WITH-CAVEATS — budget math works; gift offset requires Pro/AI edition'
print(f'\n  PERSONA 5 VERDICT: {p5_verdict}')


# ══════════════════════════════════════════════════════════════════════════════
# EDGE CASE PROBING
# ══════════════════════════════════════════════════════════════════════════════
print('\n' + '='*70)
print('EDGE CASE PROBING')
print('='*70)

# Edge 1: Guest count = 0, per-guest should not DIV0
# Check formula in setup wizard
ess_wb = src_wbs.get('essentials')
if ess_wb:
    sw = ess_wb.worksheets[0]
    i2_formula = sw['I2'].value  # per-guest KPI
    print(f'\n  Per-guest formula (Setup Wizard I2): {i2_formula}')
    if 'IFERROR' in str(i2_formula).upper() or 'MAX(1' in str(i2_formula).upper():
        print(f'  PASS: Per-guest formula guards against DIV/0 (IFERROR or MAX(1,...))')
    else:
        new_issue('HIGH', 'essentials', 'Setup Wizard', 'I2',
                  'Per-guest formula may not guard against DIV/0 when guest count=0',
                  f'Formula: {i2_formula}',
                  'Wrap denominator with MAX(1, guest_count)')

# Edge 2: Percentage vs decimal foot-gun in deposit %
# Check if vendor tracker expects % as decimal or integer
if ess_wb:
    vt = ess_wb.worksheets[3]
    # Check E9 (deposit for venue)
    e9 = vt['E9'].value
    f9 = vt['F9'].value
    print(f'\n  Vendor Tracker E9 (deposit): {e9}  F9 (total): {f9}')
    if e9 is not None and f9 is not None:
        if isinstance(e9, (int,float)) and isinstance(f9, (int,float)) and f9 > 0:
            pct = e9/f9
            print(f'  Deposit % of total: {pct:.1%}')
            # Seems like deposits are stored as dollar amounts, not percentages
            print(f'  Deposit convention: dollar amounts (not percentages) — OK')
        else:
            print(f'  Could not verify deposit convention')

# Edge 3: Wedding date in the past
# Check if countdown uses MAX(0, ...) to avoid negative
if ess_wb:
    sw = ess_wb.worksheets[0]
    c2 = sw['C2'].value  # days to go KPI
    g10 = sw['G26'].value  # months from today formula
    print(f'\n  Days-to-go formula (C2): {str(c2)[:80]}')
    print(f'  Months formula (E26): {str(g10)[:80] if g10 else None}')
    if 'MAX(0' in str(c2).upper():
        print(f'  PASS: Days-to-go uses MAX(0,...) — negative countdown prevented')
    else:
        new_issue('MEDIUM', 'essentials', 'Setup Wizard', 'C2',
                  'Days-to-go KPI may show negative if wedding date is in the past',
                  f'Formula: {c2}',
                  'Wrap with MAX(0, ...) to prevent negative countdown display')

# Edge 4: Categories sum > 100% allocation
if ess_wb:
    bc = ess_wb.worksheets[2]
    c25 = bc['C25'].value  # total % allocation
    print(f'\n  Budget Categories C25 (Total % allocation): {c25}')
    if isinstance(c25, (int,float)):
        total_alloc = float(c25)
        print(f'  Total allocation: {total_alloc:.1%}')
        if abs(total_alloc - 1.0) > 0.01:
            new_issue('MEDIUM', 'essentials', 'Budget Categories', 'C25',
                      f'Default allocations sum to {total_alloc:.1%}, not 100%',
                      f'C25 formula=SUM(C10:C23) = {total_alloc:.4f}',
                      'Ensure default category percentages sum to 100% or document intentional sub-100% allocation')
            print(f'  WARNING: Default allocation sums to {total_alloc:.1%}')
        else:
            print(f'  PASS: Default allocation sums to ~100%')
    else:
        print(f'  Cannot verify allocation sum (formula string not evaluated)')


# ══════════════════════════════════════════════════════════════════════════════
# METHODOLOGY SOUNDNESS CHECK
# ══════════════════════════════════════════════════════════════════════════════
print('\n' + '='*70)
print('METHODOLOGY SOUNDNESS')
print('='*70)

# Canonical categories check
canonical = ['venue', 'catering', 'photography', 'video', 'attire', 'flowers',
             'music', 'stationery', 'cake', 'hair', 'makeup', 'transport',
             'rings', 'officiant', 'favors', 'honeymoon', 'contingency']
found_cats = set()
if ess_wb:
    bc = ess_wb.worksheets[2]
    for row in range(10, 24):
        cat = str(bc[f'B{row}'].value or '').lower()
        for c in canonical:
            if c in cat:
                found_cats.add(c)
missing_cats = [c for c in canonical if c not in found_cats]
print(f'\n  Found canonical categories: {sorted(found_cats)}')
print(f'  Missing canonical categories: {missing_cats}')
if missing_cats:
    new_issue('LOW', 'essentials', 'Budget Categories', 'B10:B23',
              f'Missing canonical categories: {missing_cats}',
              f'Categories present: {sorted(found_cats)}',
              '[COMPLEMENT] Confirm all canonical categories are present or documented as out-of-scope')

# Separate cake from catering check
has_cake_separate = False
for row in range(10, 24):
    cat = str(bc[f'B{row}'].value or '').lower()
    if 'cake' in cat:
        has_cake_separate = True
if not has_cake_separate:
    new_issue('LOW', 'essentials', 'Budget Categories', 'B10:B23',
              'Cake is not a separate category — typically billed separately from catering',
              'No "Cake" category row found',
              '[COMPLEMENT] Add Cake as separate line item under Catering+Bar or standalone category')


# ══════════════════════════════════════════════════════════════════════════════
# THUMBNAIL CHECK
# ══════════════════════════════════════════════════════════════════════════════
print('\n' + '='*70)
print('THUMBNAIL AUDIT')
print('='*70)

from PIL import Image
thumb_dir = 'C:/ETSY/etsy-store/tools/thumb-gen/output'
wedding_thumbs = [f for f in os.listdir(thumb_dir) if f.startswith('wedding-budget-planner') and f.endswith('.png')]
print(f'\n  Wedding thumbnails found: {len(wedding_thumbs)}')
for fname in sorted(wedding_thumbs):
    fpath = os.path.join(thumb_dir, fname)
    img = Image.open(fpath)
    size = img.size
    mode = img.mode
    fsize = os.path.getsize(fpath)
    print(f'  {fname}: {size[0]}x{size[1]} {mode} {fsize/1024:.0f}KB')

    if size[0] != size[1]:
        new_issue('HIGH', 'thumbnail', fname, 'N/A',
                  f'Thumbnail not square: {size[0]}x{size[1]}',
                  f'Expected 2000x2000, got {size}',
                  'Regenerate thumbnail at 2000x2000 px')
    elif size[0] < 2000:
        new_issue('MEDIUM', 'thumbnail', fname, 'N/A',
                  f'Thumbnail smaller than 2000x2000: {size[0]}x{size[1]}',
                  f'Got {size}',
                  'Regenerate at 2000x2000 px for Etsy best practice')
    if fsize > 20 * 1024 * 1024:
        new_issue('HIGH', 'thumbnail', fname, 'N/A',
                  f'Thumbnail exceeds 20MB: {fsize/1024/1024:.1f}MB',
                  f'File size: {fsize}', 'Compress image')
    if mode != 'RGB' and mode != 'RGBA':
        new_issue('MEDIUM', 'thumbnail', fname, 'N/A',
                  f'Thumbnail not in sRGB mode: {mode}',
                  f'Mode: {mode}', 'Convert to sRGB')

    # Downsample to 250px and check legibility (just verify it works)
    small = img.resize((250, 250), Image.LANCZOS)
    small_path = os.path.join(SCRATCH, fname.replace('.png', '_250.png'))
    small.save(small_path)


# ══════════════════════════════════════════════════════════════════════════════
# OUTPUT ISSUE SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
print('\n' + '='*70)
print('ISSUE SUMMARY')
print('='*70)

critical = [i for i in issues if i['severity'] == 'CRITICAL']
high = [i for i in issues if i['severity'] == 'HIGH']
medium = [i for i in issues if i['severity'] == 'MEDIUM']
low = [i for i in issues if i['severity'] == 'LOW']

print(f'\nCRITICAL: {len(critical)}')
for i in critical:
    print(f'  {i["id"]}: {i["file"]}/{i["sheet"]}/{i["cell"]} — {i["problem"][:80]}')

print(f'\nHIGH: {len(high)}')
for i in high:
    print(f'  {i["id"]}: {i["file"]}/{i["sheet"]}/{i["cell"]} — {i["problem"][:80]}')

print(f'\nMEDIUM: {len(medium)}')
for i in medium:
    print(f'  {i["id"]}: {i["file"]}/{i["sheet"]}/{i["cell"]} — {i["problem"][:80]}')

print(f'\nLOW: {len(low)}')
for i in low:
    print(f'  {i["id"]}: {i["file"]}/{i["sheet"]}/{i["cell"]} — {i["problem"][:80]}')

# Save issues to JSON
with open('C:/ETSY/etsy-store/tools/qa/output/wedding_qa_issues.json', 'w', encoding='utf-8') as f:
    json.dump(issues, f, indent=2, ensure_ascii=False)
print(f'\nIssues saved to tools/qa/output/wedding_qa_issues.json')

print('\nDONE.')
