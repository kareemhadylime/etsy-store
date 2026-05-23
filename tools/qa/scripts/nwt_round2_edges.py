"""Round 2 edge-case re-tests."""
import json, os, shutil, subprocess
from openpyxl import load_workbook

FIXED = "C:/ETSY/etsy-store/tools/qa/fixed/net-worth-tracker-ai-edition.xlsx"
WORK = "C:/ETSY/etsy-store/tools/qa/scratch/nwt/round2-edges"
LO = "C:/Program Files/LibreOffice/program/soffice.exe"
os.makedirs(WORK, exist_ok=True)


def recalc(p):
    out = os.path.join(WORK, "recalc")
    os.makedirs(out, exist_ok=True)
    subprocess.run([LO, "--headless", "--calc", "--convert-to", "xlsx", "--outdir", out, p],
                   capture_output=True, text=True, timeout=60)
    return os.path.join(out, os.path.basename(p))


def case(name, setup_fn):
    p = os.path.join(WORK, f"{name}.xlsx")
    shutil.copy(FIXED, p)
    wb = load_workbook(p)
    setup_fn(wb)
    wb.save(p)
    new = recalc(p)
    wbc = load_workbook(new, data_only=True)
    ws_a = wbc["💼 Assets Summary"]
    ws_l = wbc["📉 Liabilities Summary"]
    ws_d = wbc["🏠 Dashboard"]
    return {
        "case": name,
        "total_assets": ws_a["N26"].value,
        "total_liabs": ws_l["N21"].value,
        "dashboard_a2_nw": ws_d["A2"].value,
        "dashboard_g2_debt_to_asset": ws_d["G2"].value,
    }


def s_zero_assets(wb):
    ws_a = wb["💼 Assets Summary"]
    for r in range(9, 25):
        ws_a.cell(r, 14).value = None
    ws_l = wb["📉 Liabilities Summary"]
    ws_l.cell(11, 14).value = 5000


def s_huge(wb):
    ws_a = wb["💼 Assets Summary"]
    for r in range(9, 25):
        ws_a.cell(r, 14).value = None
    ws_a.cell(14, 14).value = 12500000
    ws_a.cell(15, 14).value = 95000000
    ws_a.cell(17, 14).value = 8500000
    ws_a.cell(16, 14).value = 17000000


def s_single(wb):
    for r in range(9, 25):
        wb["💼 Assets Summary"].cell(r, 14).value = None
    for r in range(9, 20):
        wb["📉 Liabilities Summary"].cell(r, 14).value = None
    wb["💼 Assets Summary"].cell(9, 14).value = 50000


cases = [("zero_assets", s_zero_assets), ("huge_values", s_huge), ("single_asset", s_single)]
results = []
for name, fn in cases:
    r = case(name, fn)
    print(json.dumps(r, indent=2, default=str), flush=True)
    results.append(r)

with open("C:/ETSY/etsy-store/tools/qa/round2/edge_results.json", "w", encoding="utf-8") as fh:
    json.dump(results, fh, indent=2, default=str)
