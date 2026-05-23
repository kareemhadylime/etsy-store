"""NWT Stage A - structural integrity inspection of all three tiers.

Reads recalculated copies from tools/qa/scratch/nwt/ with data_only=True
and reports:
  - sheet list + visibility per file
  - error cells (#REF, #N/A, #DIV/0, #NAME?, #VALUE!)
  - cells where formula evaluated to None (formula likely broken or empty path)
  - defined names
  - data validations
  - first-row + summary cells per sheet (small dump for visual review)
"""
import json
import os
from collections import defaultdict
from openpyxl import load_workbook

SCRATCH = "C:/ETSY/etsy-store/tools/qa/scratch/nwt"
OUT = "C:/ETSY/etsy-store/tools/qa/round1"
os.makedirs(OUT, exist_ok=True)

FILES = {
    "essentials": "net-worth-tracker-essentials.xlsx",
    "pro": "net-worth-tracker-pro.xlsx",
    "ai": "net-worth-tracker-ai-edition.xlsx",
}

ERR_SENTINELS = ("#REF!", "#N/A", "#DIV/0!", "#NAME?", "#VALUE!", "#NUM!", "#NULL!")


def inspect(path):
    out = {}
    wb_calc = load_workbook(path, data_only=True)
    wb_form = load_workbook(path, data_only=False)
    out["sheet_names"] = wb_calc.sheetnames
    out["visibility"] = {s: wb_calc[s].sheet_state for s in wb_calc.sheetnames}
    out["defined_names"] = sorted(list(wb_calc.defined_names.keys())) if hasattr(wb_calc.defined_names, "keys") else [n.name for n in wb_calc.defined_names.definedName]
    sheet_summary = {}
    error_cells = []
    none_eval_cells = []
    formula_count_per_sheet = {}
    for sn in wb_calc.sheetnames:
        ws_c = wb_calc[sn]
        ws_f = wb_form[sn]
        fcount = 0
        errs_this = 0
        nones_this = 0
        max_row = ws_c.max_row or 0
        max_col = ws_c.max_column or 0
        for row in ws_f.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            for cell in row:
                v_form = cell.value
                if isinstance(v_form, str) and v_form.startswith("="):
                    fcount += 1
                    v_calc = ws_c.cell(row=cell.row, column=cell.column).value
                    if isinstance(v_calc, str) and any(e in v_calc for e in ERR_SENTINELS):
                        error_cells.append({"sheet": sn, "cell": cell.coordinate, "formula": v_form[:200], "value": v_calc})
                        errs_this += 1
                    elif v_calc is None:
                        none_eval_cells.append({"sheet": sn, "cell": cell.coordinate, "formula": v_form[:200]})
                        nones_this += 1
        formula_count_per_sheet[sn] = {"formulas": fcount, "errors": errs_this, "nones": nones_this, "dimensions": f"{max_row}x{max_col}"}
        # tiny dump: first 8 non-empty cells
        sample = []
        for r in ws_c.iter_rows(min_row=1, max_row=min(max_row, 30), max_col=min(max_col, 20)):
            for c in r:
                if c.value is not None and len(sample) < 12:
                    sample.append({"cell": c.coordinate, "value": str(c.value)[:80]})
        sheet_summary[sn] = sample
    out["formula_stats"] = formula_count_per_sheet
    out["error_cells"] = error_cells
    out["none_eval_cells_sample"] = none_eval_cells[:80]
    out["none_eval_cells_total"] = len(none_eval_cells)
    out["sheet_first_cells"] = sheet_summary
    out["data_validations"] = {}
    for sn in wb_form.sheetnames:
        wsf = wb_form[sn]
        dvs = []
        try:
            for dv in wsf.data_validations.dataValidation:
                dvs.append({"type": dv.type, "formula1": (dv.formula1 or "")[:200], "ranges": str(dv.sqref) if dv.sqref else ""})
        except Exception as e:
            dvs.append({"error": str(e)})
        if dvs:
            out["data_validations"][sn] = dvs
    return out


report = {}
for tier, fn in FILES.items():
    path = os.path.join(SCRATCH, fn)
    print(f"--- {tier}: {path} ---")
    r = inspect(path)
    report[tier] = r
    print(f"  sheets: {len(r['sheet_names'])}")
    print(f"  visible: {sum(1 for v in r['visibility'].values() if v == 'visible')}")
    print(f"  errors: {len(r['error_cells'])}")
    print(f"  none-eval: {r['none_eval_cells_total']}")

with open(os.path.join(OUT, "stage_a_structural.json"), "w", encoding="utf-8") as f:
    json.dump(report, f, indent=2, default=str)

print("\nWROTE:", os.path.join(OUT, "stage_a_structural.json"))
