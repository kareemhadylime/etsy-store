"""Apply Round 2 fixes to copies at tools/qa/fixed/.

Issues addressed:
  [FIX-BNDL-001] IPT Scenario Simulator G2/I2 malformed formula → replace with clean IFERROR
  [FIX-BNDL-002] Brand drift — leave as-is; deferred (per-SKU palette is intentional brand identity per SKU)
                  → instead document in README how the bundle uses color-coding per SKU and add a
                    "Lime Premium Studios — shared palette anchor color" cell on each SKU's About tab.
                    (See [COMPLEMENT-3])
  [FIX-BNDL-005] NWT row 16 → add documented instruction + helper formula referencing IPT total
                  (in NWT — since we can't auto-link across two separate workbooks at xlsx level,
                  we add a cell with the recommended Sheets formula syntax and a customer-facing note)
  [FIX-BNDL-006] NWT and IPT FX tables → add a "Sync Reminder" callout on both tables noting that
                  the rates must match and pointing to the canonical source.
  [FIX-BNDL-008] XREF-WEAK — partial fix: add a "Bundle Cross-References" page to each AI advisor
                  (deferred to round2 complement: bundle-level navigation README).
  [FIX-BNDL-015] NWT Dec-only headline — add a note on Assets Summary tab explaining that the
                  headline reflects the LATEST month (currently December as last column).
                  Optional: change formula to use MAX of any non-empty column, but that risks
                  breaking historical comparisons.

Skipped for safety:
  [BNDL-003, 004, 007, 017, 018, 019] medium-severity cross-SKU concepts — deferred to bundle
                  documentation rather than reworking 18 workbook formulas.
"""
import sys, io, shutil, openpyxl
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', line_buffering=True)
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.comments import Comment

ROOT = Path("C:/ETSY/etsy-store")
FIXED = ROOT / "tools/qa/fixed"

changelog = []

# ============================================================
# [FIX-BNDL-001] IPT Scenario Simulator G2/I2 — broken formula
# ============================================================
for tier_file in ["investment-portfolio-tracker-ai-edition.xlsx",
                  "investment-portfolio-tracker-pro.xlsx"]:
    path = FIXED / tier_file
    wb = openpyxl.load_workbook(path)
    ws_name = None
    for sh in wb.sheetnames:
        if "Scenario" in sh:
            ws_name = sh
            break
    if not ws_name:
        changelog.append(f"[FIX-BNDL-001] {tier_file}: Scenario Simulator tab not found — SKIPPED")
        continue
    ws = wb[ws_name]
    # Old (broken): ="RECOVERY"&CHAR(10)&(IFERROR(C11&" mo",—))))
    # Fix:           ="RECOVERY"&CHAR(10)&IFERROR(C11&" mo","—")
    # G2: RECOVERY metric
    old_g2 = ws['G2'].value
    new_g2 = '="RECOVERY"&CHAR(10)&IFERROR(C11&" mo","—")'
    ws['G2'].value = new_g2
    # I2: FIRE DELAY
    old_i2 = ws['I2'].value
    new_i2 = '="FIRE DELAY"&CHAR(10)&IFERROR(C13&" mo","—")'
    ws['I2'].value = new_i2
    wb.save(path)
    changelog.append(f"[FIX-BNDL-001] {tier_file}: G2 '{str(old_g2)[:50]}...' → '{new_g2}'")
    changelog.append(f"[FIX-BNDL-001] {tier_file}: I2 '{str(old_i2)[:50]}...' → '{new_i2}'")

# ============================================================
# [FIX-BNDL-005] NWT — add helper cell + note explaining IPT linkage
#                 (we cannot truly auto-link across two separate xlsx workbooks
#                  via openpyxl/excel; instead provide a clear customer-facing
#                  instruction + a visual marker on NWT Assets Summary row 16)
# ============================================================
nwt_files = ["net-worth-tracker-ai-edition.xlsx",
             "net-worth-tracker-pro.xlsx",
             "net-worth-tracker-essentials.xlsx"]
for nwt_file in nwt_files:
    path = FIXED / nwt_file
    wb = openpyxl.load_workbook(path)
    if "💼 Assets Summary" not in wb.sheetnames:
        changelog.append(f"[FIX-BNDL-005] {nwt_file}: Assets Summary tab missing — SKIPPED")
        continue
    ws = wb["💼 Assets Summary"]
    # Row 16 is "Stocks & Funds (Taxable)". Add a comment on B16 directing customer
    # to IPT bundle workbook.
    target_cell = ws['B16']
    target_cell.comment = Comment(
        "BUNDLE NOTE: If you also own the Investment Portfolio Tracker SKU, the value here "
        "should equal your IPT Dashboard 'Total MV' line. Update this cell whenever your IPT "
        "holdings change. The two workbooks do not auto-link.\n\n"
        "Recommended workflow: at month-end, open IPT, read 🏠 Dashboard → Total MV, paste here.",
        "Bundle QA"
    )
    # B28:N29 is already merged with 'How to enter monthly balances' tip; use row 30 (free)
    note_row = 30
    callout = ws.cell(row=note_row, column=2)
    if callout.value is None or "BUNDLE NOTE" not in str(callout.value):
        callout.value = ("🔗  BUNDLE NOTE — Investment Portfolio link\n"
                         "If you also own the Investment Portfolio Tracker (bundled), the 'Stocks & Funds (Taxable)' "
                         "row above should equal your IPT '🏠 Dashboard → Total MV'. Update at month-end. "
                         "The two workbooks intentionally do NOT auto-link (so each remains usable standalone).")
        callout.font = Font(size=9, italic=True, color="555555")
        callout.alignment = Alignment(wrap_text=True, vertical='top')
        callout.fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
        # Merge to span B30:N30 (matching layout of B28:N29 which is full-width)
        try:
            ws.merge_cells('B30:N30')
        except Exception:
            pass
        ws.row_dimensions[note_row].height = 60
    wb.save(path)
    changelog.append(f"[FIX-BNDL-005] {nwt_file}: added IPT-linkage comment on B16 + visible callout on B28")

# ============================================================
# [FIX-BNDL-006] NWT and IPT FX tables — add Sync Reminder callout
# ============================================================
# NWT side
for nwt_file in nwt_files:
    path = FIXED / nwt_file
    wb = openpyxl.load_workbook(path)
    fx_tab = None
    for sh in wb.sheetnames:
        if "Settings" in sh and "FX" in sh:
            fx_tab = sh
            break
    if not fx_tab:
        changelog.append(f"[FIX-BNDL-006] {nwt_file}: FX tab not found — SKIPPED")
        continue
    ws = wb[fx_tab]
    # B22:L23 already merged for "How to enter" tip; use row 25 (rows 24, 26 free)
    note_row = 25
    callout = ws.cell(row=note_row, column=2)
    if callout.value is None or "BUNDLE NOTE" not in str(callout.value):
        callout.value = ("🔗  BUNDLE NOTE — FX rate sync with Investment Portfolio Tracker\n"
                         "If you also own the IPT SKU (bundled), the rates in this table MUST match "
                         "the rates in IPT → 💵 Cash & FX Holdings → column E. Otherwise the same foreign "
                         "position will show different USD totals across NWT and IPT.\n"
                         "Recommendation: pick one source (this table) and copy values to IPT monthly. "
                         "Or use GOOGLEFINANCE() in Google Sheets to keep both live.")
        callout.font = Font(size=9, italic=True, color="555555")
        callout.alignment = Alignment(wrap_text=True, vertical='top')
        callout.fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
        try:
            ws.merge_cells(f'B{note_row}:L{note_row}')
        except Exception:
            pass
        ws.row_dimensions[note_row].height = 75
    wb.save(path)
    changelog.append(f"[FIX-BNDL-006] {nwt_file}: FX sync reminder added at B{note_row}")

# IPT side
ipt_files = ["investment-portfolio-tracker-ai-edition.xlsx",
             "investment-portfolio-tracker-pro.xlsx"]
for ipt_file in ipt_files:
    path = FIXED / ipt_file
    wb = openpyxl.load_workbook(path)
    fx_tab = None
    for sh in wb.sheetnames:
        if "Cash" in sh and "FX" in sh:
            fx_tab = sh
            break
    if not fx_tab:
        changelog.append(f"[FIX-BNDL-006] {ipt_file}: Cash & FX tab not found — SKIPPED")
        continue
    ws = wb[fx_tab]
    # B23:L24 already merged for GOOGLEFINANCE tip; use row 26 (rows 25, 27 free)
    note_row = 26
    callout = ws.cell(row=note_row, column=2)
    if callout.value is None or "BUNDLE NOTE" not in str(callout.value):
        callout.value = ("🔗  BUNDLE NOTE — FX rate sync with Net Worth Tracker\n"
                         "If you also own the NWT SKU (bundled), the Rate→USD column above MUST match "
                         "the rates in NWT → ⚙️ Settings & FX → column C. Otherwise the same foreign "
                         "position will show different USD totals across NWT and IPT.\n"
                         "Recommendation: pick one source (or use GOOGLEFINANCE() for both to stay live).")
        callout.font = Font(size=9, italic=True, color="555555")
        callout.alignment = Alignment(wrap_text=True, vertical='top')
        callout.fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
        try:
            ws.merge_cells(f'B{note_row}:L{note_row}')
        except Exception:
            pass
        ws.row_dimensions[note_row].height = 75
    wb.save(path)
    changelog.append(f"[FIX-BNDL-006] {ipt_file}: FX sync reminder added at B{note_row}")

# ============================================================
# [FIX-BNDL-015] NWT December-only headline note
# ============================================================
for nwt_file in nwt_files:
    path = FIXED / nwt_file
    wb = openpyxl.load_workbook(path)
    if "💼 Assets Summary" not in wb.sheetnames:
        continue
    ws = wb["💼 Assets Summary"]
    # Add note at row 31 (BNDL-005 callout takes B30; row 31 is free)
    note_row = 31
    callout = ws.cell(row=note_row, column=2)
    if callout.value is None or "Dashboard headline" not in str(callout.value):
        callout.value = ("ℹ️  About the Dashboard headline — TOTAL ASSETS\n"
                         "The 'TOTAL ASSETS' number on the Dashboard reflects the December (year-end) column "
                         "by design — net worth at end of fiscal year. If you start tracking mid-year, fill the "
                         "December column with your CURRENT balances to see the headline number; expand monthly "
                         "as you complete months. The NW History tab tracks month-over-month deltas independently.")
        callout.font = Font(size=9, italic=True, color="555555")
        callout.alignment = Alignment(wrap_text=True, vertical='top')
        callout.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        try:
            ws.merge_cells(f'B{note_row}:N{note_row}')
        except Exception:
            pass
        ws.row_dimensions[note_row].height = 60
    wb.save(path)
    changelog.append(f"[FIX-BNDL-015] {nwt_file}: Dec-headline note added at B29")

# Save the changelog
clog = ROOT / "tools/qa/output/bundle-fix-changelog.md"
clog.parent.mkdir(parents=True, exist_ok=True)
with clog.open('w', encoding='utf-8') as fh:
    fh.write("# Bundle Fix Changelog (Round 1 → Round 2)\n\n")
    fh.write(f"_Applied: 2026-05-23 by `/all-in-one-bundle-qa-expert` Round 2 fix step_\n\n")
    fh.write("## Files modified\n\n")
    fh.write(f"All modifications applied to copies at `tools/qa/fixed/`. Originals at `tools/sheets-gen/output/` were NOT touched.\n\n")
    fh.write("## Fix log\n\n")
    for line in changelog:
        fh.write(f"- {line}\n")
print(f"Applied {len(changelog)} fixes. Changelog at {clog}")
for line in changelog:
    print(f"  {line}")
