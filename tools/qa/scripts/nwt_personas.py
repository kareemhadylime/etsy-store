"""NWT Stage B - drive ALL 5 personas through the AI Edition workbook.

For each persona:
  1. Copy fresh backup to a working file in tools/qa/scratch/nwt/personas/
  2. Clear Assets Summary rows + Liabilities Summary rows
  3. Write persona-specific values
  4. For FIRE Calculator: set inputs (note: E12 formula is bugged - will track)
  5. Recalc via LibreOffice
  6. Read totals via openpyxl + data_only=True
  7. Compare against Python-computed reference values
  8. Capture FIRE %, debt-to-asset, liquidity ratio behavior
"""
import json
import os
import shutil
import subprocess
import sys
from openpyxl import load_workbook

BACKUP = "C:/ETSY/etsy-store/tools/qa/backups/net-worth-tracker-ai-edition.xlsx"
WORK = "C:/ETSY/etsy-store/tools/qa/scratch/nwt/personas"
LO = "C:/Program Files/LibreOffice/program/soffice.exe"
os.makedirs(WORK, exist_ok=True)

# Asset class rows in '💼 Assets Summary' (per inspection):
# Row 9: Checking | Row 10: HYSA | Row 11: Money Market | Row 12: Foreign Currency
# Row 13: Vehicles | Row 14: Real Estate Primary | Row 15: Real Estate Investment
# Row 16: Stocks Taxable | Row 17: 401k/IRA/Roth | Row 18: HSA/529
# Row 19: Metals | Row 20: Crypto | Row 21: Business Equity | Row 22: Life Insurance DB
# Row 23: Receivables | Row 24: Other
# Row 26: TOTAL ASSETS (formula =SUM(rows))
# Columns C-N = Jan-Dec; we'll write to N (December — the live column).
ASSET_ROWS = {
    "checking": 9, "hysa": 10, "money_market": 11, "foreign_currency": 12,
    "vehicles": 13, "re_primary": 14, "re_investment": 15,
    "stocks_taxable": 16, "retirement": 17, "hsa_529": 18,
    "metals": 19, "crypto": 20, "business_equity": 21,
    "life_insurance": 22, "receivables": 23, "other": 24,
}

# Liability rows:
# Row 9: Mortgage | Row 10: Auto Loan | Row 11: Credit Card | Row 12: Student Loan
# Row 13: Personal | Row 14: Business | Row 15: Family | Row 16: Medical
# Row 17: BNPL | Row 18: Tax Owed | Row 19: Other
LIAB_ROWS = {
    "mortgage": 9, "auto": 10, "credit_card": 11, "student": 12,
    "personal": 13, "business": 14, "family": 15, "medical": 16,
    "bnpl": 17, "tax_owed": 18, "other": 19,
}

PERSONAS = [
    {
        "id": "p1_yusuf",
        "name": "Yusuf (Cairo, negative-NW, EGP)",
        "currency": "EGP",
        "assets": {"checking": 8000, "hysa": 15000, "stocks_taxable": 25000, "vehicles": 180000},
        "liabilities": {"student": 220000, "credit_card": 35000, "auto": 95000},
        "annual_spend": 180000,  # EGP — rough estimate
        "expected_assets": 228000, "expected_liabs": 350000, "expected_nw": -122000,
        "expected_da": 350000/228000,
    },
    {
        "id": "p2_mariam_tarek",
        "name": "Mariam & Tarek (Cairo dual-income family)",
        "currency": "USD",
        "assets": {"checking": 4500, "hysa": 22000, "retirement": 85000+62000, "hsa_529": 18000, "stocks_taxable": 45000, "re_primary": 480000, "vehicles": 35000},
        "liabilities": {"mortgage": 310000, "auto": 14000, "credit_card": 6500},
        "annual_spend": 90000,
        "expected_assets": 751500, "expected_liabs": 330500, "expected_nw": 421000,
        "expected_da": 330500/751500,
        "expected_home_equity": 480000 - 310000,
    },
    {
        "id": "p3_kareem_hnw",
        "name": "Kareem (Egyptian-Canadian HNW, multi-currency)",
        "currency": "USD",
        # All converted to USD (no FX table in the workbook!) - this stresses multi-currency
        "assets": {
            "checking": 50750 + 122535 + 60775 + 320000,  # all cash combined
            "stocks_taxable": 1800000,
            "re_primary": 243600,  # Cairo apt
            "re_investment": 2314550,  # Dubai RE
            "vehicles": 76244,
            "other": 220000,  # Yacht
            "business_equity": 0,  # manual entry
        },
        "liabilities": {"mortgage": 326760, "business": 150000},
        "annual_spend": 250000,
        "expected_assets": 50750+122535+60775+320000+1800000+243600+2314550+76244+220000,
        "expected_liabs": 326760+150000,
        "expected_nw": None,  # we compute below
    },
    {
        "id": "p4_hany",
        "name": "Hany (pre-retiree, FI focus)",
        "currency": "USD",
        "assets": {"checking": 45000, "hysa": 80000, "retirement": 1250000, "stocks_taxable": 420000, "re_primary": 620000, "re_investment": 280000, "vehicles": 42000},
        "liabilities": {"credit_card": 2800, "tax_owed": 4500},
        "annual_spend": 95000,
        "expected_assets": 2737000, "expected_liabs": 7300, "expected_nw": 2729700,
        "expected_fi_target": 95000*25,
        "expected_fi_progress": 2729700/(95000*25),
        "expected_liquid": 45000+80000+420000,
        "expected_liquidity_ratio_months": 545000/(95000/12),
    },
    {
        "id": "p5_layla",
        "name": "Layla (tech worker, volatile crypto)",
        "currency": "USD",
        "assets": {"checking": 8500, "hysa": 35000, "retirement": 128000, "stocks_taxable": 95000, "crypto": 76500+42000+15000, "business_equity": 25000},
        "liabilities": {"mortgage": 245000, "student": 18000},
        "annual_spend": 75000,
        "expected_assets": 425000, "expected_liabs": 263000, "expected_nw": 162000,
    },
]

# Resolve P3 expected_nw
for p in PERSONAS:
    if p["expected_nw"] is None:
        p["expected_nw"] = p["expected_assets"] - p["expected_liabs"]


def run_persona(p):
    """Apply persona inputs, recalc, read outputs."""
    work_path = os.path.join(WORK, p["id"] + ".xlsx")
    shutil.copy(BACKUP, work_path)

    wb = load_workbook(work_path)

    # Clear and write Assets Summary column N (December, the live column)
    ws_a = wb["💼 Assets Summary"]
    for row_label, row_num in ASSET_ROWS.items():
        ws_a.cell(row_num, 14).value = p["assets"].get(row_label, None)

    # Clear Liabilities Summary column N
    ws_l = wb["📉 Liabilities Summary"]
    for row_label, row_num in LIAB_ROWS.items():
        ws_l.cell(row_num, 14).value = p["liabilities"].get(row_label, None)

    # FIRE Calculator inputs (correct cells per template: C10..C14 — but E12 formula is bugged C8*C9)
    ws_f = wb["🔥 FIRE Calculator"]
    ws_f["C11"] = p["annual_spend"]
    ws_f["C12"] = 25
    ws_f["C13"] = 1850  # monthly savings

    wb.save(work_path)

    # Recalc via LibreOffice
    recalc_dir = os.path.join(WORK, "recalc")
    os.makedirs(recalc_dir, exist_ok=True)
    r = subprocess.run([LO, "--headless", "--calc", "--convert-to", "xlsx",
                        "--outdir", recalc_dir, work_path], capture_output=True, text=True, timeout=60)
    recalc_path = os.path.join(recalc_dir, os.path.basename(work_path))
    if not os.path.exists(recalc_path):
        return {"persona": p["id"], "error": "recalc failed", "stderr": r.stderr[:500]}

    # Read outputs
    wbc = load_workbook(recalc_path, data_only=True)
    ws_ac = wbc["💼 Assets Summary"]
    ws_lc = wbc["📉 Liabilities Summary"]
    ws_dc = wbc["🏠 Dashboard"]
    ws_fc = wbc["🔥 FIRE Calculator"]

    total_assets = ws_ac["N26"].value
    total_liabs = ws_lc["N21"].value

    # Dashboard FIRE
    fire_e12 = ws_fc["E12"].value  # FIRE Number
    fire_e20 = ws_fc["E20"].value  # Conservative years to FIRE

    result = {
        "persona": p["id"],
        "name": p["name"],
        "inputs": {"annual_spend": p["annual_spend"]},
        "expected_assets": p["expected_assets"],
        "actual_assets_n26": total_assets,
        "assets_match": total_assets == p["expected_assets"] if total_assets is not None else False,
        "expected_liabs": p["expected_liabs"],
        "actual_liabs_n21": total_liabs,
        "liabs_match": total_liabs == p["expected_liabs"] if total_liabs is not None else False,
        "expected_nw": p["expected_nw"],
        "actual_nw": (total_assets - total_liabs) if (total_assets is not None and total_liabs is not None) else None,
        "fire_calc_e12_fire_number": fire_e12,
        "fire_calc_e20_years_cons": fire_e20,
        "dashboard_a2_total_nw": ws_dc["A2"].value,
        "dashboard_i2_fire_pct": ws_dc["I2"].value,
        "dashboard_g2_debt_to_asset": ws_dc["G2"].value,
        "dashboard_b29_fire_pct": ws_dc["B29"].value,
    }

    # Persona-specific extras
    if p["id"] == "p2_mariam_tarek":
        # Real Estate tab for home equity computation
        try:
            ws_re = wbc["🏠 Real Estate"]
            ws_re_f = wb["🏠 Real Estate"]
            result["re_row11_F"] = ws_re["F11"].value  # home equity = C11-D11
            # actual: C11=472000 (seed), D11=268000 (seed) → 204000 by default seed
        except Exception as e:
            result["re_error"] = str(e)

    if p["id"] == "p4_hany":
        result["fire_progress_expected_pct"] = p["expected_fi_progress"]

    return result


results = []
for p in PERSONAS:
    print(f"=== {p['id']} ({p['name']}) ===", flush=True)
    r = run_persona(p)
    results.append(r)
    print(json.dumps(r, indent=2, default=str), flush=True)

with open("C:/ETSY/etsy-store/tools/qa/round1/persona_results.json", "w", encoding="utf-8") as fh:
    json.dump(results, fh, indent=2, default=str)
print("\nWrote: C:/ETSY/etsy-store/tools/qa/round1/persona_results.json")
