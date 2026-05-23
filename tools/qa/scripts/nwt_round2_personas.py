"""Round 2 — drive personas through the FIXED workbooks.

Compare to Round 1 results — every NWT-### must now resolve.
"""
import json
import os
import shutil
import subprocess
from openpyxl import load_workbook

BACKUP = "C:/ETSY/etsy-store/tools/qa/fixed/net-worth-tracker-ai-edition.xlsx"
WORK = "C:/ETSY/etsy-store/tools/qa/scratch/nwt/round2-personas"
LO = "C:/Program Files/LibreOffice/program/soffice.exe"
os.makedirs(WORK, exist_ok=True)

ASSET_ROWS = {
    "checking": 9, "hysa": 10, "money_market": 11, "foreign_currency": 12,
    "vehicles": 13, "re_primary": 14, "re_investment": 15,
    "stocks_taxable": 16, "retirement": 17, "hsa_529": 18,
    "metals": 19, "crypto": 20, "business_equity": 21,
    "life_insurance": 22, "receivables": 23, "other": 24,
}
LIAB_ROWS = {
    "mortgage": 9, "auto": 10, "credit_card": 11, "student": 12,
    "personal": 13, "business": 14, "family": 15, "medical": 16,
    "bnpl": 17, "tax_owed": 18, "other": 19,
}

PERSONAS = [
    {
        "id": "p1_yusuf",
        "name": "Yusuf (Cairo, negative-NW, EGP)",
        "assets": {"checking": 8000, "hysa": 15000, "stocks_taxable": 25000, "vehicles": 180000},
        "liabilities": {"student": 220000, "credit_card": 35000, "auto": 95000},
        "annual_spend": 180000,
        "expected_assets": 228000, "expected_liabs": 350000, "expected_nw": -122000,
        "expected_fire_number": 180000*25,
    },
    {
        "id": "p2_mariam_tarek",
        "name": "Mariam & Tarek (dual-income family)",
        "assets": {"checking": 4500, "hysa": 22000, "retirement": 85000+62000, "hsa_529": 18000, "stocks_taxable": 45000, "re_primary": 480000, "vehicles": 35000},
        "liabilities": {"mortgage": 310000, "auto": 14000, "credit_card": 6500},
        "annual_spend": 90000,
        "expected_assets": 751500, "expected_liabs": 330500, "expected_nw": 421000,
        "expected_fire_number": 90000*25,
    },
    {
        "id": "p3_kareem_hnw",
        "name": "Kareem (HNW multi-currency)",
        "assets": {
            "checking": 50750 + 122535 + 60775 + 320000,
            "stocks_taxable": 1800000,
            "re_primary": 243600,
            "re_investment": 2314550,
            "vehicles": 76244,
            "other": 220000,
        },
        "liabilities": {"mortgage": 326760, "business": 150000},
        "annual_spend": 250000,
        "expected_assets": 5208454,
        "expected_liabs": 476760,
        "expected_fire_number": 250000*25,
    },
    {
        "id": "p4_hany",
        "name": "Hany (pre-retiree, FI focus)",
        "assets": {"checking": 45000, "hysa": 80000, "retirement": 1250000, "stocks_taxable": 420000, "re_primary": 620000, "re_investment": 280000, "vehicles": 42000},
        "liabilities": {"credit_card": 2800, "tax_owed": 4500},
        "annual_spend": 95000,
        "expected_assets": 2737000, "expected_liabs": 7300, "expected_nw": 2729700,
        "expected_fire_number": 95000*25,
        "expected_fi_progress": 2729700/(95000*25),
    },
    {
        "id": "p5_layla",
        "name": "Layla (volatile crypto)",
        "assets": {"checking": 8500, "hysa": 35000, "retirement": 128000, "stocks_taxable": 95000, "crypto": 76500+42000+15000, "business_equity": 25000},
        "liabilities": {"mortgage": 245000, "student": 18000},
        "annual_spend": 75000,
        "expected_assets": 425000, "expected_liabs": 263000, "expected_nw": 162000,
        "expected_fire_number": 75000*25,
    },
]
for p in PERSONAS:
    if "expected_nw" not in p:
        p["expected_nw"] = p["expected_assets"] - p["expected_liabs"]


def run_persona(p):
    work_path = os.path.join(WORK, p["id"] + ".xlsx")
    shutil.copy(BACKUP, work_path)
    wb = load_workbook(work_path)
    ws_a = wb["💼 Assets Summary"]
    for row_label, row_num in ASSET_ROWS.items():
        ws_a.cell(row_num, 14).value = p["assets"].get(row_label, None)
    ws_l = wb["📉 Liabilities Summary"]
    for row_label, row_num in LIAB_ROWS.items():
        ws_l.cell(row_num, 14).value = p["liabilities"].get(row_label, None)
    ws_f = wb["🔥 FIRE Calculator"]
    ws_f["C11"] = p["annual_spend"]
    ws_f["C12"] = 25
    ws_f["C13"] = 1850
    wb.save(work_path)

    recalc_dir = os.path.join(WORK, "recalc")
    os.makedirs(recalc_dir, exist_ok=True)
    r = subprocess.run([LO, "--headless", "--calc", "--convert-to", "xlsx",
                        "--outdir", recalc_dir, work_path], capture_output=True, text=True, timeout=60)
    recalc_path = os.path.join(recalc_dir, os.path.basename(work_path))
    if not os.path.exists(recalc_path):
        return {"persona": p["id"], "error": "recalc failed", "stderr": r.stderr[:500]}

    wbc = load_workbook(recalc_path, data_only=True)
    ws_ac = wbc["💼 Assets Summary"]
    ws_lc = wbc["📉 Liabilities Summary"]
    ws_dc = wbc["🏠 Dashboard"]
    ws_fc = wbc["🔥 FIRE Calculator"]

    total_assets = ws_ac["N26"].value
    total_liabs = ws_lc["N21"].value

    return {
        "persona": p["id"],
        "name": p["name"],
        "inputs": {"annual_spend": p["annual_spend"]},
        "expected_assets": p["expected_assets"],
        "actual_assets_n26": total_assets,
        "assets_match": total_assets == p["expected_assets"] if total_assets is not None else False,
        "expected_liabs": p["expected_liabs"],
        "actual_liabs_n21": total_liabs,
        "liabs_match": total_liabs == p["expected_liabs"] if total_liabs is not None else False,
        "expected_nw": p["expected_nw"],
        "actual_nw": (total_assets - total_liabs) if (total_assets is not None and total_liabs is not None) else None,
        "expected_fire_number": p["expected_fire_number"],
        "fire_calc_e12_fire_number": ws_fc["E12"].value,
        "fire_e16_aggressive_years": ws_fc["E16"].value,
        "fire_e18_current_years": ws_fc["E18"].value,
        "fire_e20_conservative_years": ws_fc["E20"].value,
        "dashboard_a2_total_nw": ws_dc["A2"].value,
        "dashboard_i2_fire_pct": ws_dc["I2"].value,
        "dashboard_g2_debt_to_asset": ws_dc["G2"].value,
        "dashboard_b29_fire_pct": ws_dc["B29"].value,
        "dashboard_e29_fire_number_kpi": ws_dc["E29"].value,
        "dashboard_i29_years_to_fire": ws_dc["I29"].value,
        "dashboard_c20_months_expenses": ws_dc["C20"].value,
        "dashboard_f20_liquid_nw": ws_dc["F20"].value,
        "dashboard_i20_fi_progress": ws_dc["I20"].value,
        "dashboard_l20_nw_mom_delta": ws_dc["L20"].value,
        "dashboard_b10_health_score": ws_dc["B10"].value,
        # asset mix J11..J16
        "dashboard_j11_real_estate_pct": ws_dc["J11"].value,
        "dashboard_j12_stocks_pct": ws_dc["J12"].value,
        "dashboard_j13_metals_crypto_pct": ws_dc["J13"].value,
        "dashboard_j14_cash_pct": ws_dc["J14"].value,
        "dashboard_j15_business_pct": ws_dc["J15"].value,
        "dashboard_j16_other_pct": ws_dc["J16"].value,
    }


results = []
for p in PERSONAS:
    print(f"=== {p['id']} ({p['name']}) ===", flush=True)
    r = run_persona(p)
    results.append(r)
    print(json.dumps(r, indent=2, default=str), flush=True)

with open("C:/ETSY/etsy-store/tools/qa/round2/persona_results.json", "w", encoding="utf-8") as fh:
    json.dump(results, fh, indent=2, default=str)
print("\nWrote: tools/qa/round2/persona_results.json")
