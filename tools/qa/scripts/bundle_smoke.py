"""Bundle Round 1: Smoke test all 18 SKU xlsx files + 12 SKU PDFs + 30 SKU thumbnails."""
import sys, io, os, hashlib
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', line_buffering=True)
import openpyxl, pypdf
from PIL import Image
from pathlib import Path

ROOT = Path("C:/ETSY/etsy-store")
SCRATCH = ROOT / "tools/qa/scratch"
SHEETS = ROOT / "tools/sheets-gen/output"
PDFS = ROOT / "tools/pdf-gen/output"
THUMBS = ROOT / "tools/thumb-gen/output"

SKUS = {
    "budget-tracker": {"v2": "budget-tracker-ai-edition-v2",
                       "essentials": "budget-tracker-essentials",
                       "pro": "budget-tracker-pro"},
    "debt-payoff-planner": {"ai": "debt-payoff-planner-ai-edition",
                            "essentials": "debt-payoff-planner-essentials",
                            "pro": "debt-payoff-planner-pro"},
    "sinking-funds-planner": {"ai": "sinking-funds-planner-ai-edition",
                              "essentials": "sinking-funds-planner-essentials",
                              "pro": "sinking-funds-planner-pro"},
    "net-worth-tracker": {"ai": "net-worth-tracker-ai-edition",
                          "essentials": "net-worth-tracker-essentials",
                          "pro": "net-worth-tracker-pro"},
    "investment-portfolio-tracker": {"ai": "investment-portfolio-tracker-ai-edition",
                                      "essentials": "investment-portfolio-tracker-essentials",
                                      "pro": "investment-portfolio-tracker-pro"},
    "family-education-planner": {"ai": "family-education-planner-ai-edition",
                                  "essentials": "family-education-planner-essentials",
                                  "pro": "family-education-planner-pro"},
}

ERROR_TYPES = ['#REF!', '#N/A', '#DIV/0!', '#NAME?', '#VALUE!', '#NULL!', '#NUM!']

findings = []

def smoke_xlsx(name, sku_label, tier):
    """A1-A7 protocol."""
    path = SCRATCH / f"{name}.xlsx"
    issues = []
    if not path.exists():
        issues.append(f"FILE-MISSING: recalc copy not at {path}")
        return {"issues": issues, "tabs": [], "errors": []}
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        issues.append(f"LOAD-FAIL: {e}")
        return {"issues": issues, "tabs": [], "errors": []}
    tabs = wb.sheetnames
    err_cells = []
    none_cells = 0
    formula_cells = 0
    # Also load with data_only=False to detect formula presence
    wb_f = openpyxl.load_workbook(path, data_only=False)
    for sh in wb.sheetnames:
        ws = wb[sh]
        ws_f = wb_f[sh]
        for row_f in ws_f.iter_rows():
            for c in row_f:
                if isinstance(c.value, str) and c.value.startswith('='):
                    formula_cells += 1
                    # check evaluated value
                    ev = ws[c.coordinate].value
                    if ev is None:
                        none_cells += 1
                    elif isinstance(ev, str) and ev in ERROR_TYPES:
                        err_cells.append(f"{sh}!{c.coordinate}={ev}  (formula:{c.value[:60]})")
        # also scan plain cells for error strings
        for row in ws.iter_rows(values_only=False):
            for c in row:
                if isinstance(c.value, str) and c.value in ERROR_TYPES:
                    err_cells.append(f"{sh}!{c.coordinate}={c.value}")
    if err_cells:
        issues.append(f"FORMULA-ERRORS ({len(err_cells)}): " + "; ".join(err_cells[:5]))
    if none_cells > 0:
        # Some None is acceptable for empty input cells; only flag if substantial
        pct = 100.0 * none_cells / max(formula_cells, 1)
        if pct > 25:
            issues.append(f"NONE-EVALUATED: {none_cells}/{formula_cells} formulas evaluated to None ({pct:.0f}%). Recalc may have skipped or formulas reference empty inputs.")
        elif none_cells > 10:
            issues.append(f"NONE-EVALUATED-MINOR: {none_cells}/{formula_cells} formulas eval-to-None ({pct:.1f}%) - acceptable if dependent on user inputs.")
    # named ranges
    try:
        names = list(wb.defined_names.keys())
        broken_names = []
        for n in names:
            try:
                dn = wb.defined_names[n]
                if dn.value is None:
                    broken_names.append(n)
            except Exception as e:
                broken_names.append(f"{n} ({e})")
        if broken_names:
            issues.append(f"NAMED-RANGES-BROKEN: {broken_names[:5]}")
    except Exception as e:
        issues.append(f"NAMED-RANGE-READ-FAIL: {e}")
    # GOOGLEFINANCE / IMPORTXML scan (formula text)
    sheets_only_funcs = []
    for sh in wb_f.sheetnames:
        ws_f = wb_f[sh]
        for row in ws_f.iter_rows(values_only=False):
            for c in row:
                if isinstance(c.value, str) and c.value.startswith('='):
                    fv = c.value.upper()
                    for fn in ('GOOGLEFINANCE', 'IMPORTXML', 'IMPORTDATA', 'IMPORTHTML', 'IMPORTFEED', 'SPARKLINE', 'IMAGE('):
                        if fn in fv:
                            sheets_only_funcs.append(f"{sh}!{c.coordinate}: {fn}")
                            break
    if sheets_only_funcs:
        issues.append(f"SHEETS-ONLY-FUNCS ({len(sheets_only_funcs)}): {sheets_only_funcs[:3]}")
    # SWITCH / MINIFS / MAXIFS scan (LibreOffice 26.x compatibility concern from persona notes)
    lo_concerns = []
    for sh in wb_f.sheetnames:
        ws_f = wb_f[sh]
        for row in ws_f.iter_rows(values_only=False):
            for c in row:
                if isinstance(c.value, str) and c.value.startswith('='):
                    fv = c.value.upper()
                    for fn in ('=SWITCH(', 'SWITCH(', '=MINIFS(', 'MINIFS(', '=MAXIFS(', 'MAXIFS('):
                        if fn in fv:
                            lo_concerns.append(f"{sh}!{c.coordinate}: {fn}")
                            break
    if lo_concerns:
        issues.append(f"LO-26X-CONCERN ({len(lo_concerns)}): {lo_concerns[:3]}")
    return {"issues": issues, "tabs": tabs, "errors": err_cells, "formula_count": formula_cells, "none_count": none_cells}

def smoke_pdf(path):
    """A8-A12."""
    issues = []
    if not path.exists():
        return [f"FILE-MISSING: {path}"]
    try:
        rd = pypdf.PdfReader(str(path))
        npages = len(rd.pages)
        first_text = rd.pages[0].extract_text() or ""
    except Exception as e:
        return [f"PDF-LOAD-FAIL: {e}"]
    if npages < 1:
        issues.append(f"EMPTY-PDF: {path.name}")
    # check for AI artifacts
    full_text = ""
    for p in rd.pages[:min(npages, 30)]:
        full_text += (p.extract_text() or "") + "\n"
    artifacts = ['As an AI', 'I cannot ', "I'm just an AI", '[INSERT', '[TODO', 'lorem ipsum', 'Lorem ipsum', '<placeholder']
    for a in artifacts:
        if a.lower() in full_text.lower():
            issues.append(f"AI-ARTIFACT in {path.name}: contains '{a}'")
    return issues

def smoke_thumb(path):
    """A13-A14."""
    issues = []
    if not path.exists():
        return [f"FILE-MISSING: {path.name}"]
    try:
        img = Image.open(str(path))
        w, h = img.size
        if w != 2000 or h != 2000:
            issues.append(f"DIM-FAIL: {path.name} is {w}x{h} (expected 2000x2000)")
        size_mb = path.stat().st_size / (1024*1024)
        if size_mb > 20:
            issues.append(f"SIZE-FAIL: {path.name} is {size_mb:.2f} MB (>20 MB)")
        # legibility check (downsample to 250 — just verify it loads, can't OCR here)
        small = img.resize((250, 250), Image.LANCZOS)
        small.close()
        img.close()
    except Exception as e:
        return [f"IMAGE-LOAD-FAIL: {path.name}: {e}"]
    return issues

# Run all 18 xlsx smoke tests
print("="*78)
print("BUNDLE ROUND 1 SMOKE TESTS")
print("="*78)
rows = []
for sku, tiers in SKUS.items():
    for tlabel, fname in tiers.items():
        r = smoke_xlsx(fname, sku, tlabel)
        verdict = "PASS"
        crit_keys = ['FILE-MISSING', 'LOAD-FAIL', 'FORMULA-ERRORS', 'NONE-EVALUATED:']
        for iss in r['issues']:
            if any(iss.startswith(k) for k in crit_keys):
                verdict = "FAIL"
                break
        if verdict == "PASS" and r['issues']:
            verdict = "PASS-WITH-MINOR-ISSUES"
        rows.append({"sku": sku, "tier": tlabel, "file": fname, "verdict": verdict, "issues": r['issues'], "tabs": r['tabs'], "formula_count": r.get('formula_count', 0), "none_count": r.get('none_count', 0)})
        print(f"\n--- {sku} :: {tlabel} :: {verdict}")
        if r['issues']:
            for iss in r['issues']:
                print(f"  - {iss}")
        else:
            print(f"  CLEAN: {len(r['tabs'])} tabs, {r.get('formula_count',0)} formulas")

# PDFs
pdf_map = {
    "budget-tracker": ["budget-tracker-ai-pdf.pdf", "budget-tracker-quickstart.pdf"],
    "debt-payoff-planner": ["debt-payoff-ai-pdf.pdf", "debt-payoff-quickstart.pdf"],
    "sinking-funds-planner": ["sinking-funds-ai-pdf.pdf", "sinking-funds-quickstart.pdf"],
    "net-worth-tracker": ["net-worth-ai-pdf.pdf", "net-worth-quickstart.pdf"],
    "investment-portfolio-tracker": ["investment-portfolio-ai-pdf.pdf", "investment-portfolio-quickstart.pdf"],
    "family-education-planner": ["family-education-ai-pdf.pdf", "family-education-quickstart.pdf"],
}
print("\n" + "="*78)
print("PDF SMOKE TESTS")
print("="*78)
pdf_rows = []
for sku, pdfs in pdf_map.items():
    for fname in pdfs:
        path = PDFS / fname
        iss = smoke_pdf(path)
        verdict = "PASS" if not iss else ("FAIL" if any('LOAD-FAIL' in i or 'MISSING' in i for i in iss) else "PASS-WITH-MINOR-ISSUES")
        pdf_rows.append({"sku": sku, "file": fname, "verdict": verdict, "issues": iss})
        if iss:
            print(f"\n--- {fname} :: {verdict}")
            for i in iss:
                print(f"  - {i}")
        else:
            print(f"--- {fname} :: PASS")

# Thumbnails
thumb_dirs = {
    "budget-tracker": ["budget-tracker-01-hero", "budget-tracker-02-health-score", "budget-tracker-03-methods", "budget-tracker-04-ai-advisor", "budget-tracker-05-privacy"],
    "debt-payoff-planner": ["debt-payoff-planner-01-hero", "debt-payoff-planner-02-strategy-comparison", "debt-payoff-planner-03-health-score", "debt-payoff-planner-04-methods", "debt-payoff-planner-05-privacy"],
    "sinking-funds-planner": ["sinking-funds-planner-01-hero", "sinking-funds-planner-02-priority-matrix", "sinking-funds-planner-03-vehicles", "sinking-funds-planner-04-ai-advisor", "sinking-funds-planner-05-anti-qapital"],
    "net-worth-tracker": ["net-worth-tracker-01-hero", "net-worth-tracker-02-fire-calculator", "net-worth-tracker-03-asset-mix", "net-worth-tracker-04-ai-advisor", "net-worth-tracker-05-anti-plaid"],
    "investment-portfolio-tracker": ["investment-portfolio-tracker-01-hero", "investment-portfolio-tracker-02-holdings", "investment-portfolio-tracker-03-risk-allocation", "investment-portfolio-tracker-04-ai-advisor", "investment-portfolio-tracker-05-anti-sharesight"],
    "family-education-planner": ["family-education-planner-01-hero", "family-education-planner-02-account-comparison", "family-education-planner-03-efc-aid", "family-education-planner-04-ai-advisor", "family-education-planner-05-anti-greenlight"],
}
print("\n" + "="*78)
print("THUMBNAIL SMOKE TESTS")
print("="*78)
thumb_rows = []
for sku, ts in thumb_dirs.items():
    for t in ts:
        path = THUMBS / (t + ".png")
        iss = smoke_thumb(path)
        verdict = "PASS" if not iss else ("FAIL" if any('MISSING' in i or 'LOAD-FAIL' in i for i in iss) else "PASS-WITH-MINOR-ISSUES")
        thumb_rows.append({"sku": sku, "file": t, "verdict": verdict, "issues": iss})
        if iss:
            print(f"--- {t}.png :: {verdict}  {iss}")

# Write summary to round1
import json
out = ROOT / "tools/qa/round1/part-a-smoke-tests.json"
with out.open('w', encoding='utf-8') as fh:
    json.dump({"xlsx": rows, "pdfs": pdf_rows, "thumbs": thumb_rows}, fh, indent=2, default=str)
print(f"\nWrote {out}")

# tally
print("\n" + "="*78)
print("PART A SUMMARY")
print("="*78)
xlsx_pass = sum(1 for r in rows if r['verdict'] == 'PASS')
xlsx_minor = sum(1 for r in rows if r['verdict'] == 'PASS-WITH-MINOR-ISSUES')
xlsx_fail = sum(1 for r in rows if r['verdict'] == 'FAIL')
print(f"xlsx: {xlsx_pass} PASS, {xlsx_minor} PASS-WITH-MINOR, {xlsx_fail} FAIL")
pdf_pass = sum(1 for r in pdf_rows if r['verdict'] == 'PASS')
pdf_minor = sum(1 for r in pdf_rows if r['verdict'] == 'PASS-WITH-MINOR-ISSUES')
pdf_fail = sum(1 for r in pdf_rows if r['verdict'] == 'FAIL')
print(f"pdfs: {pdf_pass} PASS, {pdf_minor} PASS-WITH-MINOR, {pdf_fail} FAIL")
thumb_pass = sum(1 for r in thumb_rows if r['verdict'] == 'PASS')
thumb_fail = sum(1 for r in thumb_rows if r['verdict'] == 'FAIL')
print(f"thumbs: {thumb_pass} PASS, {thumb_fail} FAIL")
