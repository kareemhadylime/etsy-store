"""NWT Stage C — edge-case probing.

For each edge case, use a fresh copy of the backup and write specific values.
"""
import json
import os
import shutil
import subprocess
from openpyxl import load_workbook

BACKUP = "C:/ETSY/etsy-store/tools/qa/backups/net-worth-tracker-ai-edition.xlsx"
WORK = "C:/ETSY/etsy-store/tools/qa/scratch/nwt/edges"
LO = "C:/Program Files/LibreOffice/program/soffice.exe"
os.makedirs(WORK, exist_ok=True)


def recalc(path):
    out_dir = os.path.join(WORK, "recalc")
    os.makedirs(out_dir, exist_ok=True)
    subprocess.run([LO, "--headless", "--calc", "--convert-to", "xlsx",
                    "--outdir", out_dir, path], capture_output=True, text=True, timeout=60)
    return os.path.join(out_dir, os.path.basename(path))


def edge_zero_assets():
    """Edge: zero assets / all liabilities — verify no #DIV/0!."""
    work = os.path.join(WORK, "edge_zero_assets.xlsx")
    shutil.copy(BACKUP, work)
    wb = load_workbook(work)
    ws_a = wb["💼 Assets Summary"]
    # Clear all asset rows (rows 9-24 in col N)
    for r in range(9, 25):
        ws_a.cell(r, 14).value = None
    ws_l = wb["📉 Liabilities Summary"]
    ws_l.cell(11, 14).value = 5000  # credit card
    wb.save(work)
    out = recalc(work)
    wbc = load_workbook(out, data_only=True)
    return {
        "case": "zero_assets",
        "total_assets": wbc["💼 Assets Summary"]["N26"].value,
        "total_liabs": wbc["📉 Liabilities Summary"]["N21"].value,
        "dashboard_total": wbc["🏠 Dashboard"]["A2"].value,
        "dashboard_debt_to_asset": wbc["🏠 Dashboard"]["G2"].value,
    }


def edge_huge_values():
    """Edge: 8-digit / 9-digit values (HNW persona 3)."""
    work = os.path.join(WORK, "edge_huge.xlsx")
    shutil.copy(BACKUP, work)
    wb = load_workbook(work)
    ws_a = wb["💼 Assets Summary"]
    for r in range(9, 25):
        ws_a.cell(r, 14).value = None
    ws_a.cell(14, 14).value = 12500000  # primary residence 12.5M
    ws_a.cell(15, 14).value = 95000000  # investment RE 95M
    ws_a.cell(17, 14).value = 8500000  # retirement 8.5M
    ws_a.cell(16, 14).value = 17000000  # stocks 17M
    wb.save(work)
    out = recalc(work)
    wbc = load_workbook(out, data_only=True)
    return {
        "case": "huge_values",
        "total_assets": wbc["💼 Assets Summary"]["N26"].value,
        "dashboard_total": wbc["🏠 Dashboard"]["A2"].value,
    }


def edge_negative_input():
    """Edge: enter negative balance (foot-gun) for asset."""
    work = os.path.join(WORK, "edge_negative.xlsx")
    shutil.copy(BACKUP, work)
    wb = load_workbook(work)
    ws_a = wb["💼 Assets Summary"]
    for r in range(9, 25):
        ws_a.cell(r, 14).value = None
    ws_a.cell(9, 14).value = -500  # negative checking (overdraft)
    ws_a.cell(10, 14).value = 10000
    wb.save(work)
    out = recalc(work)
    wbc = load_workbook(out, data_only=True)
    return {
        "case": "negative_input",
        "total_assets": wbc["💼 Assets Summary"]["N26"].value,
        "dashboard_total": wbc["🏠 Dashboard"]["A2"].value,
    }


def edge_string_in_amount():
    """Edge: user types '12,000' (with comma) or '$12k' — should validate."""
    work = os.path.join(WORK, "edge_string.xlsx")
    shutil.copy(BACKUP, work)
    wb = load_workbook(work)
    ws_a = wb["💼 Assets Summary"]
    for r in range(9, 25):
        ws_a.cell(r, 14).value = None
    ws_a.cell(9, 14).value = "12,000"  # string with comma
    ws_a.cell(10, 14).value = 5000
    wb.save(work)
    out = recalc(work)
    wbc = load_workbook(out, data_only=True)
    return {
        "case": "string_in_amount",
        "total_assets": wbc["💼 Assets Summary"]["N26"].value,
        "dashboard_total": wbc["🏠 Dashboard"]["A2"].value,
    }


def edge_single_asset_no_liab():
    """Edge: one asset, no liabilities — debt-to-asset = 0%."""
    work = os.path.join(WORK, "edge_single.xlsx")
    shutil.copy(BACKUP, work)
    wb = load_workbook(work)
    for r in range(9, 25):
        wb["💼 Assets Summary"].cell(r, 14).value = None
    for r in range(9, 20):
        wb["📉 Liabilities Summary"].cell(r, 14).value = None
    wb["💼 Assets Summary"].cell(9, 14).value = 50000
    wb.save(work)
    out = recalc(work)
    wbc = load_workbook(out, data_only=True)
    return {
        "case": "single_asset",
        "total_assets": wbc["💼 Assets Summary"]["N26"].value,
        "total_liabs": wbc["📉 Liabilities Summary"]["N21"].value,
        "dashboard_total": wbc["🏠 Dashboard"]["A2"].value,
        "dashboard_debt_to_asset": wbc["🏠 Dashboard"]["G2"].value,
    }


cases = [edge_zero_assets, edge_huge_values, edge_negative_input, edge_string_in_amount, edge_single_asset_no_liab]
results = []
for fn in cases:
    print(f"--- {fn.__name__} ---", flush=True)
    r = fn()
    results.append(r)
    print(json.dumps(r, indent=2, default=str), flush=True)

with open("C:/ETSY/etsy-store/tools/qa/round1/edge_results.json", "w", encoding="utf-8") as fh:
    json.dump(results, fh, indent=2, default=str)
print("\nWrote: tools/qa/round1/edge_results.json")
