"""Dig into failure modes to separate real issues from false alarms."""
import sys, io
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', line_buffering=True)
import openpyxl

ROOT = Path("C:/ETSY/etsy-store")
SCRATCH = ROOT / "tools/qa/scratch"

# 1. Check the IPT Scenario Simulator broken formula
print("="*78)
print("1. IPT Scenario Simulator G2/I2 broken formula")
print("="*78)
path = SCRATCH / "investment-portfolio-tracker-ai-edition.xlsx"
wb = openpyxl.load_workbook(path, data_only=False)
for sh in wb.sheetnames:
    if "Scenario" in sh:
        ws = wb[sh]
        for coord in ['G2', 'I2', 'C11', 'C13']:
            c = ws[coord]
            print(f"  {sh}!{coord}: {repr(c.value)[:150]}")
        # also values
        wbv = openpyxl.load_workbook(path, data_only=True)
        wsv = wbv[sh]
        for coord in ['G2', 'I2', 'C11', 'C13']:
            print(f"  EVAL {sh}!{coord} = {wsv[coord].value!r}")

# 2. Sample None-evaluated cells in DPP to assess legitimacy
print("\n" + "="*78)
print("2. DPP None-eval - are they tied to empty user inputs?")
print("="*78)
path = SCRATCH / "debt-payoff-planner-ai-edition.xlsx"
wb_f = openpyxl.load_workbook(path, data_only=False)
wb_d = openpyxl.load_workbook(path, data_only=True)
sample_none = []
sample_legit = []
for sh in wb_f.sheetnames[:10]:
    ws_f = wb_f[sh]
    ws_d = wb_d[sh]
    for r in ws_f.iter_rows(max_row=30):
        for c in r:
            if isinstance(c.value, str) and c.value.startswith('='):
                ev = ws_d[c.coordinate].value
                if ev is None and len(sample_none) < 12:
                    sample_none.append((sh, c.coordinate, c.value[:80]))
                if ev is not None and len(sample_legit) < 5:
                    sample_legit.append((sh, c.coordinate, c.value[:60], str(ev)[:30]))
print("\nDPP None-eval sample:")
for s in sample_none:
    print(f"  {s[0]}!{s[1]}: {s[2]}")
print("\nDPP working formula sample:")
for s in sample_legit:
    print(f"  {s[0]}!{s[1]}: {s[2]} -> {s[3]}")

# 3. NWT essentials - same dig
print("\n" + "="*78)
print("3. NWT essentials None-eval - real or empty-input deps?")
print("="*78)
path = SCRATCH / "net-worth-tracker-essentials.xlsx"
wb_f = openpyxl.load_workbook(path, data_only=False)
wb_d = openpyxl.load_workbook(path, data_only=True)
sample_none = []
for sh in wb_f.sheetnames:
    ws_f = wb_f[sh]
    ws_d = wb_d[sh]
    for r in ws_f.iter_rows():
        for c in r:
            if isinstance(c.value, str) and c.value.startswith('='):
                ev = ws_d[c.coordinate].value
                if ev is None and len(sample_none) < 15:
                    sample_none.append((sh, c.coordinate, c.value[:80]))
print("NWT-essentials None-eval sample:")
for s in sample_none:
    print(f"  {s[0]}!{s[1]}: {s[2]}")

# 4. Sinking Funds PDF: "I cannot "
print("\n" + "="*78)
print("4. SFP advisor PDF - what's the 'I cannot' context?")
print("="*78)
import pypdf
p = ROOT / "tools/pdf-gen/output/sinking-funds-ai-pdf.pdf"
rd = pypdf.PdfReader(str(p))
for i, page in enumerate(rd.pages):
    text = page.extract_text() or ""
    if "I cannot " in text or "I can't" in text or "i cannot" in text.lower():
        # show context
        idx = text.lower().find("i cannot")
        if idx < 0: idx = text.lower().find("i can't")
        ctx = text[max(0,idx-100):idx+200]
        print(f"  page {i+1}: ...{ctx}...")
