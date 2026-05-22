"""Dump all cells (value + formula) from a workbook to a flat text file."""
import sys, openpyxl
from openpyxl.utils import get_column_letter

path = sys.argv[1]
out = sys.argv[2] if len(sys.argv) > 2 else None

wb = openpyxl.load_workbook(path, data_only=False, keep_vba=True)
lines = []
lines.append(f"# Workbook: {path}")
lines.append(f"# Defined names: {[d for d in wb.defined_names]}")
for sname in wb.sheetnames:
    ws = wb[sname]
    state = ws.sheet_state
    lines.append(f"\n## Sheet: {sname!r}  state={state}  dim={ws.dimensions}  max_row={ws.max_row}  max_col={ws.max_column}")
    # data validations
    if ws.data_validations.dataValidation:
        for dv in ws.data_validations.dataValidation:
            lines.append(f"  DV: type={dv.type} formula1={dv.formula1!r} ranges={[str(r) for r in dv.sqref.ranges] if dv.sqref else []}")
    # conditional formatting
    for rng, rules in ws.conditional_formatting._cf_rules.items():
        for r in rules:
            try:
                f = r.formula
            except Exception:
                f = None
            lines.append(f"  CF: range={rng.sqref} type={r.type} formula={f} operator={getattr(r,'operator',None)} text={getattr(r,'text',None)}")
    # charts
    n_charts = len(ws._charts)
    lines.append(f"  Charts: {n_charts}")
    # merged
    if ws.merged_cells.ranges:
        lines.append(f"  Merged: {[str(m) for m in ws.merged_cells.ranges]}")
    # cells
    for row in ws.iter_rows():
        for c in row:
            if c.value is None and not c.data_type == 'f':
                continue
            addr = c.coordinate
            v = c.value
            t = c.data_type  # 'n' number, 's' string, 'f' formula, 'd' date, 'b' bool
            nf = c.number_format if c.number_format != 'General' else ''
            fill = c.fill.fgColor.rgb if c.fill and c.fill.fgColor and c.fill.fgColor.rgb else ''
            lines.append(f"  {addr:>5}  t={t}  nf={nf!r:<22}  fill={fill!r:<10}  val={v!r}")

text = "\n".join(lines)
if out:
    with open(out, "w", encoding="utf-8") as f:
        f.write(text)
    print(f"Wrote {out} ({len(lines)} lines)")
else:
    print(text)
