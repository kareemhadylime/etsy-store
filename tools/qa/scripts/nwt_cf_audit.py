"""NWT Stage D — Conditional Formatting audit."""
import json
import os
from openpyxl import load_workbook

SCRATCH = "C:/ETSY/etsy-store/tools/qa/scratch/nwt"

FILES = {
    "essentials": "net-worth-tracker-essentials.xlsx",
    "pro": "net-worth-tracker-pro.xlsx",
    "ai": "net-worth-tracker-ai-edition.xlsx",
}

report = {}
for tier, fn in FILES.items():
    wb = load_workbook(os.path.join(SCRATCH, fn), data_only=False)
    tier_rules = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rules_here = []
        try:
            for cf_range, cf_list in ws.conditional_formatting._cf_rules.items():
                for rule in cf_list:
                    rules_here.append({
                        "range": str(cf_range),
                        "type": rule.type,
                        "operator": getattr(rule, "operator", None),
                        "formula": [str(f) for f in (rule.formula or [])][:3],
                        "text": getattr(rule, "text", None),
                    })
        except Exception as e:
            rules_here.append({"error": str(e)})
        if rules_here:
            tier_rules[sn] = rules_here
    report[tier] = tier_rules

with open("C:/ETSY/etsy-store/tools/qa/round1/cf_rules.json", "w", encoding="utf-8") as f:
    json.dump(report, f, indent=2, default=str)
# print summary
for tier in report:
    total = sum(len(rules) for rules in report[tier].values())
    print(f"{tier}: {total} CF rules across {len(report[tier])} sheets")
print("\nWrote: tools/qa/round1/cf_rules.json")
