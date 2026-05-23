"""Bundle Round 1: Part B — bundle-level audit (B1-B6)."""
import sys, io, os, hashlib, re, zipfile
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', line_buffering=True)
import openpyxl, pypdf
from PIL import Image
from collections import Counter

ROOT = Path("C:/ETSY/etsy-store")
SCRATCH = ROOT / "tools/qa/scratch"
SHEETS = ROOT / "tools/sheets-gen/output"
PDFS = ROOT / "tools/pdf-gen/output"
THUMBS = ROOT / "tools/thumb-gen/output"

SKUS_AI = {
    "Budget Tracker": "budget-tracker-ai-edition-v2",
    "Debt Payoff Planner": "debt-payoff-planner-ai-edition",
    "Sinking Funds Planner": "sinking-funds-planner-ai-edition",
    "Net Worth Tracker": "net-worth-tracker-ai-edition",
    "Investment Portfolio Tracker": "investment-portfolio-tracker-ai-edition",
    "Family & Education Planner": "family-education-planner-ai-edition",
}

PDFS_AI = {
    "Budget Tracker": "budget-tracker-ai-pdf.pdf",
    "Debt Payoff Planner": "debt-payoff-ai-pdf.pdf",
    "Sinking Funds Planner": "sinking-funds-ai-pdf.pdf",
    "Net Worth Tracker": "net-worth-ai-pdf.pdf",
    "Investment Portfolio Tracker": "investment-portfolio-ai-pdf.pdf",
    "Family & Education Planner": "family-education-ai-pdf.pdf",
}

PDFS_QS = {
    "Budget Tracker": "budget-tracker-quickstart.pdf",
    "Debt Payoff Planner": "debt-payoff-quickstart.pdf",
    "Sinking Funds Planner": "sinking-funds-quickstart.pdf",
    "Net Worth Tracker": "net-worth-quickstart.pdf",
    "Investment Portfolio Tracker": "investment-portfolio-quickstart.pdf",
    "Family & Education Planner": "family-education-quickstart.pdf",
}

print("="*78)
print("PART B — BUNDLE-LEVEL AUDIT")
print("="*78)

# B1: BRAND & VISUAL CONSISTENCY
print("\n[B1] Brand & Visual Consistency Across 6 SKUs")
print("-"*78)
sku_palette = {}
sku_fonts = {}
sku_tab_colors = {}
for sku, name in SKUS_AI.items():
    path = SCRATCH / f"{name}.xlsx"
    if not path.exists():
        continue
    wb = openpyxl.load_workbook(path, data_only=True)
    colors_in_sku = Counter()
    fonts_in_sku = Counter()
    tab_colors = {}
    for sh in wb.sheetnames:
        ws = wb[sh]
        tc = None
        try:
            tc_raw = getattr(ws.sheet_properties.tabColor, 'rgb', None) if ws.sheet_properties.tabColor else None
            tc = tc_raw
        except Exception:
            pass
        tab_colors[sh] = tc
        for row in ws.iter_rows():
            for c in row:
                try:
                    fill = c.fill
                    if fill and fill.fgColor and fill.fgColor.rgb and isinstance(fill.fgColor.rgb, str):
                        rgb = fill.fgColor.rgb
                        if rgb not in ('00000000', 'FFFFFFFF'):
                            colors_in_sku[rgb] += 1
                except Exception:
                    pass
                try:
                    if c.font and c.font.name:
                        fonts_in_sku[c.font.name] += 1
                except Exception:
                    pass
    sku_palette[sku] = colors_in_sku.most_common(8)
    sku_fonts[sku] = fonts_in_sku.most_common(5)
    sku_tab_colors[sku] = tab_colors

print("\nTop fonts per SKU:")
for sku, fonts in sku_fonts.items():
    print(f"  {sku:33s} -> {[f[0] for f in fonts[:3]]}")

print("\nTop 5 fill colors per SKU (hex, count):")
for sku, colors in sku_palette.items():
    pal = [(c[0], c[1]) for c in colors[:5]]
    print(f"  {sku:33s} -> {pal}")

# Compare cross-SKU
all_dominant_colors = set()
sku_dom_color_set = {}
for sku, colors in sku_palette.items():
    top3 = set(c[0] for c in colors[:3])
    sku_dom_color_set[sku] = top3
    all_dominant_colors |= top3

# Brand drift score
shared = set.intersection(*sku_dom_color_set.values()) if sku_dom_color_set else set()
print(f"\nSHARED dominant colors across all 6 SKUs: {shared} (n={len(shared)})")
if len(shared) == 0:
    print("  >>> BRAND-DRIFT: No common color across all 6 SKUs - bundle feels like 6 different products")

# Font cohesion
all_dominant_fonts = Counter()
for sku, fonts in sku_fonts.items():
    if fonts:
        all_dominant_fonts[fonts[0][0]] += 1
print(f"\nDominant-font choice across 6 SKUs: {dict(all_dominant_fonts)}")
if len(all_dominant_fonts) > 1:
    print(f"  >>> FONT-DRIFT: {len(all_dominant_fonts)} different dominant fonts across the 6 SKUs")

# B2: CROSS-SKU CONCEPT RECONCILIATION (mostly narrative; we'll flag known concerns)
print("\n[B2] Cross-SKU Concept Reconciliation")
print("-"*78)
print("Manual reconciliation concerns to verify in personas:")
print("  - 'Monthly savings' - Budget vs Sinking vs Net Worth")
print("  - 'Net worth' - NWT formula vs other SKU cross-refs")
print("  - 'Total debt' - DPP total vs NWT liabilities")
print("  - 'Investment portfolio total' - IPT total MV vs NWT investment line")
print("  - 'Required monthly contribution' - SFP vs F&E annuity convention")
print("  - Currency conventions (USD default vs EGP/multi-currency)")

# B3: DOCUMENTATION COVERAGE
print("\n[B3] Documentation Coverage")
print("-"*78)
docs = ["bundle-quickstart.pdf", "bundle-setup-wizard-finance.pdf", "bundle-setup-wizard-life.pdf",
        "bundle-ai-library-finance.pdf", "bundle-ai-library-life.pdf"]
doc_pres = []
for d in docs:
    p = PDFS / d
    if p.exists():
        try:
            rd = pypdf.PdfReader(str(p))
            doc_pres.append(f"  EXISTS: {d} ({len(rd.pages)} pp)")
        except Exception as e:
            doc_pres.append(f"  PRESENT but UNREADABLE: {d}: {e}")
    else:
        doc_pres.append(f"  MISSING: {d}")
for d in doc_pres:
    print(d)

# Cross-references in advisor PDFs - scan for sibling SKU mentions
print("\nCross-references in AI advisor PDFs:")
xref_grid = {}
sku_keywords = {
    "Budget Tracker": ["budget tracker", "budget planner"],
    "Debt Payoff Planner": ["debt payoff", "debt planner", "debt tracker"],
    "Sinking Funds Planner": ["sinking fund", "sinking-fund"],
    "Net Worth Tracker": ["net worth", "net-worth"],
    "Investment Portfolio Tracker": ["investment portfolio", "portfolio tracker"],
    "Family & Education Planner": ["family & education", "education planner", "529 plan"],
}
for owner_sku, advisor in PDFS_AI.items():
    path = PDFS / advisor
    if not path.exists():
        xref_grid[owner_sku] = "MISSING-PDF"
        continue
    try:
        rd = pypdf.PdfReader(str(path))
        full = ""
        for p in rd.pages[:min(len(rd.pages), 40)]:
            full += (p.extract_text() or "").lower() + " "
    except Exception as e:
        xref_grid[owner_sku] = f"READ-ERR {e}"
        continue
    mentions = {}
    for other_sku, kws in sku_keywords.items():
        if other_sku == owner_sku:
            continue
        hits = sum(full.count(k) for k in kws)
        mentions[other_sku] = hits
    xref_grid[owner_sku] = mentions

for owner, refs in xref_grid.items():
    if isinstance(refs, dict):
        total = sum(refs.values())
        mentioned = sum(1 for v in refs.values() if v > 0)
        print(f"  {owner:33s} -> mentions {mentioned}/5 siblings (total {total} hits)")
        if mentioned < 2:
            print(f"     >>> XREF-WEAK: only mentions {mentioned} siblings")
    else:
        print(f"  {owner:33s} -> {refs}")

# B4: DELIVERY INTEGRITY
print("\n[B4] Delivery Integrity")
print("-"*78)
file_inventory = []
total_size = 0
file_hashes = {}
for sku, name in SKUS_AI.items():
    path = SHEETS / f"{name}.xlsx"
    if path.exists():
        sz = path.stat().st_size
        total_size += sz
        h = hashlib.sha256(path.read_bytes()).hexdigest()[:12]
        file_inventory.append({"file": path.name, "sku": sku, "size": sz, "hash": h})
        if h in file_hashes:
            print(f"  >>> DUPLICATE-HASH: {path.name} matches {file_hashes[h]}")
        file_hashes[h] = path.name
for pdf in list(PDFS_AI.values()) + list(PDFS_QS.values()) + docs:
    path = PDFS / pdf
    if path.exists():
        sz = path.stat().st_size
        total_size += sz
        file_inventory.append({"file": path.name, "sku": "bundle/pdf", "size": sz})
# Etsy 20MB per file check
print(f"  Total files in scope: {len(file_inventory)}")
print(f"  Total uncompressed size: {total_size/1024/1024:.2f} MB")
oversize = [f for f in file_inventory if f['size'] > 20*1024*1024]
if oversize:
    print(f"  >>> ETSY-CAP-RISK: {len(oversize)} files >20MB")
else:
    print(f"  All individual files <20 MB (Etsy per-file cap OK)")

# Naming convention scan
naming_patterns = Counter()
for f in file_inventory:
    if f['file'].endswith('.xlsx'):
        # extract pattern: "{sku}-{tier}" or "{sku}-{tier}-v{n}"
        stem = f['file'].rsplit('.', 1)[0]
        if '-essentials' in stem:
            naming_patterns['essentials'] += 1
        if '-pro' in stem:
            naming_patterns['pro'] += 1
        if '-ai-edition' in stem:
            naming_patterns['ai-edition'] += 1
        if '-v2' in stem:
            naming_patterns['v2-suffix'] += 1
print(f"  Naming patterns observed: {dict(naming_patterns)}")
if naming_patterns.get('v2-suffix', 0) > 0:
    print(f"  >>> NAMING-DRIFT: budget-tracker-ai-edition-v2 has 'v2' suffix; other 5 SKUs are just 'ai-edition' (cosmetic)")

# B5: LISTING COPY CONSISTENCY (file presence only — no Etsy API call)
print("\n[B5] Listing Copy Consistency")
print("-"*78)
# look for listing copy files in repo
listings_dir = ROOT / "content/listings"
if listings_dir.exists():
    for x in listings_dir.glob("*.md"):
        print(f"  Found: {x.relative_to(ROOT)}")
else:
    print(f"  content/listings not found; checking docs/...")
    for x in ROOT.glob("docs/**/*listing*.md"):
        print(f"  Found: {x.relative_to(ROOT)}")

# B6: AI advisor voice audit
print("\n[B6] AI-Advisor-Voice Audit")
print("-"*78)
sample_paras = {}
voice_indicators = {}
for sku, advisor in PDFS_AI.items():
    p = PDFS / advisor
    if not p.exists():
        continue
    rd = pypdf.PdfReader(str(p))
    # get first 1500 chars of text from pages 2-3 (skip cover)
    pages = rd.pages
    text = ""
    for pg in pages[1:min(len(pages), 5)]:
        text += (pg.extract_text() or "")
    sample = text[:2000]
    sample_paras[sku] = sample
    # tally
    you_count = len(re.findall(r'\byou\b', sample, re.IGNORECASE))
    your_count = len(re.findall(r'\byour\b', sample, re.IGNORECASE))
    we_count = len(re.findall(r'\bwe\b', sample, re.IGNORECASE))
    us_count = len(re.findall(r'\bus\b', sample, re.IGNORECASE))
    avg_word = sum(len(w) for w in sample.split()) / max(len(sample.split()), 1)
    voice_indicators[sku] = {
        "you+your": you_count + your_count,
        "we+us": we_count + us_count,
        "avg_word_len": round(avg_word, 2),
        "sample_len": len(sample),
    }
print("\nVoice indicators (you/we ratio, avg word len):")
for sku, ind in voice_indicators.items():
    print(f"  {sku:33s} -> you+your={ind['you+your']:3d}, we+us={ind['we+us']:3d}, avg_word_len={ind['avg_word_len']:.2f}")

# voice drift
you_we_ratios = []
for sku, ind in voice_indicators.items():
    if ind['we+us'] == 0:
        you_we_ratios.append(99.0)
    else:
        you_we_ratios.append(ind['you+your'] / max(ind['we+us'], 1))
print(f"\n  you/we ratios: {[round(r,2) for r in you_we_ratios]}")
if max(you_we_ratios) - min(you_we_ratios) > 5 and len(set(round(r) for r in you_we_ratios)) >= 3:
    print("  >>> VOICE-DRIFT: Ratio variance suggests different writers/voices")
else:
    print("  Voice ratios reasonably consistent")

# Save B-part results
import json
out = ROOT / "tools/qa/round1/part-b-bundle-audit.json"
with out.open('w', encoding='utf-8') as fh:
    json.dump({
        "b1_palette": {k: v for k, v in sku_palette.items()},
        "b1_fonts": {k: v for k, v in sku_fonts.items()},
        "b1_shared_colors": list(shared),
        "b1_font_drift": dict(all_dominant_fonts),
        "b3_xref_grid": xref_grid,
        "b4_inventory": file_inventory,
        "b4_total_mb": total_size/1024/1024,
        "b6_voice": voice_indicators,
    }, fh, indent=2, default=str)
print(f"\nWrote {out}")
