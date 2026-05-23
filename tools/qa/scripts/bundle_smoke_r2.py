"""Round 2: smoke test on tools/qa/fixed/recalc/ files."""
import sys, io, json
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', line_buffering=True)
import openpyxl

ROOT = Path("C:/ETSY/etsy-store")
RECALC = ROOT / "tools/qa/fixed/recalc"

SKUS = {
    "budget-tracker": ["budget-tracker-ai-edition-v2", "budget-tracker-essentials", "budget-tracker-pro"],
    "debt-payoff-planner": ["debt-payoff-planner-ai-edition", "debt-payoff-planner-essentials", "debt-payoff-planner-pro"],
    "sinking-funds-planner": ["sinking-funds-planner-ai-edition", "sinking-funds-planner-essentials", "sinking-funds-planner-pro"],
    "net-worth-tracker": ["net-worth-tracker-ai-edition", "net-worth-tracker-essentials", "net-worth-tracker-pro"],
    "investment-portfolio-tracker": ["investment-portfolio-tracker-ai-edition", "investment-portfolio-tracker-essentials", "investment-portfolio-tracker-pro"],
    "family-education-planner": ["family-education-planner-ai-edition", "family-education-planner-essentials", "family-education-planner-pro"],
}
ERROR_TYPES = ['#REF!', '#N/A', '#DIV/0!', '#NAME?', '#VALUE!', '#NULL!', '#NUM!']

results = []
for sku, files in SKUS.items():
    for fname in files:
        path = RECALC / f"{fname}.xlsx"
        issues = []
        if not path.exists():
            issues.append("FILE-MISSING")
            results.append({"sku": sku, "file": fname, "issues": issues})
            continue
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as e:
            issues.append(f"LOAD-FAIL {e}")
            results.append({"sku": sku, "file": fname, "issues": issues})
            continue
        err_cells = []
        for sh in wb.sheetnames:
            ws = wb[sh]
            for r in ws.iter_rows():
                for c in r:
                    if isinstance(c.value, str) and c.value in ERROR_TYPES:
                        err_cells.append(f"{sh}!{c.coordinate}={c.value}")
        if err_cells:
            issues.append(f"FORMULA-ERRORS ({len(err_cells)}): " + "; ".join(err_cells[:3]))
        # Specific verification: IPT Scenario Simulator G2/I2 should NOT be #N/A
        if "investment-portfolio-tracker" in fname and "essentials" not in fname:
            try:
                ws = wb["🎯 Scenario Simulator"]
                g2 = ws['G2'].value
                i2 = ws['I2'].value
                if isinstance(g2, str) and "#N/A" in g2:
                    issues.append(f"IPT-G2-STILL-#N/A: {g2}")
                if isinstance(i2, str) and "#N/A" in i2:
                    issues.append(f"IPT-I2-STILL-#N/A: {i2}")
                # success state
                if g2 and not any(e in str(g2) for e in ERROR_TYPES):
                    issues.append(f"IPT-G2-FIXED: now = {str(g2)[:80]}")
                if i2 and not any(e in str(i2) for e in ERROR_TYPES):
                    issues.append(f"IPT-I2-FIXED: now = {str(i2)[:80]}")
            except KeyError:
                issues.append("Scenario Simulator tab not found")
        # Specific verification: NWT B30 should have BUNDLE NOTE
        if "net-worth-tracker" in fname:
            try:
                ws = wb["💼 Assets Summary"]
                b30 = ws['B30'].value
                if b30 and "BUNDLE NOTE" in str(b30):
                    issues.append(f"NWT-BUNDLE-NOTE-PRESENT: B30 has note")
                else:
                    issues.append(f"NWT-BUNDLE-NOTE-MISSING: B30 = {repr(b30)[:80]}")
                # B25 of FX
                ws = wb["⚙️ Settings & FX"]
                b25 = ws['B25'].value
                if b25 and "BUNDLE NOTE" in str(b25):
                    issues.append(f"NWT-FX-SYNC-PRESENT: B25 has note")
                else:
                    issues.append(f"NWT-FX-SYNC-MISSING")
            except KeyError as e:
                issues.append(f"NWT-TAB-MISSING: {e}")
        # IPT FX sync B26
        if "investment-portfolio-tracker" in fname and "essentials" not in fname:
            try:
                ws = wb["💵 Cash & FX Holdings"]
                b26 = ws['B26'].value
                if b26 and "BUNDLE NOTE" in str(b26):
                    issues.append(f"IPT-FX-SYNC-PRESENT: B26 has note")
                else:
                    issues.append(f"IPT-FX-SYNC-MISSING")
            except KeyError as e:
                issues.append(f"IPT-FX-TAB-MISSING: {e}")
        results.append({"sku": sku, "file": fname, "issues": issues})

# print summary
print("="*78)
print("ROUND 2 SMOKE TEST RESULTS")
print("="*78)
for r in results:
    print(f"\n{r['sku']:30s} / {r['file']}")
    if not r['issues']:
        print("  CLEAN")
    else:
        for iss in r['issues']:
            tag = "OK" if any(s in iss for s in ['FIXED', 'PRESENT']) else "ISSUE"
            print(f"  [{tag}] {iss}")

with (ROOT / "tools/qa/round2/part-a-smoke.json").open('w', encoding='utf-8') as fh:
    json.dump(results, fh, indent=2, default=str)
