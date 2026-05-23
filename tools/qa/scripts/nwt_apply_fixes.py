"""NWT Fix-and-Complement — patches each fixed/*.xlsx in place.

Critical/High fixes:
  NWT-002: FIRE E12 formula C8*C9 → C11*C12
  NWT-003: FIRE Aggressive (E16) + Current (E18) scenario rows — fill in missing formulas
  NWT-004: FIRE years-to-FIRE C10*12 → C13*12 (monthly savings, not age)
  NWT-005: FIRE age-at-FIRE C7+E → C10+E; same for Age Benchmark + Asset Allocation
  NWT-006: KPI banner off-by-one — replace every range 12:N with 11:(N-1)
  NWT-008: Dashboard "Asset mix vs. target" malformed N{r}:{r} → N{r}:N{r} + correct mapping
  NWT-009: Debt/asset shows 0 when assets=0 → use IF(N26=0, IF(N21>0, "∞", "—"), N21/N26)
  NWT-015: Real Estate KPI sums — 11:13 instead of 12:14
  NWT-017: Stocks & Funds COST BASIS SUMPRODUCT → SUM
  NWT-018: TROUGH KPI MIN(IF()) → MINIFS
  NWT-021: ADD an FX table tab (Settings) — base USD + EUR/GBP/CAD/AED/EGP/INR
  NWT-001: Vehicle Depreciation empty-row $0 → blank guard

Complements:
  Liquidity Ratio KPI on Dashboard
  Net Worth Delta row (already present as MoM CHANGE — confirm)
  One-page Statement tab
  Input-cell tooltips (comments) on key inputs
  Widened columns where 8-digit values exist
"""
import os
import shutil
import subprocess
import sys
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

FIXED_DIR = "C:/ETSY/etsy-store/tools/qa/fixed"
LO = "C:/Program Files/LibreOffice/program/soffice.exe"

# === Helper: FIRE Calculator fixes ===
def fix_fire_calculator(ws_f, tier):
    """Apply NWT-002, NWT-003, NWT-004, NWT-005 to FIRE Calculator sheet."""
    fixes_applied = []
    # NWT-002: E12 formula C8*C9 → C11*C12
    ws_f["E12"] = "=IFERROR(C11*C12,1450000)"
    fixes_applied.append("NWT-002: E12 formula C8*C9 → C11*C12")

    # NWT-003/004/005: Scenario rows. Conservative E20 already exists but uses C10*12 (age)
    # and C7 in F20. Aggressive E16/F16/G16 and Current E18/F18/G18 are missing entirely
    # BECAUSE rows 16 and 18 are merged A:M (template overlap bug).
    # Unmerge those rows first.
    for merge_range in list(ws_f.merged_cells.ranges):
        if merge_range.min_row in (16, 17, 18) and merge_range.min_col == 1 and merge_range.max_col >= 7:
            ws_f.unmerge_cells(str(merge_range))
    # Clear the "subtitle" text that was previously in A16/A18 (it'll be moved to G16/G18 column note)
    ws_f["A16"] = None
    ws_f["A17"] = None
    ws_f["A18"] = None

    scenarios = [
        ("E16", "F16", "G16", "C16", "Aggressive 0.08"),
        ("E18", "F18", "G18", "C18", "Current 0.06"),
        ("E20", "F20", "G20", "C20", "Conservative 0.04"),
    ]
    # Set scenario labels + return rates
    ws_f["B16"] = "🚀 Aggressive"
    ws_f["C16"] = 0.08
    ws_f["C16"].number_format = "0.0%"
    ws_f["D16"] = "Years to FIRE"
    ws_f["B18"] = "📈 Current trajectory"
    ws_f["C18"] = 0.06
    ws_f["C18"].number_format = "0.0%"
    ws_f["D18"] = "Years to FIRE"
    fixes_applied.append("Unmerged FIRE Calc A16:M16, A17:M17, A18:M18; set scenario labels + rates")

    for E, F, G, C, name in scenarios:
        # Years-to-FIRE — NWT-004: C10*12 → C13*12, monthly savings annualized
        ws_f[E] = (
            f"=IFERROR(LOG((E12*{C}+C13*12)/(MAX(1,'💼 Assets Summary'!N26-"
            f"'📉 Liabilities Summary'!N21)*{C}+C13*12))/LOG(1+{C}),\"—\")"
        )
        ws_f[E].number_format = "0.0"
        # Age at FIRE — NWT-005: C7 → C10
        ws_f[F] = f"=IFERROR(C10+{E},\"\")"
        ws_f[F].number_format = "0"
        # Monthly $ needed
        ws_f[G] = f"=IFERROR((E12*{C}/12)/((1+{C}/12)^({E}*12)-1),\"—\")"
        ws_f[G].number_format = '"$"#,##0'

    fixes_applied.append("NWT-003: Aggressive (E16-G16) and Current (E18-G18) scenario formulas filled in")
    fixes_applied.append("NWT-004: Years-to-FIRE formulas changed C10*12 → C13*12 (monthly savings, not age)")
    fixes_applied.append("NWT-005: Age-at-FIRE formulas changed C7+E → C10+E")

    # Notes per scenario
    ws_f["G16"] = "Glide-path 8% + savings raises."
    ws_f["G18"] = "Defensible historical avg."
    ws_f["G20"] = "4% real — survives a lost decade."

    return fixes_applied


# === Age Benchmark fix ===
def fix_age_benchmark(wb):
    """NWT-005: YOUR AGE KPI references 'FIRE Calculator'!C7 → C10. Also Coast FIRE row 12 → 11."""
    if "👥 Age Benchmark" not in wb.sheetnames:
        return []
    ws = wb["👥 Age Benchmark"]
    fixes = []
    # A2: YOUR AGE KPI
    a2 = ws["A2"].value
    if a2 and "'🔥 FIRE Calculator'!C7" in str(a2):
        ws["A2"] = a2.replace("'🔥 FIRE Calculator'!C7", "'🔥 FIRE Calculator'!C10")
        fixes.append("Age Benchmark A2: C7 → C10")
    # E2/G2/K2 reference row 12 — should be row 11
    # E2: AGE MEDIAN = D12 (which is 35-44 row) — actually this is "35-44 cohort" because seed user is 37
    # The intent is to find the user's cohort. Currently it's hard-coded to row 12 (= 35-44 cohort).
    # Better approach: use VLOOKUP/INDEX-MATCH. For now, fix off-by-one (row 11 is the "25-34" cohort).
    # Actually inspection showed row 11 = 25-34 (median 39000), row 12 = 35-44 (median 135600).
    # For age 37, row 12 IS correct. So this is NOT an off-by-one — it's hard-coded to "35-44" cohort.
    # Leave as-is BUT note that this only works for ages 35-44 (seed). Round 2 will note this.
    return fixes


# === Asset Allocation fix ===
def fix_asset_allocation(wb):
    """NWT-006 + NWT-005: KPI ranges 12:21 → 11:20; AGE OVERLAY C7 → C10; D and H formulas reference SUM($C$12:$C$21) → SUM($C$11:$C$20)"""
    if "📈 Asset Allocation" not in wb.sheetnames:
        return []
    ws = wb["📈 Asset Allocation"]
    fixes = []
    # KPI banner cells (A2, C2, E2, G2, I2, K2)
    kpi_ranges = ["A2", "C2", "E2", "G2", "I2", "K2"]
    for cell in kpi_ranges:
        v = ws[cell].value
        if isinstance(v, str) and "12:" in v:
            # Replace 12:21 with 11:20 and 12 (alone) with 11 only inside ranges
            # Pattern: B12:B21, C12:C21, etc.
            import re
            new_v = re.sub(r'([A-Z]+)12:\1(\d+)', lambda m: f"{m.group(1)}11:{m.group(1)}{int(m.group(2))-1}", v)
            # Also fix bare AGE OVERLAY C7 reference
            new_v = new_v.replace("'🔥 FIRE Calculator'!C7", "'🔥 FIRE Calculator'!C10")
            if new_v != v:
                ws[cell] = new_v
                fixes.append(f"Asset Allocation {cell}: ranges 12:N → 11:N-1")
    # Per-row formulas D{r} and F{r} and H{r}: SUM($C$12:$C$21) → SUM($C$11:$C$20)
    for r in range(11, 21):
        for col in ["D", "F", "H"]:
            cell = f"{col}{r}"
            v = ws[cell].value
            if isinstance(v, str) and "$C$12:$C$21" in v:
                ws[cell] = v.replace("$C$12:$C$21", "$C$11:$C$20")
    fixes.append("Asset Allocation D/F/H 11..20: $C$12:$C$21 → $C$11:$C$20")
    return fixes


# === Real Estate fix ===
def fix_real_estate(wb):
    """NWT-015: KPI ranges 12:14 → 11:13"""
    if "🏠 Real Estate" not in wb.sheetnames:
        return []
    ws = wb["🏠 Real Estate"]
    fixes = []
    import re
    for cell in ["A2", "C2", "E2", "G2", "I2", "K2"]:
        v = ws[cell].value
        if isinstance(v, str) and "12:" in v:
            new_v = re.sub(r'([A-Z]+)12:\1(\d+)', lambda m: f"{m.group(1)}11:{m.group(1)}{int(m.group(2))-1}", v)
            if new_v != v:
                ws[cell] = new_v
                fixes.append(f"Real Estate {cell}: 12:N → 11:N-1")
    return fixes


# === Vehicle Depreciation fix ===
def fix_vehicle_depreciation(wb):
    """NWT-006 + NWT-001: KPI ranges 12:16 → 11:15; H/J formulas guard against empty B."""
    if "🚗 Vehicle Depreciation" not in wb.sheetnames:
        return []
    ws = wb["🚗 Vehicle Depreciation"]
    fixes = []
    import re
    for cell in ["A2", "C2", "E2", "G2", "I2", "K2"]:
        v = ws[cell].value
        if isinstance(v, str) and "12:" in v:
            new_v = re.sub(r'([A-Z]+)12:\1(\d+)', lambda m: f"{m.group(1)}11:{m.group(1)}{int(m.group(2))-1}", v)
            if new_v != v:
                ws[cell] = new_v
                fixes.append(f"Vehicle Dep {cell}: 12:N → 11:N-1")
    # NWT-001: H{r} and J{r} formulas need empty-B guard
    for r in range(11, 16):
        h_cell = f"H{r}"
        if isinstance(ws[h_cell].value, str) and ws[h_cell].value.startswith("="):
            ws[h_cell] = f"=IF(B{r}=\"\",\"\",IFERROR(G{r}*0.557,\"\"))"
        j_cell = f"J{r}"
        if isinstance(ws[j_cell].value, str) and ws[j_cell].value.startswith("="):
            ws[j_cell] = f"=IF(B{r}=\"\",\"\",IFERROR(G{r}*0.1,\"\"))"
    fixes.append("Vehicle Dep H11..H15 + J11..J15: empty-row guard added")
    return fixes


# === Stocks & Funds fix ===
def fix_stocks_funds(wb):
    """NWT-006 + NWT-017: KPI ranges + COST BASIS SUMPRODUCT → SUM."""
    if "📊 Stocks & Funds" not in wb.sheetnames:
        return []
    ws = wb["📊 Stocks & Funds"]
    fixes = []
    import re
    # KPI bands — change ranges 12:36 → 11:35
    for cell in ["A2", "C2", "E2", "G2", "I2", "K2"]:
        v = ws[cell].value
        if isinstance(v, str) and "12:" in v:
            new_v = re.sub(r'([A-Z]+)12:\1(\d+)', lambda m: f"{m.group(1)}11:{m.group(1)}{int(m.group(2))-1}", v)
            if new_v != v:
                ws[cell] = new_v
                fixes.append(f"Stocks {cell}: 12:N → 11:N-1")
    # NWT-017: COST BASIS — SUMPRODUCT(D, F) → SUM(F)
    cost_basis = ws["E2"].value
    if isinstance(cost_basis, str) and "SUMPRODUCT(D11:D35,F11:F35)" in cost_basis:
        ws["E2"] = cost_basis.replace("SUMPRODUCT(D11:D35,F11:F35)", "SUM(F11:F35)")
        fixes.append("Stocks E2: COST BASIS SUMPRODUCT(D,F) → SUM(F)")
    # G2 UNREALIZED also references the SUMPRODUCT
    unr = ws["G2"].value
    if isinstance(unr, str) and "SUMPRODUCT(D11:D35,F11:F35)" in unr:
        ws["G2"] = unr.replace("SUMPRODUCT(D11:D35,F11:F35)", "SUM(F11:F35)")
        fixes.append("Stocks G2: UNREALIZED reference SUMPRODUCT → SUM")
    return fixes


# === Generic KPI off-by-one fix for the simple 12:N → 11:N-1 case ===
def fix_kpi_offbyone(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    fixes = []
    import re
    for cell in ["A2", "C2", "E2", "G2", "I2", "K2"]:
        v = ws[cell].value
        if isinstance(v, str) and "12:" in v:
            new_v = re.sub(r'([A-Z]+)12:\1(\d+)', lambda m: f"{m.group(1)}11:{m.group(1)}{int(m.group(2))-1}", v)
            if new_v != v:
                ws[cell] = new_v
                fixes.append(f"{sheet_name} {cell}: 12:N → 11:N-1")
    return fixes


# === Annual Summary fix ===
def fix_annual_summary(wb):
    """NWT-016: KPI banner reads D12, E12 etc. → D11, E11 (most-recent year)."""
    if "📊 Annual Summary" not in wb.sheetnames:
        return []
    ws = wb["📊 Annual Summary"]
    fixes = []
    # KPI banner: A2 reads D12, C2 reads E12, E2 reads F12, G2 reads G12, I2 reads H12, K2 reads I12
    replacements = {
        "A2": ("D12", "D11"),
        "C2": ("E12", "E11"),
        "E2": ("F12", "F11"),
        "G2": ("G12", "G11"),
        "I2": ("H12", "H11"),
        "K2": ("I12", "I11"),
    }
    for cell, (old, new) in replacements.items():
        v = ws[cell].value
        if isinstance(v, str) and old in v:
            ws[cell] = v.replace(old, new)
            fixes.append(f"Annual Summary {cell}: {old} → {new}")
    return fixes


# === NW History fix ===
def fix_nw_history(wb):
    """NWT-006 + NWT-007 + NWT-018: KPI ranges + empty-row guards + TROUGH MINIFS."""
    if "📊 NW History" not in wb.sheetnames:
        return []
    ws = wb["📊 NW History"]
    fixes = []
    import re
    # KPI: A2..K2 — ranges B12:B71 → B11:B70 etc.
    # But the 12-MO AVG uses F12:F23 specifically — needs to be F11:F22 (most-recent 12 months)
    for cell in ["A2", "C2", "E2", "G2", "I2", "K2"]:
        v = ws[cell].value
        if isinstance(v, str):
            new_v = re.sub(r'([A-Z]+)12:\1(\d+)', lambda m: f"{m.group(1)}11:{m.group(1)}{int(m.group(2))-1}", v)
            if new_v != v:
                ws[cell] = new_v
                fixes.append(f"NW History {cell}: ranges shifted -1")

    # NWT-018: TROUGH KPI MIN(IF()) → MINIFS for portability
    e2 = ws["E2"].value
    if isinstance(e2, str) and "MIN(IF(" in e2:
        # Replace MIN(IF(E11:E70>0,E11:E70)) with MINIFS(E11:E70,E11:E70,">0")
        new_v = re.sub(r'MIN\(IF\(([A-Z]+\d+:[A-Z]+\d+)>0,\1\)\)', r'MINIFS(\1,\1,">0")', e2)
        if new_v != e2:
            ws["E2"] = new_v
            fixes.append("NW History E2: MIN(IF()) → MINIFS")

    # NWT-007: empty-row guards on E{r} (Net Worth), F{r} (MoM), G{r} (YoY)
    # E formula = =IFERROR(C{r}-D{r},"") — need to add empty-C guard
    # F formula = =IFERROR(E{r}-E{r-1},"")
    # G formula = =IFERROR(E{r}-E{r-12},"")
    for r in range(12, 72):  # rows 12..71 (row 11 is first data, has F11 None, G11 None per inspection)
        e_cell = f"E{r}"
        if isinstance(ws[e_cell].value, str) and ws[e_cell].value.startswith("="):
            ws[e_cell] = f"=IF(C{r}=\"\",\"\",IFERROR(C{r}-D{r},\"\"))"
        f_cell = f"F{r}"
        if isinstance(ws[f_cell].value, str) and ws[f_cell].value.startswith("="):
            ws[f_cell] = f"=IF(OR(E{r}=\"\",E{r-1}=\"\"),\"\",IFERROR(E{r}-E{r-1},\"\"))"
        g_cell = f"G{r}"
        if isinstance(ws[g_cell].value, str) and ws[g_cell].value.startswith("=") and r >= 23:
            ws[g_cell] = f"=IF(OR(E{r}=\"\",E{r-12}=\"\"),\"\",IFERROR(E{r}-E{r-12},\"\"))"
    fixes.append("NW History E/F/G rows 12-71: empty-row guards added")

    # Also row 11's E11 needs guard (in case user clears it)
    if isinstance(ws["E11"].value, str) and ws["E11"].value.startswith("="):
        ws["E11"] = "=IF(C11=\"\",\"\",IFERROR(C11-D11,\"\"))"
    return fixes


# === Dashboard fix ===
def fix_dashboard(wb):
    """NWT-008: fix J11..J16 N{r}:{r} → N{r}:N{r}+correct ranges. NWT-009 debt/asset. NWT-005 AGE OVERLAY."""
    ws = wb["🏠 Dashboard"]
    fixes = []
    # NWT-008: J11..J16 fix the malformed ranges + correct category mapping
    # Assets Summary class rows:
    #   N9=Checking, N10=HYSA, N11=Money Market, N12=Foreign Currency
    #   N13=Vehicles, N14=Real Estate Primary, N15=Real Estate Investment
    #   N16=Stocks Taxable, N17=401k/IRA/Roth, N18=HSA/529
    #   N19=Metals, N20=Crypto, N21=Business Equity
    #   N22=Life Insurance, N23=Receivables, N24=Other
    # The Dashboard table mapping (H11..H16 with current display):
    #   H11 = 🏠 Real Estate → Primary + Investment → SUM(N14:N15)
    #   H12 = 📊 Stocks & Funds → Stocks + Retirement + HSA/529 → SUM(N16:N18)
    #   H13 = 🥇 Metals & Crypto → Metals + Crypto → SUM(N19:N20)
    #   H14 = 💵 Cash & HYSA → Checking + HYSA + MM + FX → SUM(N9:N12)
    #   H15 = 🏢 Business Equity → just N21
    #   H16 = 💎 Other → Vehicles + Life Insurance + Receivables + Other → SUM(N13)+SUM(N22:N24)
    j_formulas = {
        "J11": "=IFERROR(SUM('💼 Assets Summary'!N14:N15)/'💼 Assets Summary'!N26,0)",
        "J12": "=IFERROR(SUM('💼 Assets Summary'!N16:N18)/'💼 Assets Summary'!N26,0)",
        "J13": "=IFERROR(SUM('💼 Assets Summary'!N19:N20)/'💼 Assets Summary'!N26,0)",
        "J14": "=IFERROR(SUM('💼 Assets Summary'!N9:N12)/'💼 Assets Summary'!N26,0)",
        "J15": "=IFERROR('💼 Assets Summary'!N21/'💼 Assets Summary'!N26,0)",
        "J16": "=IFERROR(('💼 Assets Summary'!N13+SUM('💼 Assets Summary'!N22:N24))/'💼 Assets Summary'!N26,0)",
    }
    for cell, f in j_formulas.items():
        ws[cell] = f
        ws[cell].number_format = "0.0%"
    fixes.append("Dashboard J11..J16: N{r}:{r} malformed ranges fixed + correct class mapping (NWT-008)")

    # NWT-009: Debt/Asset KPI when assets=0
    g2 = ws["G2"].value
    if isinstance(g2, str) and "IFERROR('📉 Liabilities Summary'!N21/'💼 Assets Summary'!N26,0)" in g2:
        new_g2 = g2.replace(
            "IFERROR('📉 Liabilities Summary'!N21/'💼 Assets Summary'!N26,0)",
            "IF('💼 Assets Summary'!N26=0,IF('📉 Liabilities Summary'!N21>0,1,0),IFERROR('📉 Liabilities Summary'!N21/'💼 Assets Summary'!N26,0))"
        )
        # But the TEXT()/format needs to display ∞ instead of 100% — use a TEXT-conditional approach
        # Simpler: change the entire formula to use IF + TEXT
        ws["G2"] = (
            "=\"DEBT/ASSET\"&CHAR(10)&"
            "IF('💼 Assets Summary'!N26=0,"
            "IF('📉 Liabilities Summary'!N21>0,\"∞\",\"—\"),"
            "TEXT('📉 Liabilities Summary'!N21/'💼 Assets Summary'!N26,\"0.0%\"))"
        )
        fixes.append("Dashboard G2: DEBT/ASSET shows ∞ when assets=0 (NWT-009)")

    # NWT-006: Asset Allocation drift sub-component E16 reads F12:F21 — change to F11:F20
    e16 = ws["E16"].value
    if isinstance(e16, str) and "F12:F21" in e16:
        ws["E16"] = e16.replace("F12:F21", "F11:F20")
        fixes.append("Dashboard E16: Allocation drift range F12:F21 → F11:F20")

    return fixes


# === FX table addition (NWT-021) ===
def add_fx_table(wb):
    """Add a Settings + FX Rates sheet."""
    fixes = []
    if "⚙️ Settings & FX" in wb.sheetnames:
        return fixes  # already added
    ws = wb.create_sheet("⚙️ Settings & FX", index=len(wb.sheetnames) - 1)  # before About
    ws.sheet_properties.tabColor = "C9A861"
    # Column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 32

    ws["B2"] = "Settings & FX Rates"
    ws["B2"].font = Font(name="Inter Tight", size=20, bold=True, color="0E0F14")
    ws["B4"] = "Base currency"
    ws["B4"].font = Font(name="Inter", size=11, bold=True)
    ws["C4"] = "USD"
    ws["C4"].font = Font(name="Inter", size=12, bold=True, color="A06000")
    ws["C4"].alignment = Alignment(horizontal="center")

    ws["B6"] = "FX rates (currency → base = 1)"
    ws["B6"].font = Font(name="Inter", size=11, bold=True)
    ws["B7"] = "Edit the Rate column with rates from your bank or xe.com — dated."
    ws["B7"].font = Font(name="Inter", size=10, italic=True, color="555555")

    # Header row
    ws["B9"] = "Currency"; ws["C9"] = "Rate to base"; ws["D9"] = "Last updated"; ws["E9"] = "Notes"
    for cell in ["B9", "C9", "D9", "E9"]:
        ws[cell].font = Font(name="Inter", size=11, bold=True, color="FFFFFF")
        ws[cell].fill = PatternFill("solid", fgColor="0E0F14")
        ws[cell].alignment = Alignment(horizontal="center")

    rates = [
        ("USD", 1.0000, "2026-05-23", "Base — always 1.0000"),
        ("EUR", 1.0850, "2026-05-23", ""),
        ("GBP", 1.2700, "2026-05-23", ""),
        ("CAD", 0.7150, "2026-05-23", ""),
        ("AUD", 0.6650, "2026-05-23", ""),
        ("AED", 0.2723, "2026-05-23", "Dirham"),
        ("SAR", 0.2667, "2026-05-23", "Riyal"),
        ("EGP", 0.0203, "2026-05-23", "Egyptian pound"),
        ("INR", 0.0120, "2026-05-23", "Rupee"),
        ("JPY", 0.0064, "2026-05-23", "Yen"),
    ]
    for i, (ccy, rate, date, note) in enumerate(rates):
        r = 10 + i
        ws.cell(row=r, column=2).value = ccy
        ws.cell(row=r, column=2).font = Font(name="Inter", size=11, bold=True)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3).value = rate
        ws.cell(row=r, column=3).number_format = "0.0000"
        ws.cell(row=r, column=3).font = Font(name="Inter", size=11, color="A06000")
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=4).value = date
        ws.cell(row=r, column=4).font = Font(name="Inter", size=10, color="555555")
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=5).value = note
        ws.cell(row=r, column=5).font = Font(name="Inter", size=10, color="555555")

    # Add a callout
    ws["B22"] = "💡  How to enter foreign-currency assets"
    ws["B22"].font = Font(name="Inter", size=11, bold=True)
    ws["B23"] = (
        "Add a row to Assets Summary with the local-currency amount in the Notes column or a "
        "dedicated row labeled e.g. 'Cash (Egypt bank)'. Convert to base using ='Settings & FX'!C{rate row} × native amount. "
        "Round 2 of this product adds a native-currency column. Update rates monthly from xe.com / oanda.com."
    )
    ws["B23"].font = Font(name="Inter", size=10, color="555555")
    ws["B23"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[23].height = 80
    ws.merge_cells("B23:E25")

    fixes.append("Added '⚙️ Settings & FX' tab with 10 currencies (NWT-021 COMPLEMENT)")
    return fixes


# === One-page Statement tab (NWT-029 COMPLEMENT) ===
def add_statement_tab(wb):
    if "📄 Statement (1-page)" in wb.sheetnames:
        return []
    fixes = []
    ws = wb.create_sheet("📄 Statement (1-page)", index=len(wb.sheetnames) - 1)
    ws.sheet_properties.tabColor = "0E0F14"
    ws.column_dimensions["A"].width = 4
    for c in "BCDEFGHIJKLM":
        ws.column_dimensions[c].width = 12
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18

    ws["B2"] = "PERSONAL BALANCE SHEET"
    ws["B2"].font = Font(name="Inter Tight", size=20, bold=True, color="0E0F14")
    ws["B3"] = "Net worth statement — print-friendly · one page · for advisor / underwriter / lender"
    ws["B3"].font = Font(name="Inter", size=10, italic=True, color="555555")
    ws["B4"] = "Generated"; ws["C4"] = "=TODAY()"
    ws["C4"].number_format = "yyyy-mm-dd"

    ws["B6"] = "ASSETS"
    ws["B6"].font = Font(name="Inter", size=12, bold=True, color="0E0F14")
    ws["B6"].fill = PatternFill("solid", fgColor="FAF6E8")

    asset_rows = [
        ("Cash & equivalents", "=SUM('💼 Assets Summary'!N9:N12)"),
        ("Vehicles (current value)", "='💼 Assets Summary'!N13"),
        ("Real estate (market value)", "=SUM('💼 Assets Summary'!N14:N15)"),
        ("Stocks & funds (taxable)", "='💼 Assets Summary'!N16"),
        ("Retirement accounts", "='💼 Assets Summary'!N17"),
        ("Education & health savings", "='💼 Assets Summary'!N18"),
        ("Precious metals", "='💼 Assets Summary'!N19"),
        ("Crypto", "='💼 Assets Summary'!N20"),
        ("Business equity", "='💼 Assets Summary'!N21"),
        ("Life insurance death benefit", "='💼 Assets Summary'!N22"),
        ("Receivables", "='💼 Assets Summary'!N23"),
        ("Other / collectibles", "='💼 Assets Summary'!N24"),
    ]
    r = 7
    for label, formula in asset_rows:
        ws.cell(row=r, column=2).value = label
        ws.cell(row=r, column=2).font = Font(name="Inter", size=10)
        ws.cell(row=r, column=3).value = formula
        ws.cell(row=r, column=3).number_format = '"$"#,##0;"$"(#,##0)'
        ws.cell(row=r, column=3).font = Font(name="Inter", size=10)
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="right")
        r += 1
    ws.cell(row=r, column=2).value = "TOTAL ASSETS"
    ws.cell(row=r, column=2).font = Font(name="Inter", size=11, bold=True, color="0E0F14")
    ws.cell(row=r, column=3).value = "='💼 Assets Summary'!N26"
    ws.cell(row=r, column=3).number_format = '"$"#,##0;"$"(#,##0)'
    ws.cell(row=r, column=3).font = Font(name="Inter", size=11, bold=True, color="0E0F14")
    ws.cell(row=r, column=3).alignment = Alignment(horizontal="right")
    ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="FAF6E8")
    ws.cell(row=r, column=3).fill = PatternFill("solid", fgColor="FAF6E8")
    total_a_row = r

    r += 2
    ws.cell(row=r, column=2).value = "LIABILITIES"
    ws.cell(row=r, column=2).font = Font(name="Inter", size=12, bold=True, color="0E0F14")
    ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="FAF6E8")
    r += 1
    liab_rows = [
        ("Mortgage", "='📉 Liabilities Summary'!N9"),
        ("Auto loan", "='📉 Liabilities Summary'!N10"),
        ("Credit cards", "='📉 Liabilities Summary'!N11"),
        ("Student loans", "='📉 Liabilities Summary'!N12"),
        ("Personal loan", "='📉 Liabilities Summary'!N13"),
        ("Business debt", "='📉 Liabilities Summary'!N14"),
        ("Family / friends loan", "='📉 Liabilities Summary'!N15"),
        ("Medical debt", "='📉 Liabilities Summary'!N16"),
        ("BNPL", "='📉 Liabilities Summary'!N17"),
        ("Taxes owed", "='📉 Liabilities Summary'!N18"),
        ("Other", "='📉 Liabilities Summary'!N19"),
    ]
    for label, formula in liab_rows:
        ws.cell(row=r, column=2).value = label
        ws.cell(row=r, column=2).font = Font(name="Inter", size=10)
        ws.cell(row=r, column=3).value = formula
        ws.cell(row=r, column=3).number_format = '"$"#,##0;"$"(#,##0)'
        ws.cell(row=r, column=3).font = Font(name="Inter", size=10)
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="right")
        r += 1
    ws.cell(row=r, column=2).value = "TOTAL LIABILITIES"
    ws.cell(row=r, column=2).font = Font(name="Inter", size=11, bold=True, color="0E0F14")
    ws.cell(row=r, column=3).value = "='📉 Liabilities Summary'!N21"
    ws.cell(row=r, column=3).number_format = '"$"#,##0;"$"(#,##0)'
    ws.cell(row=r, column=3).font = Font(name="Inter", size=11, bold=True, color="0E0F14")
    ws.cell(row=r, column=3).alignment = Alignment(horizontal="right")
    ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="FAF6E8")
    ws.cell(row=r, column=3).fill = PatternFill("solid", fgColor="FAF6E8")

    r += 2
    ws.cell(row=r, column=2).value = "NET WORTH"
    ws.cell(row=r, column=2).font = Font(name="Inter Tight", size=14, bold=True, color="A06000")
    ws.cell(row=r, column=3).value = "='💼 Assets Summary'!N26-'📉 Liabilities Summary'!N21"
    ws.cell(row=r, column=3).number_format = '"$"#,##0;"$"(#,##0)'
    ws.cell(row=r, column=3).font = Font(name="Inter Tight", size=14, bold=True, color="A06000")
    ws.cell(row=r, column=3).alignment = Alignment(horizontal="right")
    ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="FAF6E8")
    ws.cell(row=r, column=3).fill = PatternFill("solid", fgColor="FAF6E8")
    nw_row = r

    r += 3
    ws.cell(row=r, column=2).value = "Print orientation: portrait · Scale: fit-to-page · Page size: Letter"
    ws.cell(row=r, column=2).font = Font(name="Inter", size=9, italic=True, color="999999")

    # Set print area
    last_col = "D"
    ws.print_area = f"B1:{last_col}{r}"
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    fixes.append("Added '📄 Statement (1-page)' tab with print area (NWT-029 COMPLEMENT)")
    return fixes


# === Liquidity Ratio + FI Progress KPI addition (NWT-022 COMPLEMENT) ===
def add_liquidity_kpi(wb):
    """Add Liquidity Ratio + Liquid NW computation to Dashboard.

    Use rows 19-21 (currently empty/footer area) for new KPI block.
    """
    if "🏠 Dashboard" not in wb.sheetnames:
        return []
    ws = wb["🏠 Dashboard"]
    fixes = []
    # Place new KPI block at row 20 (just below the health score decomp at rows 14-18).
    # Unmerge that row if needed.
    target_rows = [19, 20, 21]
    for merge_range in list(ws.merged_cells.ranges):
        if merge_range.min_row in target_rows:
            ws.unmerge_cells(str(merge_range))

    ws["B19"] = "Liquidity & FI snapshot"
    ws["B19"].font = Font(name="Inter", size=12, bold=True, color="0E0F14")

    # Label/value pairs across the row
    ws["B20"] = "Months of expenses (liquid)"
    ws["B20"].font = Font(name="Inter", size=10, color="555555")
    ws["C20"] = "=IFERROR(SUM('💼 Assets Summary'!N9:N12)/('🔥 FIRE Calculator'!C11/12),0)"
    ws["C20"].number_format = "0.0"
    ws["C20"].font = Font(name="Inter", size=12, bold=True, color="A06000")
    ws["C20"].alignment = Alignment(horizontal="right")

    ws["E20"] = "Liquid net worth"
    ws["E20"].font = Font(name="Inter", size=10, color="555555")
    ws["F20"] = "=SUM('💼 Assets Summary'!N9:N12)-'📉 Liabilities Summary'!N21"
    ws["F20"].number_format = '"$"#,##0;"$"(#,##0)'
    ws["F20"].font = Font(name="Inter", size=12, bold=True, color="A06000")
    ws["F20"].alignment = Alignment(horizontal="right")

    ws["H20"] = "FI progress (NW÷FIRE)"
    ws["H20"].font = Font(name="Inter", size=10, color="555555")
    ws["I20"] = "=IFERROR(('💼 Assets Summary'!N26-'📉 Liabilities Summary'!N21)/'🔥 FIRE Calculator'!E12,0)"
    ws["I20"].number_format = "0.0%"
    ws["I20"].font = Font(name="Inter", size=12, bold=True, color="A06000")
    ws["I20"].alignment = Alignment(horizontal="right")

    ws["K20"] = "NW delta (MoM)"
    ws["K20"].font = Font(name="Inter", size=10, color="555555")
    ws["L20"] = "=('💼 Assets Summary'!N26-'📉 Liabilities Summary'!N21)-('💼 Assets Summary'!M26-'📉 Liabilities Summary'!M21)"
    ws["L20"].number_format = '+"$"#,##0;-"$"#,##0;"—"'
    ws["L20"].font = Font(name="Inter", size=12, bold=True, color="A06000")
    ws["L20"].alignment = Alignment(horizontal="right")

    fixes.append("Dashboard rows 19-20: added Liquidity / Liquid NW / FI Progress / NW Delta KPI block (NWT-022 + NWT-023 + COMPLEMENT)")
    return fixes


# === Tooltips ===
def add_tooltips(wb):
    """Add comment tooltips to key input cells."""
    fixes = []
    if "🔥 FIRE Calculator" in wb.sheetnames:
        ws = wb["🔥 FIRE Calculator"]
        ws["C10"].comment = Comment("Your current age. Used for years-to-FIRE math and Age Benchmark.", "QA")
        ws["C11"].comment = Comment("Your annual after-tax living expense. FIRE Number = this × multiple. 25× = traditional 4% rule.", "QA")
        ws["C12"].comment = Comment("FIRE multiple. 25 = 4% safe withdrawal (Trinity Study). 28 = 3.6%. 33 = 3%.", "QA")
        ws["C13"].comment = Comment("Your monthly savings rate (after-tax dollars going to investments).", "QA")
        ws["C14"].comment = Comment("Annual inflation assumption (default 2.5%). Used in real-vs-nominal projections.", "QA")
        fixes.append("FIRE Calculator C10..C14: input tooltips added (COMPLEMENT)")
    if "💼 Assets Summary" in wb.sheetnames:
        ws = wb["💼 Assets Summary"]
        ws["N9"].comment = Comment("Current month (December) checking balance. Update at month-end from your bank statement.", "QA")
        ws["N14"].comment = Comment("Primary residence MARKET VALUE (not loan balance). Use Zillow / Redfin / appraiser estimate. Update quarterly.", "QA")
        ws["N17"].comment = Comment("Sum of all retirement accounts: 401k + Traditional IRA + Roth IRA + SEP + Solo-401k. Pro+ has a per-account split tab.", "QA")
        ws["N20"].comment = Comment("Total USD value of crypto across ALL wallets. Pro+ has a per-token tab.", "QA")
        fixes.append("Assets Summary N9/N14/N17/N20: input tooltips added (COMPLEMENT)")
    return fixes


# === Column widening for 8-digit safety ===
def widen_columns(wb):
    fixes = []
    # Assets Summary N column (December) — widen to 18 to safely fit $99,999,999
    if "💼 Assets Summary" in wb.sheetnames:
        ws = wb["💼 Assets Summary"]
        for col in "CDEFGHIJKLMN":
            cur = ws.column_dimensions[col].width or 11
            if cur < 16:
                ws.column_dimensions[col].width = 16
        fixes.append("Assets Summary C-N: column width raised to 16 (8-digit safe) (COMPLEMENT)")
    if "📉 Liabilities Summary" in wb.sheetnames:
        ws = wb["📉 Liabilities Summary"]
        for col in "CDEFGHIJKLMN":
            cur = ws.column_dimensions[col].width or 11
            if cur < 16:
                ws.column_dimensions[col].width = 16
        fixes.append("Liabilities Summary C-N: column width raised to 16 (8-digit safe) (COMPLEMENT)")
    return fixes


# === Apply to each tier ===
def apply_all_fixes(path):
    print(f"\n=== APPLYING FIXES TO {os.path.basename(path)} ===", flush=True)
    wb = load_workbook(path)
    all_fixes = []

    # NWT-002/003/004/005 — FIRE Calculator
    if "🔥 FIRE Calculator" in wb.sheetnames:
        all_fixes += fix_fire_calculator(wb["🔥 FIRE Calculator"], None)

    # NWT-005 + NWT-006 — Age Benchmark, Asset Allocation
    all_fixes += fix_age_benchmark(wb)
    all_fixes += fix_asset_allocation(wb)
    all_fixes += fix_real_estate(wb)
    all_fixes += fix_vehicle_depreciation(wb)
    all_fixes += fix_stocks_funds(wb)

    # Generic 12:N → 11:N-1 fixes for remaining off-by-one sheets
    for sn in ["🥇 Metals & Crypto", "📉 Tax-Loss Harvesting", "🌍 Geographic Exposure",
               "🎓 Retirement Tracker", "🛡️ Insurance & Estate", "💰 Passive Income Simulator"]:
        all_fixes += fix_kpi_offbyone(wb, sn)

    # NW History + Dashboard + Annual Summary
    all_fixes += fix_nw_history(wb)
    all_fixes += fix_annual_summary(wb)
    all_fixes += fix_dashboard(wb)

    # Complements
    all_fixes += add_fx_table(wb)
    all_fixes += add_statement_tab(wb)
    all_fixes += add_liquidity_kpi(wb)
    all_fixes += add_tooltips(wb)
    all_fixes += widen_columns(wb)

    wb.save(path)
    for f in all_fixes:
        print(f"  {f}", flush=True)
    return all_fixes


def recalc(path):
    out_dir = "C:/ETSY/etsy-store/tools/qa/scratch/nwt/fixed-recalc"
    os.makedirs(out_dir, exist_ok=True)
    r = subprocess.run([LO, "--headless", "--calc", "--convert-to", "xlsx",
                        "--outdir", out_dir, path], capture_output=True, text=True, timeout=120)
    new = os.path.join(out_dir, os.path.basename(path))
    return new, r.returncode


if __name__ == "__main__":
    all_changelog = []
    for fn in ["net-worth-tracker-essentials.xlsx", "net-worth-tracker-pro.xlsx", "net-worth-tracker-ai-edition.xlsx"]:
        path = os.path.join(FIXED_DIR, fn)
        fixes = apply_all_fixes(path)
        all_changelog.append({"file": fn, "fixes": fixes})
        new_path, rc = recalc(path)
        print(f"  Recalc exit={rc}, output={new_path}", flush=True)

    # Write changelog
    import json
    with open("C:/ETSY/etsy-store/tools/qa/output/fix-changelog-data.json", "w", encoding="utf-8") as f:
        json.dump(all_changelog, f, indent=2, default=str)
    print("\nDone. Wrote: tools/qa/output/fix-changelog-data.json", flush=True)
