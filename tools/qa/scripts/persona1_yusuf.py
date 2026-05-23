"""Persona 1 — Yusuf, 28, early career, EGP — Budget + Debt + NWT.
Verifies cross-SKU reconciliation: debt total in NWT = debt total in DPP, budget debt-min = DPP min.
"""
import sys, io, subprocess, os, shutil
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', line_buffering=True)
import openpyxl

ROOT = Path("C:/ETSY/etsy-store")
SCRATCH = ROOT / "tools/qa/scratch"
SHEETS = ROOT / "tools/sheets-gen/output"
SOFFICE = "C:/Program Files/LibreOffice/program/soffice.exe"

# Setup: copy source xlsx → p1 scratch files
P1_BUDGET = SCRATCH / "r1-p1-yusuf-budget.xlsx"
P1_DEBT = SCRATCH / "r1-p1-yusuf-debt.xlsx"
P1_NWT = SCRATCH / "r1-p1-yusuf-nwt.xlsx"
shutil.copy(SHEETS / "budget-tracker-ai-edition-v2.xlsx", P1_BUDGET)
shutil.copy(SHEETS / "debt-payoff-planner-ai-edition.xlsx", P1_DEBT)
shutil.copy(SHEETS / "net-worth-tracker-ai-edition.xlsx", P1_NWT)

print("="*78)
print("PERSONA 1 — Yusuf (28, EGP, early career) — Budget + Debt + NWT")
print("="*78)

# Reference values (from persona spec)
REF = {
    "net_cashflow": 28000 - 18500 - 4200,  # 5,300
    "savings_rate": 5300 / 28000,  # 18.93%
    "total_debt": 220000 + 35000 + 95000,  # 350,000
    "total_min": 4200,
    "total_assets": 23000 + 25000 + 180000,  # 228,000
    "net_worth": 228000 - 350000,  # -122,000
}
print(f"\nReference values:")
print(f"  Net cashflow:    EGP {REF['net_cashflow']:,}")
print(f"  Savings rate:    {REF['savings_rate']*100:.2f}%")
print(f"  Total debt:      EGP {REF['total_debt']:,}")
print(f"  Total monthly min: EGP {REF['total_min']:,}")
print(f"  Total assets:    EGP {REF['total_assets']:,}")
print(f"  Net worth:       EGP {REF['net_worth']:,}")

# Inspect each workbook for input cells we need to write to
def show_input_layout(path, label, tabs_of_interest):
    print(f"\n--- {label}: input cells ---")
    wb = openpyxl.load_workbook(path, data_only=False)
    for tab in tabs_of_interest:
        # find the most-likely input tab
        matches = [s for s in wb.sheetnames if tab.lower() in s.lower()]
        if not matches:
            print(f"  no tab matching '{tab}'")
            continue
        for sh in matches[:1]:
            ws = wb[sh]
            print(f"  Tab '{sh}': max_row={ws.max_row}, max_col={ws.max_column}")
            # show first 20 rows of cols A-G
            for r in range(1, min(ws.max_row+1, 22)):
                vals = []
                for cc in range(1, min(ws.max_column+1, 8)):
                    v = ws.cell(row=r, column=cc).value
                    if v is None: continue
                    vals.append(f"{openpyxl.utils.get_column_letter(cc)}{r}={str(v)[:30]}")
                if vals:
                    print(f"    row {r}: {' | '.join(vals[:5])}")

show_input_layout(P1_BUDGET, "BUDGET ai", ["Setup", "Income", "Expense", "Budget", "Input"])
show_input_layout(P1_DEBT, "DEBT ai", ["Debt List", "Setup", "Input"])
show_input_layout(P1_NWT, "NWT ai", ["Assets", "Liabilities", "Balance Sheet", "Setup", "Input"])
