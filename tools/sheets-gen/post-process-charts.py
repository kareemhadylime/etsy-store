#!/usr/bin/env python3
"""
Post-process Budget Tracker .xlsx files to add native Excel charts.

ExcelJS (the JS generator) does not support `worksheet.addChart`. This script
opens each generated workbook with openpyxl, embeds charts, and saves back.

Charts added (per tier where the host sheet exists):
  Dashboard:
    - Bar chart  (Budget vs Actual)  — G10:J18 driven by H/I cols
    - Doughnut   (Category mix)       — from Expense Categories!B11:D23
  Financial Health Score (AI Edition only):
    - Bar chart  (Sub-scores 0-100)   — components D21:D25 with B21:B25 labels
  Annual Summary:
    - Line chart (12-month trajectory) — Income + Expense rows over C-N

Usage:
    python tools/sheets-gen/post-process-charts.py [path1.xlsx path2.xlsx ...]
With no args, defaults to the three published outputs.
"""
from __future__ import annotations

import io
import re
import shutil
import sys
import zipfile
from pathlib import Path
from typing import Iterable

# Force UTF-8 for stdout so the → / ✓ glyphs print on Windows cp1252 consoles.
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
else:  # pragma: no cover — pre-3.7 fallback
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import load_workbook


# Branding constants — must match the JS generator's workbook properties.
# Unified single brand for all client-facing surfaces.
BRAND = 'Lime Premium Studios'
LIME_LOGO_PATH = Path(__file__).resolve().parent / 'assets' / 'lime-logo-128.png'

from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image as XlImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU


DASHBOARD_SHEET = "🏠 Dashboard"
CATEGORIES_SHEET = "📂 Expense Categories"
HEALTH_SHEET = "🏆 Financial Health Score"
ANNUAL_SHEET = "📊 Annual Summary"

# Product registry — filename prefix → which chart routines to run.
# Every product gets _add_lime_logo() on every tab; charts are product-specific.
def _budget_tracker_charts(wb):
    """Charts for Budget Tracker workbooks."""
    added = 0
    if DASHBOARD_SHEET in wb.sheetnames:
        dash = wb[DASHBOARD_SHEET]
        _add_budget_vs_actual_chart(dash); added += 1
        _add_category_mix_chart(wb, dash); added += 1
    if HEALTH_SHEET in wb.sheetnames:
        _add_subscores_chart(wb[HEALTH_SHEET]); added += 1
    if ANNUAL_SHEET in wb.sheetnames:
        _add_annual_line_chart(wb[ANNUAL_SHEET]); added += 1
    return added


def _debt_payoff_charts(wb):
    """Charts for Debt Payoff Planner workbooks.

    Reads live from 📋 Debt List columns B (name), D (balance), E (APR).
    - Dashboard: doughnut of balances (top 10 debts by balance)
    - Strategy Comparison: bar chart of annual interest if available
    """
    added = 0
    if '📋 Debt List' in wb.sheetnames and DASHBOARD_SHEET in wb.sheetnames:
        dl = wb['📋 Debt List']
        dash = wb[DASHBOARD_SHEET]
        chart = DoughnutChart()
        chart.title = 'Debt mix by balance'
        chart.style = 26
        chart.dataLabels = DataLabelList(showPercent=True)
        # Debt List balances live at D11:D30, names at B11:B30 (input spine).
        data = Reference(dl, min_col=4, max_col=4, min_row=10, max_row=20)
        cats = Reference(dl, min_col=2, max_col=2, min_row=11, max_row=20)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 9; chart.width = 11
        dash.add_chart(chart, 'B22')
        added += 1
    return added


PRODUCT_CHART_BUILDERS = {
    'budget-tracker':       _budget_tracker_charts,
    'debt-payoff-planner':  _debt_payoff_charts,
}


def _add_budget_vs_actual_chart(ws) -> None:
    """Bar chart of Budget (col H) and Actual (col I) per category (col G), rows 10-23.
    Spans all 13 categories (was 8 — N20 fix; previously missed Healthcare / Education /
    Emergency Fund / Investing / Other, the entire 50/30/20 savings band)."""
    chart = BarChart()
    chart.type = "bar"
    chart.style = 11
    chart.title = "Budget vs Actual"
    chart.y_axis.title = "Category"
    chart.x_axis.title = "Dollars"
    chart.legend.position = "b"

    data = Reference(ws, min_col=8, max_col=9, min_row=10, max_row=23)  # H10:I23 (header + 13 cats)
    cats = Reference(ws, min_col=7, max_col=7, min_row=11, max_row=23)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 11
    chart.width = 17

    ws.add_chart(chart, "G26")


def _add_category_mix_chart(wb, dashboard_ws) -> None:
    """Doughnut chart of monthly budget by category, sourced from Expense Categories tab."""
    if CATEGORIES_SHEET not in wb.sheetnames:
        return
    cats_ws = wb[CATEGORIES_SHEET]
    chart = DoughnutChart()
    chart.title = "Where the money goes (by budget)"
    chart.style = 26
    chart.dataLabels = DataLabelList(showPercent=True)

    data = Reference(cats_ws, min_col=4, max_col=4, min_row=10, max_row=23)  # D10:D23 (header row 10 + 13 cats)
    cats = Reference(cats_ws, min_col=2, max_col=2, min_row=11, max_row=23)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 9
    chart.width = 11

    dashboard_ws.add_chart(chart, "B22")


def _add_subscores_chart(ws) -> None:
    """Horizontal bar chart of the 5 Health Score sub-scores 0-100."""
    chart = BarChart()
    chart.type = "bar"
    chart.style = 12
    chart.title = "Sub-scores (0–100)"
    chart.y_axis.title = "Component"
    chart.x_axis.title = "Score"
    chart.legend = None

    # Component names in col B rows 21-25; sub-score values in col D rows 21-25.
    data = Reference(ws, min_col=4, max_col=4, min_row=21, max_row=25)
    cats = Reference(ws, min_col=2, max_col=2, min_row=21, max_row=25)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.x_axis.scaling.min = 0
    chart.x_axis.scaling.max = 100
    chart.height = 7
    chart.width = 13

    ws.add_chart(chart, "G19")


def _add_annual_line_chart(ws) -> None:
    """Line chart of Income / Expense / Net over the 12 months in cols C-N."""
    chart = LineChart()
    chart.title = "12-month trajectory"
    chart.style = 13
    chart.y_axis.title = "Dollars"
    chart.x_axis.title = "Month"
    chart.legend.position = "b"

    # Rows 11-13 = Income / Expenses / Net. Cols C-N = Jan-Dec.
    data = Reference(ws, min_col=2, max_col=14, min_row=11, max_row=13)
    chart.add_data(data, titles_from_data=True, from_rows=True)
    # Month labels live in row 10 cols C-N.
    cats = Reference(ws, min_col=3, max_col=14, min_row=10, max_row=10)
    chart.set_categories(cats)
    chart.height = 9
    chart.width = 22

    ws.add_chart(chart, "B17")


def _add_lime_logo(ws) -> None:
    """Anchor the 32×32 Lime parent-brand logo in cell A1 of every visible tab.
    openpyxl's re-save strips the original ExcelJS-embedded images, so we re-add
    the same PNG from sheets-gen/assets/lime-logo-128.png on every tab here."""
    if not LIME_LOGO_PATH.exists():
        return
    img = XlImage(str(LIME_LOGO_PATH))
    img.width = 32
    img.height = 32
    # Anchor in cell A1 with a small inset so the icon sits inside the charcoal band.
    anchor = OneCellAnchor(
        _from=AnchorMarker(col=0, colOff=pixels_to_EMU(4), row=0, rowOff=pixels_to_EMU(4)),
        ext=XDRPositiveSize2D(cx=pixels_to_EMU(32), cy=pixels_to_EMU(32)),
    )
    img.anchor = anchor
    ws.add_image(img)


def _stamp_app_xml(path: Path) -> None:
    """Inject <Company> + branded <Application> into docProps/app.xml. openpyxl rewrites
    app.xml on save and drops the Company tag that ExcelJS originally wrote, so we patch
    the zip directly after the chart pass to make sure Excel's File→Info shows the brand."""
    tmp = path.with_suffix(path.suffix + '.brand.tmp')
    with zipfile.ZipFile(path, 'r') as zin, zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == 'docProps/app.xml':
                xml = data.decode('utf-8')
                # Rewrite Application + Company + Manager all to the unified brand.
                # Use regex (not "if not in xml") so we replace whatever openpyxl wrote.
                xml = re.sub(r'<Application>[^<]*</Application>',
                             f'<Application>{BRAND} — Budget Tracker</Application>', xml)
                if '<Company>' in xml:
                    xml = re.sub(r'<Company>[^<]*</Company>', f'<Company>{BRAND}</Company>', xml)
                else:
                    xml = xml.replace('</Properties>', f'<Company>{BRAND}</Company></Properties>')
                if '<Manager>' in xml:
                    xml = re.sub(r'<Manager>[^<]*</Manager>', f'<Manager>{BRAND}</Manager>', xml)
                else:
                    xml = xml.replace('</Properties>', f'<Manager>{BRAND}</Manager></Properties>')
                data = xml.encode('utf-8')
            zout.writestr(item, data)
    tmp.replace(path)


def _detect_product(filename: str) -> str | None:
    """Match the filename to a product key in PRODUCT_CHART_BUILDERS."""
    for key in PRODUCT_CHART_BUILDERS:
        if filename.startswith(key):
            return key
    return None


def process_workbook(path: Path) -> int:
    if not path.exists():
        print(f"  ! skip (missing): {path}")
        return 0

    wb = load_workbook(path)
    product_key = _detect_product(path.name)
    builder = PRODUCT_CHART_BUILDERS.get(product_key)

    added = builder(wb) if builder else 0
    if not builder:
        print(f"  ! no chart routine for {path.name} — Lime branding only")

    # Re-stamp the Lime logo on every sheet (openpyxl's re-save strips the original).
    logos_added = 0
    for ws in wb.worksheets:
        _add_lime_logo(ws)
        logos_added += 1

    wb.save(path)
    _stamp_app_xml(path)
    print(f"  ✓ {path.name} — {added} chart(s), Lime logo on {logos_added} tabs, branding stamped")
    return added


def main(argv: list[str]) -> int:
    here = Path(__file__).resolve().parent
    default_outputs = [
        here / "output" / "budget-tracker-essentials.xlsx",
        here / "output" / "budget-tracker-pro.xlsx",
        here / "output" / "budget-tracker-ai-edition-v2.xlsx",
        here / "output" / "debt-payoff-planner-essentials.xlsx",
        here / "output" / "debt-payoff-planner-pro.xlsx",
        here / "output" / "debt-payoff-planner-ai-edition.xlsx",
    ]
    targets: Iterable[Path] = [Path(a).resolve() for a in argv[1:]] if len(argv) > 1 else default_outputs

    print("→ Adding native Excel charts + Lime branding via openpyxl...")
    total = 0
    for t in targets:
        total += process_workbook(t)
    print(f"\n✓ Added {total} chart(s) total across all products")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
